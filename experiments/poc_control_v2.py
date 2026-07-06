#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""制御盤エンジン(正: 分岐回路コード方式)。1負荷=1回路コード。
主回路パターン→回路種別, kW→容量ステップ(切上), ●/○→デバイス変種。"""
import app, json, re, io
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB=app.DB; byCode=app.byCode
cache=json.load(open('pc_cache.json',encoding='utf-8'))

# --- 分岐回路コードを (変種, 種別, kW) で索引化 ---
def parse_bunki(name):
    # 例: 分岐回路(ELB・AX) L-S(AM付) 2.2  /  分岐回路 INV(支給品) 3.7
    m=re.match(r'分岐回路\s*(\([^)]*\))?\s*(.+?)\s*([\d.]+)\s*(KW|kW)?\s*$', name)
    if not m: return None
    dev=(m.group(1) or '').strip('()')          # ELB・AX / MCB・AX / ELB / ''(=MCCB基本)
    typ=m.group(2).strip()                       # L-S(AM付) / スターデルタ / INV(支給品) 等
    kw=float(m.group(3))
    return dev, typ, kw
INDEX={}
for d in DB:
    if '分岐回路' not in d['name']: continue
    p=parse_bunki(d['name'])
    if p: INDEX[p]=d['code']
def steps_for(dev, typ):
    return sorted(k for (dv,ty,k) in INDEX if dv==dev and ty==typ)

# 主回路パターン → 回路種別ラベル(DB表記に合わせる)
PAT2TYPE={'C':'L-S(AM付)','E':'L-S(AM付)','D':'スターデルタ','F':'スターデルタ','G':'スターデルタ','H':'スターデルタ',
          'I':'INV','J':'INV','K':'INV(スターデルタ)','L':'INV'}
def dev_variant(kind, remote):
    # ●=MCCB基本 / ○=ELB。遠方(操作A)=補助接点AX付き
    if kind=='○': return 'ELB・AX' if remote else 'ELB'
    return 'MCB・AX' if remote else ''
def find_code(typ, kw, dev):
    for cand_dev in ([dev,''] if dev else ['']):   # AX無しへフォールバック
        steps=steps_for(cand_dev, typ)
        if not steps: continue
        pick=next((s for s in steps if kw<=s+1e-9), steps[-1])   # 型ごとのステップで切上げ
        c=INDEX.get((cand_dev, typ, pick))
        if c: return c, pick, ('◎' if cand_dev==dev else '○')
    return '', None, '△'

# --- 適用表(キャッシュ) ---
rows=[]
for k in ('t_p1','t_p2L','t_p2R'): rows+=cache[k].get('rows',[])
loads=[r for r in rows if (r.get('load') or '')!='予備' and (r.get('main') or r.get('breaker'))]

bom=Counter(); detail={}
for r in loads:
    main=(r.get('main') or '').strip().upper()
    kw=r.get('kw'); kind=r.get('kind'); brk=(r.get('breaker') or '').strip()
    try: kwf=float(re.sub(r'[^\d.]','',str(kw)) or 0)
    except: kwf=0
    if main in PAT2TYPE and kwf>0:                     # 電動機回路 → 分岐回路コード
        typ=PAT2TYPE[main]; remote=False               # 遠方は操作記号A時。表からは簡易にFalse
        code,pick,conf=find_code(typ,kwf,dev_variant(kind,remote))
        key=f'分岐回路 {typ} {pick}kW' if pick else f'分岐回路 {typ} {kwf}kW(容量外)'
    elif brk:                                          # 電源/コンセント等 直接MCCB → 遮断器サイズで
        sel=app.select_one(brk.replace('MCCB','MCB').replace('AF','AF ').split('/')[0]+ '/'+brk.split('/')[-1] if '/' in brk else brk,'動力制御盤')
        # 簡易: 遮断器そのものを分岐MCBとして
        sel=app.select_one(('MCB '+brk),'動力制御盤')
        code=sel.get('code',''); conf=sel.get('conf','△'); key=f'{r.get("load","")}({brk})'
    else:
        code,conf='','△'; key=f'{r.get("load","")}(要確認)'
    bom[key]+=1; detail[key]=(code,conf)
# 盤本体
npanels=len(set(r.get('panel') for r in loads))
bom['制御盤本体(自立鋼板)']=npanels; detail['制御盤本体(自立鋼板)']=('','△')

print(f"実負荷{len(loads)}行 / 盤{npanels}面 / BOM {len(bom)}品目")
coded=sum(1 for k in bom if detail[k][0])
print(f"コード付与: {coded}/{len(bom)}品目\n")
print(f"{'回路/部品':<30}{'数':>4} {'コード':<8}{'判定':<4}品名")
out=[]
for key,q in bom.most_common():
    code,conf=detail[key]; nm=byCode.get(code,{}).get('name','') if code else ''
    out.append((key,q,code,nm,conf))
    print(f"  {key:<30}{q:>4} {code or '—':<8}{conf:<4}{nm[:22]}")

# Excel
FONT='Meiryo'; thin=Side(style='thin',color='BBBBBB'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
cf={'◎':'E8F0E8','◉':'DCEBF7','○':'FFF8E0','△':'FCE4E4'}
wb=Workbook(); ws=wb.active; ws.title='動力制御盤BOM'
ws.append(['回路/部品(展開)','数量','選定コード','正式品名(DB)','判定'])
for c in ws[1]: c.font=Font(name=FONT,bold=True,color='FFFFFF'); c.fill=PatternFill('solid',start_color='1E3A28'); c.border=bd; c.alignment=Alignment(horizontal='center',wrap_text=True)
for key,q,code,nm,conf in out:
    ws.append([key,q,code or '—',nm,conf]); row=ws[ws.max_row]
    for c in row: c.font=Font(name=FONT,size=9); c.border=bd
    row[4].fill=PatternFill('solid',start_color=cf.get(conf,'FFFFFF')); row[4].alignment=Alignment(horizontal='center')
for i,w in enumerate([34,6,12,28,6],1): ws.column_dimensions[chr(64+i)].width=w
ws.freeze_panes='A2'; wb.save('尼崎_動力制御盤_BOM.xlsx')
print("\n-> 尼崎_動力制御盤_BOM.xlsx")
