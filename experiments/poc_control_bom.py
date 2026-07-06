#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""制御盤PoC 完成版: pc_cache.json(抽出済) → pattern_master.json(編集可) →
展開 → コード付与(既存エンジン流用) → Excel。追加API呼び出しなし。"""
import app, json, os, io, re
from collections import Counter, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

cache=json.load(open('pc_cache.json',encoding='utf-8'))

# --- 1. パターンマスタを編集可能ファイルに(なければ生成) ---
MASTER='pattern_master.json'
if not os.path.exists(MASTER):
    master={
        "_説明":"御社標準に合わせて編集可。main=主回路A〜L, ctrl=操作1〜11。partsは{kind,qty}。電動機Mは負荷=対象外で除外。",
        "main":{p['id']:[pp for pp in p.get('parts',[]) if '電動機' not in pp.get('kind','') and pp.get('kind')!='M']
                for p in cache['main_patterns'].get('patterns',[])},
        "ctrl":{str(p['id']):p.get('parts',[]) for p in cache['ctrl_patterns'].get('patterns',[])}}
    json.dump(master, open(MASTER,'w',encoding='utf-8'), ensure_ascii=False, indent=1)
    print(f"[生成] {MASTER}")
master=json.load(open(MASTER,encoding='utf-8'))
main_m, ctrl_m = master['main'], master['ctrl']

# --- 2. 適用表(キャッシュ) → 実負荷 ---
rows=[]
for k in ('t_p1','t_p2L','t_p2R'): rows+=cache[k].get('rows',[])
loads=[r for r in rows if (r.get('load') or '')!='予備' and (r.get('main') or r.get('breaker'))]

# --- 3. 展開 → BOM(desc->qty) ---
bom=OrderedDict()
def add(desc,q):
    bom[desc]=bom.get(desc,0)+q
for r in loads:
    kind='ELB' if r.get('kind')=='○' else 'MCCB'
    brk=(r.get('breaker') or '').strip(); kw=r.get('kw','')
    for pp in main_m.get(r.get('main'),[]) or [{'kind':'MCCB','qty':1}]:
        k=pp.get('kind',''); q=pp.get('qty',1)
        if '電動機' in k or k=='M': continue
        if 'MCCB' in k or 'ELB' in k:
            add(f'{kind} {brk}' if brk else f'{kind} 分岐(要枠確認) {kw}kW', q)
        else: add(k, q)
    for pp in ctrl_m.get(str(r.get('ctrl') or ''),[]): add(pp.get('kind',''), pp.get('qty',1))
add('制御盤本体(自立鋼板)', len(set(r.get('panel') for r in loads)))

# --- 4. コード付与(既存 select_one 流用) ---
# 部品種類→照合クエリ(DB表記に寄せる)
QMAP={'電磁接触器MC':'電磁接触器','電磁接触器':'電磁接触器','サーマルリレーTHR':'サーマルリレー','サーマルTHR':'サーマルリレー',
      'THR':'サーマルリレー','変流器CT':'CT','CT':'CT','電流計A':'電流計','A':'電流計','補助リレー':'補助リレー',
      'タイマ':'タイマ','切換スイッチ':'切換スイッチ','端子台':'端子台','押ボタン':'押ボタン'}
def code_of(desc):
    if desc.startswith(('MCCB','ELB')):
        if '要枠' in desc: return '','△',f'分岐MCB 容量要確認({desc.split()[-1]})'
        sel=app.select_one(desc.replace('MCCB','MCB'), '動力制御盤')
        return sel.get('code',''), sel.get('conf','△'), sel.get('note','')
    if '制御盤本体' in desc: return '','△','盤製作・別途(標準色/自立鋼板)'
    q=QMAP.get(desc, desc)
    sel=app.select_one(q,'動力制御盤')
    return sel.get('code',''), sel.get('conf','△'), sel.get('note','')

rows_out=[]
for desc,q in bom.items():
    if not desc.strip(): continue
    code,conf,note=code_of(desc)
    rows_out.append((desc,q,code,app.byCode.get(code,{}).get('name','') if code else '',conf,note))

print(f"BOM {len(rows_out)}品目 / 実負荷{len(loads)}行 / 盤{len(set(r.get('panel') for r in loads))}面")
coded=sum(1 for x in rows_out if x[2])
print(f"コード付与: {coded}/{len(rows_out)}品目\n")
print(f"{'部品':<26}{'数':>4} {'コード':<8}{'判定':<4}品名")
for desc,q,code,nm,conf,note in rows_out:
    print(f"  {desc:<26}{q:>4} {code or '—':<8}{conf:<4}{nm[:22]}")

# --- 5. Excel ---
FONT='Meiryo'; thin=Side(style='thin',color='BBBBBB'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
cf={'◎':'E8F0E8','◉':'DCEBF7','○':'FFF8E0','△':'FCE4E4'}
wb=Workbook(); ws=wb.active; ws.title='動力制御盤BOM'
ws.append(['部品(展開)','数量','選定コード','正式品名(DB)','判定','根拠・確認事項'])
for c in ws[1]: c.font=Font(name=FONT,bold=True,color='FFFFFF'); c.fill=PatternFill('solid',start_color='1E3A28'); c.border=bd; c.alignment=Alignment(horizontal='center',wrap_text=True)
for desc,q,code,nm,conf,note in rows_out:
    ws.append([desc,q,code or '—',nm,conf,note])
    row=ws[ws.max_row]
    for c in row: c.font=Font(name=FONT,size=9); c.border=bd; c.alignment=Alignment(vertical='center',wrap_text=True)
    row[4].fill=PatternFill('solid',start_color=cf.get(conf,'FFFFFF')); row[4].alignment=Alignment(horizontal='center')
for i,w in enumerate([30,6,12,28,6,40],1): ws.column_dimensions[chr(64+i)].width=w
ws.freeze_panes='A2'
wb.save('尼崎_動力制御盤_BOM.xlsx')
print("\n-> 尼崎_動力制御盤_BOM.xlsx")
