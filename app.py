#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
積算コード選定システム v1.6  Flask版（図面 → コード選定）
  ブラウザから図面PDF/画像をアップ → Claude Visionで抽出 → v9.2 DB照合＋AI絞り込み → 結果表示・Excel
設計原則:
  AIはコードを生成しない。コードは v9.2 DB の実在コード または 確定/学習ルール からのみ。
  確信が持てない/読取不明瞭は △(要確認)。◎の誤りゼロを最優先。
起動:
  export ANTHROPIC_API_KEY=sk-ant-...
  pip install flask anthropic pdf2image openpyxl pillow
  python3 app.py        # → http://localhost:8000
"""
import os, re, io, json, base64, unicodedata, datetime, zipfile, tempfile, hmac, secrets, functools, contextvars
from flask import Flask, request, jsonify, send_file, Response
from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

HERE=os.path.dirname(os.path.abspath(__file__))
DB=json.load(open(os.path.join(HERE,'db.json'),encoding='utf-8'))
byCode={d['code']:d for d in DB}
# 部品別必要属性表(属性駆動の抽出チェック・選定確信度判定で共有)
try:
    ATTR_TABLE=json.load(open(os.path.join(HERE,'attr_table.json'),encoding='utf-8'))
except Exception:
    ATTR_TABLE={}
MODEL=os.environ.get('SELECT_MODEL','claude-opus-4-8')
app=Flask(__name__)
app.config['MAX_CONTENT_LENGTH']=40*1024*1024  # 40MB
app.secret_key=os.environ.get('APP_SECRET', secrets.token_hex(16))
APP_PASSWORD=os.environ.get('APP_PASSWORD','')  # 未設定なら認証オフ（ローカル試用向け）

from flask import session, redirect

def login_required(f):
    @functools.wraps(f)
    def w(*a,**k):
        if not APP_PASSWORD:        # パスワード未設定時は素通り（ローカル用）
            return f(*a,**k)
        if session.get('auth'):
            return f(*a,**k)
        # API呼び出しは401、画面はログインへ
        if request.path.startswith('/api/'):
            return jsonify(error='未認証'),401
        return redirect('/login')
    return w

LOGIN_HTML="""<!DOCTYPE html><html lang=ja><head><meta charset=UTF-8>
<meta name=viewport content="width=device-width,initial-scale=1"><title>ログイン</title>
<style>body{font-family:'Yu Gothic',Meiryo,sans-serif;background:#f7f5ef;display:flex;
align-items:center;justify-content:center;height:100vh;margin:0}
.box{background:#fff;border:1px solid #d6d1c4;border-radius:10px;padding:32px;width:320px}
h1{font-size:16px;color:#1e3a28;margin:0 0 18px}input{width:100%;padding:10px;border:1px solid #d6d1c4;
border-radius:6px;font-size:14px;box-sizing:border-box}button{width:100%;margin-top:12px;padding:11px;
background:#1e3a28;color:#fff;border:0;border-radius:6px;font-size:14px;font-weight:700;cursor:pointer}
.e{color:#c0504d;font-size:12px;margin-top:10px}</style></head>
<body><form class=box method=post action=/login>
<h1>積算コード選定システム</h1>
<input type=password name=pw placeholder=パスワード autofocus>
<button>ログイン</button>
{ERR}</form></body></html>"""

@app.route('/login',methods=['GET','POST'])
def login():
    if not APP_PASSWORD: return redirect('/')
    if request.method=='POST':
        if hmac.compare_digest(request.form.get('pw',''),APP_PASSWORD):
            session['auth']=True; return redirect('/')
        return Response(LOGIN_HTML.replace('{ERR}','<div class=e>パスワードが違います</div>'),mimetype='text/html')
    return Response(LOGIN_HTML.replace('{ERR}',''),mimetype='text/html')

@app.route('/logout')
def logout():
    session.clear(); return redirect('/login')


def client():
    if not os.environ.get('ANTHROPIC_API_KEY'):
        raise RuntimeError('ANTHROPIC_API_KEY 未設定')
    return Anthropic()

def norm(s):
    if not s: return ''
    s=unicodedata.normalize('NFKC',str(s)).lower()
    return re.sub(r'[ー\-\s\u3000()（）\[\]、,．.]','',s)

# ===== フェーズ1: 図面抽出（Vision）=====
EXTRACT_PROMPT="""あなたは配電盤・制御盤・分電盤の設計図面を読む専門家です。
この図面（単線結線図・盤リスト・分電盤結線図など）から、盤ごとに搭載される電気機器を抽出してください。

【最重要・抽出の第一原則: 機器の「塊」を正しく束ねる】
図面では、1つの機器が「記号(四角/丸/×印等)」＋「その周囲に縦書き・横書きで
分散した文字(機器名・電圧・電流・遮断容量・極数・容量)」で構成される。
機器を抽出する際は、記号の上下左右(特に左側・直上)に近接する文字を、
その機器1つの仕様としてすべて束ねること。数値が複数行に分かれていても1機器とする。
例(実際の図面の配置):
- 「VCB」「7.2kV」「600A」「12.5kA」が記号の脇に縦4行 → name="VCB 7.2kV 600A 12.5kA"
- 「DS×3」「7.2kV」「600A」が縦3行 → name="DS 7.2kV 3P 600A", qty=1
  ※「×3」は三相(R/S/T の3極)を表す記号であり、数量ではない。3極(3P)の断路器が1台。
    単線結線図の「×3」「x3」「×3個」の類は原則キュービクル内の3相分を示すもので、
    機器の台数は特記がない限り1とする(qty=1)。
- 「VT(PF付き)」「7.2kV」「1A」「6.6kV/110V」「50VA」 → 1つのVTとして束ねる
- 「LBS」「7.2kV200A」 → name="LBS 7.2kV 200A"
- 「T:1φ100kVA」「P:6600V」「S:210V/105V」 → 1つの変圧器として束ねる
記号の近くにある数値を「別の要素」として切り離さず、必ず最寄りの機器に結合する。
機器名だけ拾って定格(数値)を落とすことが最も多い失敗。記号の周囲の数値を見落とさない。

【電気図記号は機器名の代わりになる（重要）】
機器は「機器名の文字(VCB等)」だけでなく「電気図記号」でも表される。
記号そのものが機器を表すので、文字の機器名が無くても記号があれば機器として抽出する。
主な図記号と機器の対応:
- 〇を2つ重ねた記号(一次・二次巻線) = 変圧器(TR)。脇に「T:◯φ◯kVA P:◯V S:◯V」と表記
- ×印を四角で囲む = VCB(真空遮断器)
- 斜め線・接点記号 = DS(断路器)
- 丸の中に貫通線 = CT(変流器)、丸2つ = VT
- 縦の平行線 = コンデンサ(SC)、コイル記号 = リアクトル(SR)
- 箱に斜線・ヒューズ付き = LBS(高圧負荷開閉器)
- 丸の中に文字(V/A/W/I>等) = 計器・継電器
変圧器のように機器名の文字が無く記号と「T:」表記だけの場合も、必ず機器として抽出する。

【自己検証: 機器名も図記号も無い数値の塊は「結合漏れ」のサイン】
抽出を終える前に必ず自己チェックすること。図面上に「機器名(VCB/DS/CT等)も
図記号も伴わない、電圧・電流・遮断容量などの数値だけの塊」が浮いて残っていたら、
それは独立機器ではなく、近くの機器の定格を結合し損ねた証拠である。最寄りの機器に結合する。
※ただし近くに図記号(変圧器の〇二重等)があれば、それは正常な機器なので結合漏れではない。
逆に、機器名や記号はあるが定格が空の機器があれば、その記号の周囲を見直して
近くの数値(電圧・電流・遮断容量・極数・容量)を探し、結合すること。
特に高圧主回路のVCB・DS・LBS・PAS・VCT・VTは必ず定格を伴うはずなので、
定格が無いまま機器名だけになっていたら読み落としを疑い、周囲を再確認する。

ルール:
- 盤名ごとにグループ化（例: 高圧受電盤, 低圧動力盤No1, P-AC11, L-1W, T-1）
- 各機器の「品名・仕様(型番/容量/極数/AF等を図面表記のまま)」と「数量」
- 主幹/分岐/計器/操作回路/支給品(TR/SC/SR)/SPD/函体/端子盤を漏れなく
- 読み取れない項目は推測せず "unclear":true

【ケーブル・電線類は計上対象外（機器ではない）】
FP・FPT・CVT・CV・CVV・IV・VVF・KIV等の電線・ケーブル類は機器ではないので、
盤搭載機器として抽出しない("cable":true を付けるか、そもそも出力しない)。
例: 「6kV FPT60」「CVT38²」「CV60²」等はケーブルなので計上対象外。
ただし負荷名称行に付随するケーブルサイズ表記(L-1A CVT60²等)は従来どおり負荷明細扱い。

【LP等の分電盤+制御盤一体盤は母線(主幹)ごとに分けて出力】
「LP」「LP＋数字」等の盤は、1つの盤の中に「制御用母線」と「分電用母線」の2つの
主幹(入力系統)を持つ一体盤のことがある。図面上、母線(主幹)が2つ描かれていたら、
母線ごとに別の盤として出力し、盤名に区画を付記すること:
- 制御用母線(3P動力・制御機器側。主幹M)MCB/M)ELB＋3P分岐)の区画 → panel="(元の盤名) (制御)"
- 分電用母線(照明・コンセント等の小容量2P分岐が主。主幹＋2P分岐)の区画 → panel="(元の盤名) (分電)"
各母線について、その主幹(M)MCB/M)ELB)・配下の分岐・M)LUG(盤頭端子)を、その区画に紐付ける。
主幹(入力系統)が2つ見えたら必ず2区画に分ける。主幹が1つしか無ければ分けない。
※LP以外でも、明らかに制御母線と分電母線が別々に立っている盤は同様に(制御)/(分電)で分ける。

【高圧コンデンサ盤の構成（密集機器の分離・重要）】
高圧コンデンサ盤は1系統あたり次の4機器で構成される。縦に密集して描かれるが、
必ず4つの別々の機器として分離して抽出すること(まとめて1～2機器にしない)。
- PF(限流ヒューズ): 例 40A
- VMC(真空電磁接触器): 例 200A
- SR(直列リアクトル): 例 4.97kvar ※SCの約6%の小容量
- SC(高圧コンデンサ本体): 例 7.02kV 3φ 79.8kvar
各機器の定格を取り違えないこと。特に「SCの容量(79.8kvar)」と「VMCの定格電流(200A)」と
「SRの容量(4.97kvar)」は別物。これらを1つの機器名に混ぜてはいけない。
コンデンサ盤が複数系統(SC-1, SC-2)あれば、系統ごとに4機器を繰り返し抽出する。
VMCやSRが抜け落ちやすいので、PFとSCだけでなくVMC・SRも必ず拾うこと。

【PF(パワーヒューズ)とF(計器ヒューズ)は別機器・分離する】
「PF」はパワーヒューズ(限流ヒューズ、高圧主回路の機器、計上対象)。
「F」は計器ヒューズ(計器保護用の小ヒューズ)。両者は別機器なので、図面で近くにあっても
1つにまとめず、別々の機器として分離して抽出すること。
  例: 図面に「PF」と「F 3A×2」が近接 → name="PF"(または定格付き) と name="F 3A", qty=2 の2機器に分離
「F×N」「F N個」はFが計器ヒューズN個の意味。PFと混同して「PF Fx2」のように
1機器に合体させてはいけない。

【部品別・必要属性表（全カテゴリ網羅）】
部品名を特定したら、対応する属性を図面(記号の脇・機器表・凡例・仕様表)から
必ず探して name に含めること。属性が読めれば正しいコード選定に直結する。
《高圧機器》
- VCB: 電圧/定格電流(400/600/1200A)/遮断容量(8/12.5/20/25kA)/操作(手動/電動/引出)
- VCS・VMC: 電圧/定格電流(200/400A)/形式(電磁/引出/PF付)
- DS(断路器): 電圧/極数(1P/3P)/定格電流(200/400/600A)
- LBS: 電圧/定格電流(200A)/PF有無/G感度(G75A以下等)
- PAS: 電圧/定格電流(200/300A)/種別(SOG/UGS/G-RY付)
- PF(限流ヒューズ): 定格電流(A)
《変成器・計器》
- CT: 変流比(150/5A等)/負担(VA)/形式(丸型/モールド)
- VT: 電圧(一次kV)/負担(VA)/形式(モールド)
- WHM(電力量計): 相数(1φ/3φ)/変流比/電流/検定有無
- ZCT: 定格電流(200/400/600A)
- 計器(VM/AM/VS/AS): 形式(普通角/広角/付加)
《変圧器・コンデンサ》
- TR(変圧器): 相数(1φ/3φ)/容量(kVA) ※二次電圧は選定に不要
- SC(高圧コンデンサ): 電圧/容量(kVA)
- SC(低圧コンデンサ): 電圧(V)/容量(kvar or kW)
- SR(リアクトル): 相数/容量(kvar) ※定格電流(A)でなく容量(kvar)
- LCユニット: L%/電圧(V)/容量(kvar)
《遮断器・開閉器（低圧）》
- B)MCB/ELB/MCCB(分岐): 極数(2P/3P/4P)/AF/AT(定格電流)
- M)MCB(負荷主幹・遮断器): 極数/定格電流/欠相保護有無 ※頭に「MCCB/MCB」表記＋遮断器記号
- M)LUG(主幹端子・断路のみ): 極数/定格電流 ※頭に断路スイッチ記号はあるが「MCCB/MCB」表記が無い＝端子受け
- ACB: 極数/AF/AT
- MCB(汎用)/MCCB: 極数/定格電流
- MCTT: 極数/定格電流/交流直流
- 高速電源切替器: 極数/定格電流/操作
- SSC: 極数/定格電流
《電動機・起動》
- MGS(電磁開閉器): 容量(kW)/電圧(V)
- SM/起動器/スターリアクトル: 容量(kW)/電圧/極数
- INV(インバータ): 容量(kW)/電圧(V)
- ACリアクトル/スターリアクトル: 容量(kW)/電圧(V)
《端子・函体・その他》
- 端子盤/TB/制御端子: 極数(P)/定格電流
- 保安器函/接地端子盤: 極数(P)
- 函体: 種別(屋内/屋外/標準)・寸法増し有無
- SPD: クラス(I/II)/極数/遮断容量(kA)
  ※SPDや計器が入力(系統)の母線から分岐する箇所の「頭」を必ず見る:
    ・頭に断路スイッチ記号(×)があり「MCCB/MCB」表記が無い → その頭を『M)LUG [極数]P [系統主幹容量]A』として1件計上(端子受け)。
    ・頭にスイッチ自体が無い(直結) → M)LUGは計上しない。
    ・頭に「MCCB/MCB」表記＋遮断器記号がある(負荷側主幹) → 従来どおりM)MCBで拾う。
  ※SPD用分離器(セパレータ)はMCCBではない。分離器は分離器として別項目で出す(遮断器に誤読しない)。
- 継電器(OCR/DGR/RPR/OVGR等): 形式(静止型/方向性等)
- TV機器: 種別(分配器/増幅器/混合器)・ポート数
- コンセント: 極数/定格電流/E付・プラグ付
《区分の共通ルール》
- TR/SC/SR等: 図面に「支給」明記→支給品/「スペース」明記→スペース/明記なし→購入
  ※高圧TR/SC/SRは購入品コードが無いので明記なしでも支給品系
読めない属性は推測せず省く。名称だけで終わらせず、その部品の属性を狙って拾うこと。

【同一機器の重複抽出を避ける】
1つの機器に付属する制御・保護要素(PASのSOG制御箱、VCBの操作機構等)を、
別個の機器として二重に出さないこと。SOGはPASの一部、と判断する。
品名中の「×N」「xN」「XN」は数量N個を意味する。"qty" にその数を入れること。
  例: 「SC 7.2kV 50kvar(SC×3)」→ name="SC 7.2kV 50kvar", qty="3"
  例: 「PF×3 40A」→ name="PF 40A", qty="3" / 「CT×2」→ qty="2"
ただし「3φ」の3は相数であり数量ではない。
また「2.2kW×2」のように単位(kW/kVA等)直後の×Nは負荷構成の説明なので数量にしない。

【配電盤リストの「負荷名称行」の扱い（最重要）】
配電盤リスト(電灯盤/動力盤の表)では、1つの分岐回路の下に、その回路が給電する複数の負荷が
ぶら下がって列挙されることがあります。表の構造を次の2種類に区別してください:
- 【機器行】「接続用遮断器(AF/AT)」列に値がある行（例: 1L1 MCB3P 225/150、2M1 MCB3P 225/225）
  → これが計上対象の分岐機器。通常どおり "name" に機器仕様を入れる。
- 【負荷明細行】AF/AT列が空欄で、負荷名称(L-2A, M-1A, EV1, 直圧1, PF-1等)と容量(kVA/kW)・
  ケーブルサイズ(CVT38²等)だけの行
  → これは直前の機器行(親分岐MCB)が給電する負荷の内訳。独立した機器ではない。
  → "load_detail":true を付け、"parent" に親の分岐回路記号(例:"1L1")を入れる。
  → "name" には負荷名称と容量をそのまま入れてよいが、機器仕様は付けない。
重要: 負荷名称(L-1A,L-2A,M-1A等)は「機器名」ではなく「その回路が何に給電するかの名称」です。
これを独立した機器として抽出しないでください。親の分岐MCBが計上単位です。
ただし負荷名称行に明示的に機器(「MCB分岐」等)が併記されている場合のみ機器行として扱います。

【電圧帯の判定（重要）】
単線結線図では変圧器(TR)を境に電圧が変わります。各機器について、接続されている線の電圧を判定し "volt" に入れてください:
- 高圧(TR一次側、6.6kV/7.2kV等) → "volt":"HV"
- 低圧 400V級(415V/440V、三相3線400V級) → "volt":"400V"
- 低圧 200V級(210V/220V) → "volt":"200V"
- 低圧 100V級(105V/100V) → "volt":"100V"
- 図から判断できない/確信が持てない → "volt":"" (空欄。推測しない)
低圧でも200V級と400V級で品番が異なるため、TR二次側の電圧表記(210V/415V等)を必ず確認してください。
特にCT・VT・計器・電磁開閉器(MGS)・モータ回路(L-S/INV)は電圧でコードが変わります。線の接続と電圧表記をたどって判定してください。

【動力制御盤の主回路パターン集＋適用表（たすき掛け・最重要）】
動力制御盤は「左に主回路パターン集(記号A〜L＝各回路の型)」「右に適用表(各負荷がどの記号を使うか)」で構成される。
1) パターン集(凡例)を読む: 各記号(A,B,C…)が何の回路型か(直入L-S/スターデルタ/インバータ/電源/コンセント等)を
   読み取り、盤ごとに "legend":{"A":"型の説明","C":"型の説明",...} を出す。**割付は図面ごとに違う**ので必ず図面の
   凡例から読む(標準の思い込みで決めない)。パターン集が読めない場合のみ "legend":{} とする。
2) 適用表の各負荷(行)を items に出す: "symbol"=主回路記号(A〜L)、"kw"=容量(数値)、
   "breaker"=分岐遮断器の別(●=MCCB は "●"、○=ELB は "○"、表記どおり)。負荷名は "name"。
動力盤でない(記号が無い)場合は "symbol":"" とする。「予備」行は name:"予備" とする。

【分電盤(電灯/照明分電盤)の分岐回路表（重要・取りこぼし厳禁）】
分電盤は「回路番号／分岐開閉器(列: 1P・2P・ELB・1G・1H 等に○印)／R-RY／負荷名称・用途／容量VA／備考」の表で構成される。
各回路(行)を items に1つずつ出す。**分岐開閉器の○印がどの列に付くかで遮断器種別が決まる**ので、必ず読み取る:
1) 図の凡例(例「1P:MCB1P 50AF/20AT ／ 2P:MCB2P 50AF/20AT ／ ELB:ELB2P 50AF/20AT」)を確認する。
2) 各回路の○印が付いた列を読み、遮断器種別を **name の末尾に必ず付ける**(例 "電灯 事務室 280VA MCB2P"、
   "コンセント 廊下 400VA ELB2P"、"予備 500VA MCB2P")。1P列→"MCB1P"、2P列→"MCB2P"、ELB列→"ELB2P"。
3) "qty": 同一回路の口数。負荷名/備考に「Rx5」「×5」「N回路」等の乗数があればその数、無ければ"1"。
4) 「予備」行も分岐開閉器の○があれば種別を付ける(例 "予備 500VA MCB2P")。○が無ければ name:"予備"。
分岐開閉器の○列がどうしても読めない行のみ "unclear":true とし種別を付けない(推測しない)。

【受変電の保護セット（重要）】
高圧受変電の単線図では、保護リレーと、その検出器・試験端子が、線で繋がって/近接してまとまって描かれ、1つの保護セットを構成します。次のセット関係を読み取り、付属側に "group" でセット名(親リレー)を入れてください:
- 地絡方向保護: DGR(親) ← ZCT・ZPD(検出器)。例: ZCTに "group":"DGR"
- 過電流保護: OCR(親) ← CT・CTT(検出器・試験端子)。例: CTに "group":"OCR"
- 逆電力/地絡過電圧: RPR・OVGR(親) ← VT・VTT・ZPD
- 計量: WHM/Wh(親) ← VCT・VT・CT
リレー本体(DGR/OCR/RPR/OVGR/WHM)は "group" を自分自身の名前にします。線の繋がり・近接配置でセットを判断し、どのリレーに属すか分かるものは必ず "group" を付けてください。判断できない場合は "group":"" とします。

同じ図面に対しては常に同じ結果を返すよう、決定的・一貫して判断すること。
推測や曖昧な解釈で揺らがず、図面に書かれた事実に厳密に基づいて抽出する。

【配電盤(受変電)は盤単位の "set_attrs" も出す（セットコード選定用）】
受電盤/饋電盤/低圧電灯盤/低圧動力盤/スコットトランス盤/段積 の盤には set_attrs を付ける。
制御盤(動力)・分電盤・端子盤は set_attrs 省略(または settype:"")。
- settype: "低圧"(TR二次の電灯/動力/スコット) / "高圧"(受電・饋電) / "段積" / ""
- meter(計器種別): 推測しない。器具表等で明確な時のみ "広角"/"普通角"/"マルチ"、単線図だけで確信が無ければ ""(空)。
- phase: 1φ→"1φ3W" / 3φ→"3φ3W" / スコットTR→"スコット"
- cap: 変圧器のKVA(数値+KVA)。読めなければ ""。
- vcb: "8KA"/"12.5KA"(高圧のみ)。 op: "手動"/"電動"/"電動引出"(VCB上下に黒四角＋回転矢印記号なら電動引出)。
- unclear_specs: 読み切れなかった仕様名の配列(例 ["meter","cap"])
出力は次のJSONのみ（説明文・マークダウン禁止）:
{"panels":[{"panel":"盤名","set_attrs":{"settype":"","meter":"","phase":"","cap":"","vcb":"","op":"","unclear_specs":[]},"legend":{},"items":[{"name":"品名仕様","qty":"数量","volt":"HV|400V|200V|100V|","symbol":"A〜L|","kw":"容量数値|","breaker":"●|○|","group":"親リレー名|","load_detail":false,"parent":"","unclear":false}]}]}"""

# ===== DXF: 文字を座標つきで抽出し、Claudeで盤・機器に構造化 =====
def extract_dxf(cli, data_bytes):
    import ezdxf
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as tf:
        tf.write(data_bytes); path=tf.name
    try:
        doc=ezdxf.readfile(path)
    except Exception as e:
        raise RuntimeError(f"DXF読込失敗: {e}")
    finally:
        try: os.unlink(path)
        except: pass
    msp=doc.modelspace()
    texts=[]
    for e in msp:
        try:
            if e.dxftype()=="TEXT":
                texts.append((round(e.dxf.insert.x,1),round(e.dxf.insert.y,1),e.dxf.text))
            elif e.dxftype()=="MTEXT":
                texts.append((round(e.dxf.insert.x,1),round(e.dxf.insert.y,1),e.text))
        except Exception:
            continue
    if not texts:
        return {"panels":[]}
    texts.sort(key=lambda v:(-v[1],v[0]))
    lines="\n".join(f"({x},{y}) {t}" for x,y,t in texts)
    prompt=(EXTRACT_PROMPT +
        "\n\n以下はDXF図面から抽出した文字（座標(x,y)付き）です。"
        "座標の近さから盤と機器の対応を判断し、盤ごとに機器を構造化してください。\n\n"+lines[:60000])
    with cli.messages.stream(model=MODEL,max_tokens=64000,
        messages=[{"role":"user","content":prompt}]) as stream:
        msg=stream.get_final_message()
    txt="".join(b.text for b in msg.content if b.type=="text").strip()
    txt=re.sub(r"^```(json)?|```$","",txt,flags=re.M).strip()
    try: return json.loads(txt)
    except Exception:
        m=re.search(r"\{.*\}",txt,re.S)
        if m:
            try: return json.loads(m.group(0))
            except Exception: pass
        # 応答が出力上限で途中で切れた場合は分かりやすく通知
        if getattr(msg,'stop_reason',None)=='max_tokens':
            raise ValueError('図面の機器数が多く抽出結果が出力上限を超えました。盤を分割して投入してください。')
        return {"panels":[]}

# ===== 入力1件を抽出（拡張子で振り分け）=====
def extract_one(cli, fname, data_bytes):
    low=fname.lower()
    if low.endswith(".pdf"):  return extract_pdf_hires(cli, data_bytes)
    if low.endswith(".png"):  return extract(cli, data_bytes, "image/png")
    if low.endswith((".jpg",".jpeg")): return extract(cli, data_bytes, "image/jpeg")
    if low.endswith(".dxf"):  return extract_dxf(cli, data_bytes)
    return None  # 未対応はスキップ

def extract_pdf_hires(cli, data_bytes):
    """PDFを高解像度画像に変換し、ページごとにVisionへ送る。
    APIにPDFをそのまま渡すより解像度を制御でき、密集機器・細かい容量の
    読み取り精度が上がる。複数ページは各々抽出してpanelsを統合する。"""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        # PyMuPDF未導入なら従来方式(PDFそのまま)にフォールバック
        return extract(cli, data_bytes, "application/pdf")
    all_panels=[]
    doc=fitz.open(stream=data_bytes, filetype="pdf")
    for page in doc:
        # 3倍解像度でレンダリング(細部の文字が潰れないように)
        pix=page.get_pixmap(matrix=fitz.Matrix(3,3))
        png=pix.tobytes("png")
        res=extract(cli, png, "image/png")
        for p in res.get("panels",[]):
            all_panels.append(p)
    return {"panels":all_panels}

# ===== ZIP含む入力をまとめて抽出（複数図面を一括）=====
def extract_input(cli, fname, data_bytes):
    low=fname.lower()
    all_panels=[]
    if low.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(data_bytes)) as z:
            for info in z.infolist():
                if info.is_dir(): continue
                inner=info.filename
                if inner.startswith("__MACOSX") or "/." in inner: continue
                if not inner.lower().endswith((".pdf",".png",".jpg",".jpeg",".dxf")): continue
                try:
                    res=extract_one(cli, inner, z.read(info))
                except Exception as e:
                    all_panels.append({"panel":f"[抽出失敗] {os.path.basename(inner)}",
                        "items":[{"name":str(e),"qty":"","unclear":True}]}); continue
                if res:
                    base=os.path.basename(inner)
                    for p in res.get("panels",[]):
                        p["panel"]=f"{base} / {p.get('panel','')}"
                        all_panels.append(p)
        return {"panels":all_panels}
    res=extract_one(cli, fname, data_bytes)
    if res is None:
        raise RuntimeError("未対応の形式です（PDF/PNG/JPG/DXF/ZIPのみ）")
    return res

def extract(cli, data_bytes, media):
    src={"type":"base64","media_type":media,"data":base64.standard_b64encode(data_bytes).decode()}
    block={"type":"document","source":src} if media=="application/pdf" else {"type":"image","source":src}
    with cli.messages.stream(model=MODEL,max_tokens=64000,
        messages=[{"role":"user","content":[block,{"type":"text","text":EXTRACT_PROMPT}]}]) as stream:
        msg=stream.get_final_message()
    txt="".join(b.text for b in msg.content if b.type=="text").strip()
    txt=re.sub(r'^```(json)?|```$','',txt,flags=re.M).strip()
    try: return json.loads(txt)
    except Exception:
        m=re.search(r'\{.*\}',txt,re.S)
        if m:
            try: return json.loads(m.group(0))
            except Exception: pass
        if getattr(msg,'stop_reason',None)=='max_tokens':
            raise RuntimeError('図面の機器数が多く抽出結果が出力上限を超えました。盤を分割して投入してください。')
        raise RuntimeError('抽出JSON解析失敗')

# ===== フェーズ2: 属性方式 選定エンジン（候補生成→ルール絞り込み）=====
def R(code,conf,note): return dict(code=code,name=byCode.get(code,{}).get('name','') if code else '',conf=conf,note=note)

def _norm2(s):
    import re as _re
    return _re.sub(r'[()（）.]','',norm(s))

def split_qty_suffix(raw):
    """品名中の「x2 ×2 X2」等を数量として分離。
    ただし「2.2kW×2」のように単位(kW/kVA/kvar/V/A)直後の×Nは
    負荷構成の説明なので数量にしない。機器個数を表す×Nのみ分離する。"""
    import re as _re
    s=str(raw)
    # 例外: 高圧DS(断路器)の「A×3」は三相(3極)を表す表記で台数ではない。
    # 「DS ... 600A×3」「DS×3 ... 600A」等は数量分離せず、3P(3極)として明示する
    # (×3が明示されている＝三相と断定できるため、後段で3P品に確定させる)。
    if _re.search(r'(?<![a-z])ds(?![a-z])', s, _re.I) and _re.search(r'[xX×]\s*3\b', s):
        s2=_re.sub(r'[xX×]\s*3\b','', s).strip(' ,、')
        if not _re.search(r'[123]p', s2, _re.I):   # 極数未記載なら3Pを補う
            s2=s2+' 3P'
        return s2, None
    for m in _re.finditer(r'[xX×]\s*(\d{1,3})\b', s):
        pre=s[:m.start()].rstrip()
        # 容量単位(kw/kva/kvar)直後は負荷構成→数量にしない。
        # 定格A(例 600A×3)直後は台数なので数量にする。
        if _re.search(r'(kw|kva|kvar|kva?r)$', pre, _re.I):
            continue
        return (s[:m.start()]+s[m.end():]).strip(' ,、'), m.group(1)
    return s, None

_OPTION_WORDS=['spd用','ct付','ax付','2e付','1e付','am付','as付','pf付','al付','zct付','sog付','g-ry付',
              'td付','広角','赤針付','spd付','th付','ドアsw付','検付','通信','方向性',
              '引出','電動','手動','コージェネ','標準','高性能','漏電アラーム']
def split_main_opt(name):
    """メイン機器名と付属(オプション)を分離"""
    import re as _re
    s=str(name); options=[]
    for p in _re.findall(r'[（(]([^（）()]*)[）)]', s): options.append(p.strip())
    main_str=_re.sub(r'[（(][^（）()]*[）)]','', s).strip()
    for tok in _re.split(r'[\s\u3000,、]', main_str):
        if tok.endswith('付') and len(tok)>1:
            options.append(tok); main_str=main_str.replace(tok,'').strip()
    nf=norm(s)
    for ow in _OPTION_WORDS:
        if norm(ow) in nf and ow not in [norm(o) for o in options]: options.append(ow)
    return main_str.strip(), options

def _voltband(volt, name):
    n=norm(name)
    if volt in ('HV','400V','200V','100V'): return volt
    if any(k in n for k in ['6kv','6.6kv','7.2kv','12.5']): return 'HV'
    if '415' in n or '440' in n: return '400V'
    if '210' in n or '220' in n: return '200V'
    if '105' in n or '100v' in n: return '100V'
    return ''
def _dbvolt(v):
    if not v: return ''
    if 'kv' in v.lower(): return 'HV'
    for x in ('400','200','100'):
        if x in v: return x+'V'
    return ''
def _get_kw(n):
    import re as _re
    m=_re.search(r'(\d+\.\d+|\d+)\s*kw', n); return m.group(1) if m else None
def _get_kva(n):
    import re as _re
    m=_re.search(r'(\d+\.\d+|\d+)\s*kva[r]?', n); return m.group(1) if m else None

# メイン機器キー (kw, label, prefix一致が必要か)
_MAIN_KEYS=[
 ('t/u','T/U',True),('リモコンsw','リモコンSW',True),
 ('mdf','MDF',True),('端子盤','端子盤',True),
 ('分配器','分配器',True),('増幅器','増幅器',True),('混合器','混合器',True),
 ('uhf','UHFアンテナ',False),('bs-110','BS-110CSアンテナ',False),('アンテナ','アンテナ',False),
 ('インターホン','インターホン',True),('ネットワークカメラ','ネットワークカメラ',False),('カメラ','カメラ',False),
 ('カードリーダー','カードリーダー',True),('tvbox','TV BOX',False),('tv-','TV',False),
 ('自動力率調整','自動力率調整器',True),('自動力率','自動力率調整器',True),('apfc','自動力率調整器',False),
 ('マルチ指示計','マルチ指示計',True),('mda','マルチ指示計',True),('多機能電力計','マルチ指示計',True),('マルチt/d','マルチT/D',True),
 ('電圧計','電圧計',True),('電流計','電流計',True),('電力計','電力計',True),('力率計','力率計',True),
 ('vm','VM',True),('am','AM',True),('vs','VS',True),('as','AS',True),
 ('whm','WHM',False),('mgs','MGS',True),('pas','PAS',True),('vcb','VCB',True),('vct','VCT',True),
 ('vmc','VCS',True),('vcs','VCS',True),('sog','SOG',True),
 ('lbs','LBS',True),('vgb','VCB',True),('ds','DS',True),('ch','CH',True),('l-s','L-S',False),('インバータ','INV',False),('inv','INV',False),
 ('ocr','OCR',True),('dgr','DGR',True),('rpr','RPR',True),('ovgr','OVGR',True),
 ('lg-ry','LG-RY',True),('lgr','LGR',True),('zpd','ZPD',True),('zctt','ZCTT',True),('zct','ZCT',True),
 ('vtt','VTT',True),('ctt','CTT',True),
 ('m)lug','M)LUG',True),('m)mcb','M)MCB',True),('m)elb','M)ELB',True),('b)mcb','B)MCB',True),('b)elb','B)ELB',True),
 ('mcb','MCB',True),('elb','ELB',True),('vt','VT',True),('ct','CT',True),('pf','PF',True),('la','LA',True),('sc','SC',True),('srx','SR',True),('sr','SR',True),('tr','TR',True),
 ('fl-10w','FL-10W',True),('pbs','PBS',True),('spd','SPD',True),('mctt','MCTT',True),
 ('換気扇','換気扇',True),('コンセント','コンセント',True),('伝送','伝送',True),
 ('リモコンリレー','R.RY',False),('リモコントランス','R.TR',False),('端子台','TB',True)]

_PARTS = [
    {'name':'T/U','aliases':['t/u'],'prefix':True},
    {'name':'リモコンSW','aliases':['リモコンsw'],'prefix':True},
    {'name':'MDF','aliases':['mdf'],'prefix':True},
    {'name':'端子盤','aliases':['端子盤'],'prefix':True},
    {'name':'分配器','aliases':['分配器'],'prefix':True},
    {'name':'増幅器','aliases':['増幅器'],'prefix':True},
    {'name':'混合器','aliases':['混合器'],'prefix':True},
    {'name':'UHFアンテナ','aliases':['uhf'],'prefix':False},
    {'name':'BS-110CSアンテナ','aliases':['bs-110'],'prefix':False},
    {'name':'アンテナ','aliases':['アンテナ'],'prefix':False},
    {'name':'インターホン','aliases':['インターホン'],'prefix':True},
    {'name':'ネットワークカメラ','aliases':['ネットワークカメラ'],'prefix':False},
    {'name':'カメラ','aliases':['カメラ'],'prefix':False},
    {'name':'カードリーダー','aliases':['カードリーダー'],'prefix':True},
    {'name':'TV BOX','aliases':['tvbox'],'prefix':False},
    {'name':'TV','aliases':['tv-'],'prefix':False},
    {'name':'自動力率調整器','aliases':['自動力率調整', '自動力率', 'apfc'],'prefix':True},
    {'name':'マルチ指示計','aliases':['マルチ指示計','mda','多機能電力計','デジタルマルチメーター'],'prefix':True},
    {'name':'マルチT/D','aliases':['マルチt/d'],'prefix':True},
    {'name':'電圧計','aliases':['電圧計'],'prefix':True},
    {'name':'電流計','aliases':['電流計'],'prefix':True},
    {'name':'電力計','aliases':['電力計'],'prefix':True},
    {'name':'力率計','aliases':['力率計'],'prefix':True},
    {'name':'VM','aliases':['vm'],'prefix':True},
    {'name':'AM','aliases':['am'],'prefix':True},
    {'name':'VS','aliases':['vs'],'prefix':True},
    {'name':'AS','aliases':['as'],'prefix':True},
    {'name':'WHM','aliases':['whm'],'prefix':False},
    {'name':'MGS','aliases':['mgs'],'prefix':True},
    {'name':'PAS','aliases':['pas'],'prefix':True},
    {'name':'VCB','aliases':['vcb', 'vgb'],'prefix':True},
    {'name':'VCT','aliases':['vct'],'prefix':True},
    {'name':'VCS','aliases':['vcs','vmc'],'prefix':True},
    {'name':'SOG','aliases':['sog'],'prefix':True},
    {'name':'LBS','aliases':['lbs'],'prefix':True},
    {'name':'DS','aliases':['ds'],'prefix':True},
    {'name':'CH','aliases':['ch'],'prefix':True},
    {'name':'L-S','aliases':['l-s'],'prefix':False},
    {'name':'INV','aliases':['インバータ', 'inv'],'prefix':False},
    {'name':'OCR','aliases':['ocr'],'prefix':True},
    {'name':'DGR','aliases':['dgr'],'prefix':True},
    {'name':'RPR','aliases':['rpr'],'prefix':True},
    {'name':'OVGR','aliases':['ovgr'],'prefix':True},
    {'name':'LG-RY','aliases':['lg-ry'],'prefix':True},
    {'name':'LGR','aliases':['lgr'],'prefix':True},
    {'name':'ZPD','aliases':['zpd'],'prefix':True},
    {'name':'ZCTT','aliases':['zctt'],'prefix':True},
    {'name':'ZCT','aliases':['zct'],'prefix':True},
    {'name':'VTT','aliases':['vtt'],'prefix':True},
    {'name':'CTT','aliases':['ctt'],'prefix':True},
    {'name':'M)LUG','aliases':['m)lug'],'prefix':True},
    {'name':'M)MCB','aliases':['m)mcb'],'prefix':True},
    {'name':'M)ELB','aliases':['m)elb'],'prefix':True},
    {'name':'B)MCB','aliases':['b)mcb'],'prefix':True},
    {'name':'B)ELB','aliases':['b)elb'],'prefix':True},
    {'name':'MCB','aliases':['mcb'],'prefix':True},
    {'name':'ELB','aliases':['elb'],'prefix':True},
    {'name':'VT','aliases':['vt'],'prefix':True},
    {'name':'CT','aliases':['ct'],'prefix':True},
    {'name':'PF','aliases':['pf'],'prefix':True},
    {'name':'LA','aliases':['la'],'prefix':True},
    {'name':'SC','aliases':['sc'],'prefix':True},
    {'name':'SR','aliases':['sr','srx'],'prefix':True},
    {'name':'TR','aliases':['tr'],'prefix':True},
    {'name':'FL-10W','aliases':['fl-10w'],'prefix':True},
    {'name':'PBS','aliases':['pbs'],'prefix':True},
    {'name':'SPD','aliases':['spd'],'prefix':True},
    {'name':'MCTT','aliases':['mctt'],'prefix':True},
    {'name':'換気扇','aliases':['換気扇'],'prefix':True},
    {'name':'コンセント','aliases':['コンセント'],'prefix':True},
    {'name':'伝送','aliases':['伝送'],'prefix':True},
    {'name':'R.RY','aliases':['リモコンリレー'],'prefix':False},
    {'name':'R.TR','aliases':['リモコントランス'],'prefix':False},
    {'name':'TB','aliases':['端子台','m)tb','b)tb'],'prefix':True},
]

def _detect_main(main_str):
    import re as _re
    n=norm(main_str)
    # 最優先: B)/M) 接頭辞付きの遮断器(MCB/ELB/MCCB/ELCB/LUG)は、負荷名称より先に判定。
    # 「主幹/主/主開閉器」等の和名接頭辞が付いても遮断器語を拾えるようにする(主幹 MCCB 等)。
    m=_re.match(r'(?:主幹用?|主開閉器?|主)?\s*(b\)|m\))?(mccb|mcb|elcb|elb|lug)', n)
    if m:
        kind=m.group(2)
        label={'mccb':'MCB','mcb':'MCB','elcb':'ELB','elb':'ELB','lug':'LUG'}[kind]
        pre=(m.group(1) or '')
        return kind, (pre.upper()+label if pre else label), True
    # 分電盤分岐: 負荷名が先頭に付く形(例「電灯 ELCB2P 50/20AT」「予備 MCCB2P 50/20AT」)。
    # 遮断器+極数(2P/3P/4P)が名称途中にあれば分岐遮断器として拾う(負荷名で埋もれるのを救済)。
    m2=_re.search(r'(mccb|elcb|mcb|elb)[^a-z]{0,2}[1-4]\s*p', n)
    if m2:
        kind=m2.group(1)
        return kind, {'mccb':'MCB','mcb':'MCB','elcb':'ELB','elb':'ELB'}[kind], True
    # 遮断器キーワードが無くても「NP + NNNAF(/MMAT)」形式は成形遮断器(MCCB)分岐とみなす。
    # 動力盤/コンセント盤の分岐(例「フォーク用コンセント 3P 100AF/60AT」「ACP 3P 225AF/125AT」)。
    # AF(フレーム定格)は成形遮断器固有の表記なので、負荷名が先頭でも遮断器分岐と確定できる。
    # 「NP+NNNAF」または「NP+NNN/MMM」(枠/トリップ対、AF表記なし。例「分岐 3P225/225」)を分岐遮断器とみなす。
    m3=_re.search(r'[1-4]\s*p.{0,5}\d+\s*af', n) or _re.search(r'[1-4]\s*p\s*\d{2,4}\s*/\s*\d{2,4}(?![\d.]|v)', n)
    if m3:
        kind='elb' if _re.search(r'(elb|elcb|漏電|漏保|el)(?![a-z])', n) else 'mcb'
        return kind, {'mcb':'MCB','elb':'ELB'}[kind], True
    # 主幹/主開閉器: 極数表記が無くてもAF枠(NNNAF)や枠/トリップ対(NNN/MMM)を持つ主幹遮断器はMCB主幹として拾う。
    # (開閉器＝MCCBを主幹遮断器として使う図面。AF枠は成形遮断器固有表記なので確定できる。極数既定3P。)
    if _re.search(r'主幹|主開閉|開閉器|主機', n):
        m4=_re.search(r'\d+\s*af', n) or _re.search(r'\d{2,4}\s*/\s*\d{2,4}(?![\d.]|v)', n)
        if m4:
            kind='elb' if _re.search(r'(elb|elcb|漏電|漏保)', n) else 'mcb'
            return kind, {'mcb':'MCB','elb':'ELB'}[kind], True
    # 変圧器(TR): 「T:」始まり、または 相数φ+KVA を持つものはTRとして最優先判定。
    # (二次電圧の "S:210V" が "6600vs210" のように VS と誤マッチするのを防ぐ)
    if _re.match(r't\s*[:：]', n) or (_re.search(r'[13]φ', n) and _re.search(r'\d+\s*kva', n) and 'kvar' not in n):
        # SC/SR(コンデンサ/リアクトル)は別系列なので除外
        if not _re.search(r'(^|[^a-z])(sc|sr)([^a-z]|$)', n):
            return 'tr','TR',False
    # 部品→別名リストで照合（同じ部品の英語/日本語/略称をまとめて判定）
    for part in _PARTS:
        for alias in part['aliases']:
            nk=norm(alias)
            if len(nk)<=2:
                # 短い別名(as/vs/am/vm/ct/vt等)は単語境界でのみマッチ
                if _re.search(r'(?<![a-z])'+_re.escape(nk)+r'(?![a-z])', n):
                    # 計器別名(vm/am/vs/as)は、負荷回路名の途中に含まれる文字と誤マッチしやすい。
                    # 実計器は名称の先頭に来る(「VM 広角」等)。VA負荷があり かつ 別名が先頭でない
                    # (前に別の語がある)場合は負荷回路名の一部と見なし、計器としない。
                    if nk in ('vm','am','vs','as') and _re.search(r'\d+\s*va', n) and n.find(nk)>0:
                        continue
                    # DB照合には別名でなく正式名(part['name'])を返す。
                    # (例: VMCで検出→DB品名は'VCS'始まりなので'VCS'で照合する)
                    return part['name'], part['name'], part['prefix']
            else:
                if nk in n:
                    return part['name'], part['name'], part['prefix']
    return None,None,False

# ===== 制御盤 分岐回路(22-29系)選定: 回路種別＋kW(切上)＋●○/AX で1コード =====
# (poc_control_v2 由来。小数kWを正しく扱う=旧_set_codeの「2.2→22」バグを解消。スターデルタ対応)
def _parse_bunki(name):
    m=re.match(r'分岐回路\s*(\([^)]*\))?\s*(.+?)\s*([\d.]+)\s*(?:KW|kW)?\s*$', name)
    if not m: return None
    return (m.group(1) or '').strip('()'), m.group(2).strip(), float(m.group(3))
_BUNKI_INDEX={}   # (dev, typ, kw, volt) -> code。200V=22-25系 / 400V=26-29系。
for _d in DB:
    if '分岐回路' in _d.get('name',''):
        _p=_parse_bunki(_d['name'])
        if _p:
            _volt='400V' if _d['code'][:2] in ('26','27','28','29') else '200V'
            _BUNKI_INDEX[_p+(_volt,)]=_d['code']
def _bunki_steps(dev,typ,volt): return sorted(k for (dv,ty,k,vo) in _BUNKI_INDEX if dv==dev and ty==typ and vo==volt)
def _bunki_find(typ,kw,dev,volt):
    for cd in ([dev,''] if dev else ['']):   # AX/変種無しへフォールバック
        steps=_bunki_steps(cd,typ,volt)
        if not steps: continue
        pick=next((s for s in steps if kw<=s+1e-9), steps[-1])
        c=_BUNKI_INDEX.get((cd,typ,pick,volt))
        if c: return c,('◎' if cd==dev else '○'),pick
    return '','△',None

def _set_code(name, vb):
    """制御盤の負荷名(例「2.2KW (L-S)」「INV 3.7kW」「スターデルタ 15kW」)→分岐回路コード。
    回路種別＋kW(型ごとステップで切上)＋●○(MCCB/ELB)/AX。小数kWを正確に扱う。"""
    s=str(name); U=s.upper()
    # kW: 小数を保持して抽出(normは小数点を削るため使わない)
    km=re.search(r'([\d.]+)\s*KW', U)
    if not km: km=re.search(r'(?<![\d.])([\d]+\.[\d]+|[\d]+)(?![\d.])\s*$', s.strip())
    if not km: return None,None
    try: kw=float(km.group(1))
    except: return None,None
    if kw<=0: return None,None
    # 回路種別(DB表記に合わせる)。長音ー(U+30FC)とハイフン両対応。
    if re.search(r'スタ[ー\-]?デルタ|ｽﾀ[ｰ\-]?ﾃﾞﾙﾀ|STAR[\- ]?DELTA|Y[\-]?Δ', U) and 'INV' not in U: typ='スターデルタ'
    elif 'INV' in U or 'インバータ' in s:
        typ='INV(スターデルタ)' if re.search(r'スタ|ｽﾀ|Δ|STAR', U) else 'INV'
    elif re.search(r'L[ー\-]?S', U): typ='L-S(AM付)'
    else: return None,None
    volt='400V' if vb=='400V' else '200V'
    if '支給' in s:  # INV支給品等
        for cand in (typ+'(支給品)', typ):
            if any(ty==cand for (_dv,ty,_k,_v) in _BUNKI_INDEX): typ=cand; break
    # ●=MCCB基本('') / ○=ELB / 遠方操作=AX
    dev=''
    if 'ELB' in U or '○' in s: dev='ELB'
    if 'AX' in U or '遠方' in s: dev=('ELB・AX' if dev=='ELB' else 'MCB・AX')
    code,conf,pick=_bunki_find(typ,kw,dev,volt)
    if code: return code,f'分岐回路 {typ} {pick}kW {volt}'
    return '',f'分岐回路 {typ} {kw}kW {volt}・容量外/該当なし要確認'

# ===== 制御盤 適用表→分岐回路コード（たすき掛け変換）=====
# ①パターン集の凡例(記号→回路種別)を覚え ②適用表の記号を読んだら即その型に変換 ③型+kW+●○でコード。
# 記号→回路種別の既定対応(図面に凡例があればそれで上書き。図面ごとに割付が違う場合に備える)
_PATTERN_TYPE_DEFAULT={
    'A':'L-S(AM付)','B':'L-S(AM付)','C':'L-S(AM付)','E':'L-S(AM付)',
    'D':'スターデルタ','F':'スターデルタ','G':'スターデルタ','H':'スターデルタ',
    'I':'INV','J':'INV','L':'INV','K':'INV(スターデルタ)'}
def _normalize_type(t):
    """凡例のタイトルを DB表記(L-S(AM付)/スターデルタ/INV/INV(スターデルタ))に正規化。
    直接遮断器のみ(電源/コンセント)は 'MCCB'(=モーター分岐回路でない印)を返す。"""
    u=str(t or '').upper(); s=str(t or '')
    if re.search(r'スタ[ー\-]?デルタ|ｽﾀ[ｰ\-]?ﾃﾞﾙﾀ|STAR|[YTΥ][ー\-]?Δ|Δ始動',u+s):
        return 'INV(スターデルタ)' if 'INV' in u else 'スターデルタ'
    if 'INV' in u or 'インバータ' in s: return 'INV'
    # 「電源送り/電源供給/送り/操作電源」はモーター始動回路でなく単なる遮断器(電源送り)。
    # 電流計付(B=電源送り(MCCB,電流計付))でも送りは送り→MCCB扱い(電流計でL-S誤判定しないよう先に判定)。
    if re.search(r'電源送り|電源供給|(?<![直])送り|操作電源|操作用電源', s) and not re.search(r'始動|直入|L[ー\-]?S(?![a-z])', s+u):
        return 'MCCB'
    if re.search(r'直入|電流計|[0-9０-９]回路|[0-9０-９]台|L[ー\-]?S',s+u): return 'L-S(AM付)'
    if re.search(r'MCCB|ELB|遮断器',u) and 'のみ' in s: return 'MCCB'   # 直接遮断器=電源/コンセント回路
    return str(t or '').strip()
def _dev_from_breaker(brk, ax=False):
    b=str(brk or ''); U=b.upper()
    elb=('○' in b) or ('ELB' in U) or ('ELCB' in U) or ('漏電' in b)
    if elb: return 'ELB・AX' if ax else 'ELB'
    return 'MCB・AX' if ax else ''
def control_apply(load_rows, legend=None, volt='200V'):
    """動力制御盤 適用表→分岐回路コード。
    load_rows: [{'load':負荷名,'pattern':主回路記号,'kw':容量,'breaker':'●/○/MCCB/ELB','ax':bool,'spare':bool}]
    legend: 図面のパターン凡例 {記号:型}(無ければ既定 _PATTERN_TYPE_DEFAULT)。volt:'200V'/'400V'
    戻り: [{'load','code','conf','note'}]。◎/○=確定, △=記号不明/kW不明/容量外(→確認ゲート)。"""
    leg=dict(_PATTERN_TYPE_DEFAULT)
    for k,v in (legend or {}).items(): leg[str(k).strip().upper()]=_normalize_type(v)
    out=[]
    for r in load_rows:
        load=r.get('load','') or ''
        if r.get('spare') or str(load).strip() in ('予備','スペース','ｽﾍﾟｰｽ',''):
            continue
        sym=str(r.get('pattern','')).strip().upper()
        try: kw=float(re.sub(r'[^\d.]','',str(r.get('kw','') or 0)) or 0)
        except: kw=0
        typ=leg.get(sym)
        if not typ: out.append({'load':load,'code':'','conf':'△','note':f'記号"{sym}"→回路種別不明(凡例要確認)'}); continue
        if typ=='MCCB':   # 電源/コンセント等 直接遮断器 → 遮断器サイズで個別選定(分岐回路でない)
            brk=str(r.get('breaker') or '').strip()
            if brk:
                sel=select_one('MCB '+brk,'動力制御盤')
                out.append({'load':load,'code':sel.get('code',''),'conf':sel.get('conf','△'),'note':f'直接遮断器 {brk}'})
            else:
                out.append({'load':load,'code':'','conf':'△','note':'直接遮断器 容量要確認'})
            continue
        if kw<=0: out.append({'load':load,'code':'','conf':'△','note':'kW不明→確認'}); continue
        dev=_dev_from_breaker(r.get('breaker'), r.get('ax'))
        code,conf,pick=_bunki_find(typ,kw,dev,volt)
        if code: out.append({'load':load,'code':code,'conf':conf,'note':f'{typ}({sym}) {pick}kW {volt}'})
        else: out.append({'load':load,'code':'','conf':'△','note':f'{typ} {kw}kW 容量外→確認'})
    return out

def gen_candidates(name, volt='', panel=''):
    name=re.sub(r'(?i)vgb','VCB',str(name))  # VGBはVCBの別表記
    """属性で候補を生成。戻り: (meta, [DB行,...])"""
    import re as _re
    qname,qty=split_qty_suffix(name)
    vb=_voltband(volt,qname)
    main_str,opts=split_main_opt(qname)
    main_kw,main_label,prefix=_detect_main(main_str)
    n=norm(qname)
    kw=_get_kw(n)
    # 容量(kva/kvar)は小数点を保持するため元文字列から抽出(normは小数点を削るため)
    _km=_re.search(r'(\d+\.?\d*)\s*k?va[r]?', qname, _re.I)
    kva=_km.group(1) if _km else _get_kva(n)
    mr=_re.search(r'(\d+)/5a',n); ratio=mr.group(1) if mr else None
    af=_re.search(r'(\d+)af',n); af=af.group(1) if af else None
    dvals=attr_value_set(qname, '')
    meta=dict(qty=qty,qname=qname,main=main_label,main_kw=main_kw,opts=opts,vb=vb,kw=kw,kva=kva,ratio=ratio,af=af,dvals=dvals)
    out=[]
    for d in DB:
        if d.get('settype'): continue  # 配電盤セットコードは専用セレクタ(setcode)で選定。個別選定からは除外。
        dn=norm(d['name'])
        if main_kw:
            if prefix:
                if not dn.startswith(norm(main_kw)): continue
            else:
                if norm(main_kw) not in dn: continue
        sc=10
        dvb=_dbvolt(d['volt'])
        if vb and dvb:
            if vb==dvb: sc+=6
            else: continue
        # 値ベース属性照合（極数・容量・AF・変流比・電圧を値で突合）
        vsc,_cv=value_score(dvals, d)
        sc+=vsc
        sc+=option_bonus(opts, d['name'], qname)
        # 図面にオプション語が無い場合は「標準/既定」を優先（特殊型を後ろへ）
        if not opts and any(k in d['name'] for k in ['スペース','(SP)','コージェネ','高性能']):
            sc-=3
        out.append((sc,d))
    out.sort(key=lambda x:-x[0])
    return meta,[d for s,d in out[:8]]

# 第2段: 候補をルールで1つに絞り、信頼度を付ける
def _extra_meter_code(name):
    """追加計器(IWM無効電力計/PFM力率計)を計器種別(広角/普通角)でコード確定。
    VM/AM等の基本計器は従来の別名照合で処理されるためここでは追加計器のみ。"""
    n=norm(name); hiro=('広角' in name) or bool(re.search(r'(?<![a-z])広(?![角])', name))
    if re.search(r'(?<![a-z])iwm(?![a-z])|無効電力計?', n):
        c='42030' if hiro else '42020'
        return c if c in byCode else None
    if re.search(r'(?<![a-z])pfm(?![a-z])|力率計', n):
        c='42032' if hiro else '42022'
        return c if c in byCode else None
    return None

def refine(meta, cands, name, panel, prev_is_main=False, volt=''):
    n=norm(name); vb=meta['vb']
    H=lambda *xs: all(norm(x) in n for x in xs)

    # --- 会社確認: TB文脈判定 ---
    is_tb=(meta['main']=='TB')
    if is_tb:
        if prev_is_main:
            amp=_get_amp(n)
            ctrl=any(k in norm(panel) for k in ['制御','動力','pac','m1','m2','m3','m4'])
            if ctrl:
                cmap={'50':'50901','100':'50902','200':'50903','225':'50903','400':'50904'}
                if amp in cmap and cmap[amp] in byCode: return R(cmap[amp],'○','主幹用TB(制御盤二次TB)単独計上(会社確認)')
            mtb={'50':'68509','100':'68109','225':'68209','200':'68209','400':'68409','600':'68609'}
            if amp in mtb and mtb[amp] in byCode: return R(mtb[amp],'○','主幹用TB(M)TB)単独計上(会社確認)')
            return R('','△','主幹用TB 単独計上・容量要確認(会社確認)')
        else:
            return R('','○','分岐B)系のTBは本体内包・計上せず(会社確認)')

    # --- SPD用MCCB: 容量はSPDの種類で決まる。容量表記が無ければ△(要確認)で目安提示 ---
    if 'spd用' in n and ('mcb' in n or 'mccb' in n):
        af,at=_amp_af(n)
        if not af:
            cls='クラスII' if re.search(r'ii|2', n) else ''
            return R('','△','SPD用MCCB 容量表記なし→要確認(目安: 低圧分電盤クラスII=2P/3P 20〜50A, 主幹近く=3P 50〜100A)')
        # 容量表記があれば通常のMCBとして選定(下のMCBロジックへ)

    # --- M)LUG(主幹端子): SPD等の分岐頭に断路スイッチのみ(MCCB表記なし)=端子受け。
    #     負荷頭の遮断器(M)MCB)とは別物。端子台系統(40/50/60)×容量(切上)でLUGコードを計上(○)。
    #     スイッチ無し(直結)の場合は抽出側でM)LUGを出さない方針(茂泉様確定)。---
    if meta['main'] in ('M)LUG','LUG') or norm(name).startswith(('m)lug','lug')):
        lc=_lug_code(name, panel, meta)
        if lc: return R(lc,'○','主幹LUG(SPD分岐頭・断路のみ/端子受け・%s)'%byCode.get(lc,{}).get('name','')[:16])
        return R('','△','M)LUG 容量・盤種別要確認')

    # --- 主幹/分岐 MCB・ELB（MCBが主部品。M)=主幹, B)=分岐）---
    # 盤頭の主幹だが遮断器型が非定型でmeta未検出のもの(実測:東部P-5/P-6「主幹 EM-CET200 3φ3W 200V」等)。
    # 盤頭のM)遮断器として容量を最近傍上位で丸め、M)MCB/ELB/LUG候補を○提示(型・容量は確認ゲート)。行き止まり△解消。
    if re.search(r'主幹|主開閉器', str(name)) \
       and meta.get('main') not in ('M)MCB','M)ELB','B)MCB','B)ELB','MCB','ELB','M)LUG') \
       and _panel_kind(panel) in ('ctrl','haiden','bunden'):
        _am=re.search(r'(?:CET|CB|EA|EM|CS|CE)[\s\-]?(\d{2,4})', str(name), re.I) or re.search(r'(\d{2,4})\s*A(?![A-Za-z])', norm(name))
        if _am:
            _at=int(_am.group(1)); _FD=[(100,'1'),(225,'2'),(400,'4'),(600,'6'),(800,'8')]
            _fd=next((d for a,d in _FD if a>=_at), '8')
            _ser=_KIND_SERIES.get(_panel_kind(panel),'50')
            # 既定はM)LUG(盤頭端子受け・x09)。実見積書=東部の制御盤主幹は全てM)LUG(50109/50209/50409で20件、
            # 非AXのM)MCBは0件)。型式非定型の主幹は端子受け扱いが定番。M)MCB/ELBは候補提示で確認ゲート。
            _mc=[_ser+_fd+s for s in ('09','03','06') if (_ser+_fd+s) in byCode]
            if _mc:
                _r=R(_mc[0],'○','盤頭の主幹(型式非定型)→M)端子/遮断器候補・型/容量は確認ゲート(既定=%s)'%byCode.get(_mc[0],{}).get('name','')[:16])
                _r['candidates']=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in _mc]
                return _r
    if meta['main'] in ('M)MCB','M)ELB','B)MCB','B)ELB','MCB','ELB'):
        # 分電盤の最終分岐(負荷VA表記+MCB/ELB明示・遮断器枠なし)はコンパクト分岐を優先判定。
        # ※_mcb_codeより先に判定(「280VA」等のVA数値をAF枠と誤解して誤コードを出すのを防ぐ)。
        cb=_compact_branch(name, panel)
        if cb: return R(cb[0],cb[1],cb[2])
        code=_mcb_code(name, panel, meta)
        if code:
            if code not in byCode: return R('','△', _mcb_note(name,panel))
            # 制御盤(50/51系)でAXを「警報付」から推定した個別遮断器は、AX付の有無・主幹/分岐・極数の変種に
            # 誤りが混じり得る(実測:東部で225AF主幹をB)分岐、600A/2P等の取り違え)。◎誤答ゼロのため○止め+
            # AX/非AX両候補を提示し確認ゲートへ委ねる(コード自体は保持=一致率は不変)。明示AX/非AXの遮断器は従来◎。
            _soft_ax = (_panel_kind(panel)=='ctrl'
                        and bool(re.search(r'警報付|警報出力', str(name)))
                        and not re.search(r'AX付?|(?<![A-Za-z])AX(?![A-Za-z])|中欠|欠相|補助接点', str(name)))
            _r=R(code,'○' if _soft_ax else '◎', _mcb_note(name,panel)+('（警報付→AX付と推定・AX有無/主幹分岐は確認ゲート）' if _soft_ax else ''))
            if _soft_ax and len(code)==5 and code[1]=='1':
                _nax=code[0]+'0'+code[2:]
                _r['candidates']=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in (code,_nax) if c in byCode]
            return _r
        # 主幹ELB(『ELB付 主幹』): 変種(EL・中欠/AX)が図面に明記されないため素のM)ELBを○(要確認)。
        me=_main_elb_code(name, panel, meta)
        if me: return R(me,'○','主幹ELB(%s・EL/中欠/AX変種は要確認)'%byCode.get(me,{}).get('name','')[:18])
        # 主幹(M)MCB/M)ELB)で容量が読めない場合も、行き止まり△でなく盤種のM)MCB候補(100/225/400/600A)を
        # 提示して○(容量は確認ゲートで確定・既定100A)。確定率100%方針。
        # MCB/ELBで容量が読めない場合も、行き止まり△でなく盤種のM)MCB/M)ELB候補(100/225/400/600A)を
        # 提示して○(容量は確認ゲートで確定・既定100A)。確定率100%方針。
        _ser=_KIND_SERIES.get(_panel_kind(panel),'60')
        _is_elb=bool(re.search(r'EL[CB]', str(name), re.I))
        # ELBはM)ELB(x06)優先だが、系統にM)ELBが無ければM)MCB(x03)へフォールバック(型は確認ゲート)。
        _suf=('106','206','406','606','103','203','403','603') if _is_elb else ('103','203','403','603')
        _mc=[_ser+cap for cap in _suf if (_ser+cap) in byCode]
        if _mc:
            _r=R(_mc[0],'○','容量が図面から不明・確認ゲートで容量/極数確定(既定100A)')
            _r['candidates']=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in _mc]
            return _r
        return R('','△','MCB/ELB 容量・盤種別要確認')
    # --- 制御盤 分岐回路(L-S/スターデルタ/INV) 最優先 ---
    # 回路種別＋kW＋●○が読めれば決定的にコード確定→○(回路種別の推定余地を残し安全側)。容量外は△。
    sc_code,sc_note=_set_code(name,vb)
    if sc_code: return R(sc_code,'○',sc_note)
    if sc_code=='' and sc_note: return R('','△',sc_note)

    # --- 高圧/低圧CT 変流比判定 ---
    if meta['main']=='CT' and meta['ratio']:
        r=int(meta['ratio'])
        lv={'10':'72000','15':'72001','100':'72002','200':'72003','300':'72004','400':'72005','500':'72006','600':'72007'}
        # 72系(低圧CT)は変流比を最近傍上位で丸める(150/5A等の非標準比が44系へ落ちる誤りを防ぐ)。
        _lv72_list=[(10,'72000'),(15,'72001'),(100,'72002'),(200,'72003'),(300,'72004'),(400,'72005'),(500,'72006'),(600,'72007')]
        def _lv72(rr):
            for a,c in _lv72_list:
                if rr<=a and c in byCode: return c
            return ''
        # 44系 低圧CT(受変電低圧配電盤の主計器用・全変流比)。600A超は44系のみ(72系は600Aまで)。
        lv44=[(300,'44141'),(400,'44142'),(500,'44143'),(600,'44144'),(800,'44145'),(1000,'44146'),
              (1200,'44147'),(1500,'44148'),(2000,'44149'),(3000,'44150'),(4000,'44151'),(5000,'44152')]
        def _lv44(rr):
            for a,c in lv44:
                if rr<=a and c in byCode: return c
            return ''
        _hvc='44121' if r<=40 else '44122' if r<=75 else '44123' if r<=200 else ''
        # 受変電低圧配電盤(電灯盤/動力盤=haiden)の計器CTは44系(実見積書:城山/六本木の電灯・動力盤)。
        _pn=norm(panel)
        _is_haiden_lv = bool(re.search(r'電灯盤|動力盤', _pn)) and not re.search(r'制御|分電', _pn)
        if vb=='HV':
            if _hvc: return R(_hvc,'○',f'高圧CT 変流比{meta["ratio"]}/5A')
            # 高圧CTの標準範囲(≤200)外は44系(全変流比)へ最近傍上位でフォールバック→○(行き止まり△解消)。
            c44=_lv44(r)
            if c44: return R(c44,'○',f'高圧CT 変流比{meta["ratio"]}/5A(標準範囲外→44系最近傍・要確認)')
            return R('44123','○','高圧CT 変流比が標準範囲外・確認ゲートで容量確定')
        if vb in('200V','400V','100V'):
            if r>600 or (_is_haiden_lv and r>=300):
                c44=_lv44(r)
                if c44: return R(c44,'○',f'低圧CT {meta["ratio"]}/5A(44系)')
            _c72=_lv72(r)
            if _c72: return R(_c72,'○',f'低圧CT {meta["ratio"]}/5A(72系・最近傍上位)')
            c44=_lv44(r)
            if c44: return R(c44,'○',f'低圧CT {meta["ratio"]}/5A(44系)')
        # 電圧帯不明→盤種で最善推定(受変電/高圧盤=高圧CT / それ以外=低圧CT)。行き止まり△を無くす。
        if re.search(r'受電|受変電|高圧|饋電|コンデンサ', str(panel)):
            if _hvc: return R(_hvc,'○',f'CT 変流比{meta["ratio"]}/5A(受変電→高圧CT・電圧帯要確認)')
        if r>600 or (_is_haiden_lv and r>=300):
            c44=_lv44(r)
            if c44: return R(c44,'○',f'CT 変流比{meta["ratio"]}/5A(低圧44系・電圧帯要確認)')
        _c72=_lv72(r)
        if _c72: return R(_c72,'○',f'CT 変流比{meta["ratio"]}/5A(低圧72系・電圧帯要確認)')
        c44=_lv44(r)
        if c44: return R(c44,'○',f'CT 変流比{meta["ratio"]}/5A(低圧44系・電圧帯要確認)')
        if _hvc: return R(_hvc,'○',f'CT 変流比{meta["ratio"]}/5A(高圧CT最善推定・電圧帯要確認)')
        return R('','△','CT 電圧帯不明・要確認')

    # --- TR/SC/SR 容量選定: 仕様値「以上」かつ「最も近い」コードを選ぶ ---
    # 区分: 図面に「支給」明記→支給品 / 「スペース」明記→スペース / 既定→購入品
    #   ただし高圧TR/SC/SRには購入品コードが無いため、既定でも支給品系を使う。
    # 容量は仕様以上で最小(最近傍上位)。完全一致→◎ / 繰上げ→○ / 超過→△
    _grp=None
    if meta['main']=='TR': _grp='TR'
    elif meta['main']=='SC': _grp='SC'
    elif meta['main']=='SR': _grp='SR'
    if _grp and meta.get('kva'):
        try:
            want=float(meta['kva'])
            grp_name=_grp; vb=meta.get('vb')
            nm_in=str(name)
            nn_in=norm(nm_in)
            is_shikyu=('支給' in nm_in)
            is_space=('スペース' in nm_in or 'SP' in nm_in)
            # 区分ラベル(品名接頭の判定用)
            if is_shikyu: kbn, kbn_label = '支給品','支給品'
            elif is_space: kbn, kbn_label = 'スペース','スペース'
            else: kbn, kbn_label = '購入','購入品'
            # 高圧/低圧は品名の電圧表記から直接判定(vbは不正確なため使わない)。
            #   高圧: 6.6kV/6600V/7.2kV/7.02kV/3.3kV 等の kV表記、または 200kVA級の高圧SC
            #   低圧: 200V/400V/105V/210V の明示
            is_lv=bool(re.search(r'(^|[^.\d])(100|105|200|210|400|415)v(?![a-z])', nn_in))
            # kV(高圧電圧)判定: kvの直後がa(kva)やr等の英字でないこと。kvar/kvaを除外。
            is_hv=bool(re.search(r'\d\.?\d*kv(?![a-z])', nn_in)) or bool(re.search(r'6600v|3300v|6900v|7020v|7200v', nn_in))
            if is_lv: is_hv=False  # 低圧電圧明示が最優先
            if not is_hv and not is_lv:
                # 電圧表記が無い場合: TR/高圧SCはkVA表記(高圧)、低圧SC/SRはkvar+電圧で既に判定済
                is_hv = (grp_name in('TR',) ) or (grp_name=='SC' and 'kvar' not in nn_in)
            # TRの一次が高圧(6.6kV/6600V/3.3kV/7.2kV等)なら高圧トランスとして確定。
            # 二次側の低圧電圧(210V/105V等)に惑わされない(is_lvの打消しより優先)。
            # ※norm後は小数点が消えるため "66kv"(6.6kV)/"72kv"(7.2kV)/"33kv"(3.3kV)で照合。
            # 一次/二次比の表記(例 6600/210V)は二次の低圧Vだけ拾って低圧誤判定しやすい。
            # 4桁一次(3300/6600/6900等)＋"/"＋二次 の比表記があれば高圧一次と確定。
            if grp_name=='TR' and re.search(r'(6600v|66kv|3300v|33kv|72kv|702kv|\dkv/|\d{4}\s*/\s*\d)', nn_in):
                is_hv=True; is_lv=False
            # 系統(ライン)の電圧文脈で判定: 高圧コンデンサ盤のSR/SCは、
            # 前後の接続機器(SC本体・VMC)が高圧7.02kVなので、機器に234V/200V等の
            # 低圧表記があっても系統は高圧。低圧表記に引きずられず高圧(45系)とする。
            # コンデンサ盤(受変電の進相コンデンサ)は既定で高圧45系。低圧進相コンデンサ盤(53系)は
            # 盤名に「低圧」が明示されるので、それ以外の「コンデンサ盤」は高圧扱い。
            # ※実見積書照合で東部SCが「50kvar 200V」を低圧53系と誤選択(正解45系)だった反省に基づく緩和。
            if grp_name in ('SR','SC') and 'コンデンサ' in str(panel) and '低圧' not in str(panel):
                is_hv=True; is_lv=False
            unit='kVA' if grp_name=='TR' or (grp_name=='SC' and is_hv) else 'kvar'

            # 高圧で購入指定だが購入品が無い→支給品代替。ラベルを明示。
            if is_hv and kbn=='購入':
                kbn_label='支給品(高圧は購入品なし)'

            def _match_series(d):
                """この行が(grp/区分/電圧帯/相数)に合致するか"""
                n0=d['name']
                if not n0.startswith(grp_name): return False
                hv_row=(d['code'][:2]=='45')
                if is_hv and not hv_row: return False
                if (not is_hv) and hv_row: return False
                # 区分: 支給品/スペース/購入(=どちらの語も付かない素のコード)
                has_shikyu='支給品' in n0; has_space='スペース' in n0 or '(SP)' in n0
                if kbn=='支給品' and not has_shikyu: return False
                if kbn=='スペース' and not has_space: return False
                if kbn=='購入':
                    if is_hv:
                        # 高圧は購入品が無いので支給品を代替採用
                        if not has_shikyu: return False
                    else:
                        # 低圧購入品は素のコード(支給品/スペース語なし)
                        if has_shikyu or has_space: return False
                # 低圧の電圧(200V/400V)一致
                if not is_hv:
                    vm=re.search(r'(200|400)v', nn_in)
                    if vm and f"{vm.group(1)}V" not in n0: return False
                # TRの相数
                if grp_name=='TR':
                    if re.search(r'3φ',nn_in) and '3φ3W' not in n0: return False
                    if re.search(r'1φ',nn_in) and '1φ3W' not in n0: return False
                # SRのL%(直列リアクトルのリアクタンス%)。「L=6%」「(6%)」「6%」いずれの表記も拾う。
                # L%が読めたら、その%のコードだけに絞る(6%指定なのにL=13%を選ぶ誤りを防ぐ)。
                if grp_name=='SR' and is_hv:
                    lm=re.search(r'l\s*=?\s*(\d+)\s*%', nn_in) or re.search(r'(?<![\d.])(\d+)\s*%', nn_in)
                    if lm and f"L={lm.group(1)}%" not in n0: return False
                return True

            pool=[]
            for d in DB:
                if not _match_series(d): continue
                mm=re.search(r'(\d+\.?\d*)\s*kva[r]?', d['name'], re.I)
                if mm: pool.append((float(mm.group(1)), d['code']))
            if pool:
                # SC/SR(支給品)は銘板実測値(7020V時31.9kvar等)が公称値の前後になり、切下/切上のどちらを
                # 採るかは案件で異なる(茂泉様)→前後の公称値を候補に確認ゲート化。既定=切下(公称値に近い)。
                if grp_name in ('SC','SR'):
                    exact=[c for v,c in pool if abs(v-want)<1e-6]
                    if exact: return R(exact[0],'◎',f'{grp_name}容量一致({want:g}{unit}・{kbn_label})')
                    below=sorted([(v,c) for v,c in pool if v<=want+1e-6], key=lambda x:-x[0])
                    above=sorted([(v,c) for v,c in pool if v>=want-1e-6])
                    fl=below[0] if below else None; ce=above[0] if above else None
                    if fl or ce:
                        d=(fl or ce)
                        _r=R(d[1],'○',f'{grp_name}容量 銘板{want:g}{unit}→切下/切上を確認ゲートで選択(既定=切下{d[0]:g}{unit}・{kbn_label})')
                        _r['_round_gate']=[x[1] for x in (fl,ce) if x and x[1] in byCode]
                        return _r
                ge=sorted([(v,c) for v,c in pool if v>=want-1e-6])
                if ge:
                    v0,c0=ge[0]
                    if abs(v0-want)<1e-6:
                        return R(c0,'◎',f'{grp_name}容量一致({v0:g}{unit}・{kbn_label})')
                    return R(c0,'○',f'{grp_name}容量繰上げ(仕様{want:g}→{v0:g}{unit}・{kbn_label})')
                vmax=max(v for v,_ in pool)
                return R('','△',f'{grp_name}容量{want:g}{unit}が{kbn_label}DB最大({vmax:g}{unit})超過・要確認')
            return R('','△',f'{grp_name} {kbn_label}・該当容量コードなし・要確認')
        except (ValueError,TypeError):
            pass

    # --- メイン機器が辞書で特定できない場合は△（誤った確定を出さない）---
    # 重要原則: 機器が特定できないときは選定コード欄を空(—)にする。
    # 候補先頭を入れると無関係コード(SR/低圧TR/SC/MGS/LA等)が選定欄に出て誤誘導するため。
    if not meta.get('main'):
        # 追加計器(IWM無効電力計/PFM力率計)は計器種別(普通角/広角)でコード確定。
        _emc=_extra_meter_code(name)
        if _emc: return R(_emc,'◎',byCode.get(_emc,{}).get('name',''))
        # 分電盤は「1負荷=1分岐回路」(社内方式・実見積書で確認)。遮断器記号が無い裸のVA負荷でも、
        # 分電盤ならコンパクト分岐(60012/60014)で確定できる。湿式・接触注意負荷はELB、乾式はMCB。
        _bl=_bunden_load_branch(name, panel)
        if _bl: return R(_bl,'○','分電盤の負荷回路→コンパクト分岐(1負荷=1回路・%s)'%('ELB' if _bl=='60014' else 'MCB'))
        # 注: 遮断器記号(MCB/ELB等)が付く分岐は _compact_branch がコンパクトで確定する(上流で処理)。
        # ここに来るのは遮断器種別が全く無い「裸のVA負荷/機器名」=分岐か否かも図面から判別不能。
        # 分電盤以外(制御盤/配電盤)では負荷ごとに勝手に分岐を立てると過剰計上になるため安全側△(人が確認)。
        if cands:
            hint=cands[0]['code']
            return R('','△',f'機器未特定・要確認(参考候補{len(cands)}件: 先頭={hint})')
        return R('','△','該当機器が辞書に無い・要確認')
    # 高圧DS(断路器): 極数が明記されていれば、その極数の標準品で確定する。
    # 単線図の「DS×3」は三相を表す(=3P 1台)。split_qty_suffixで3P化される。
    _dsname=meta.get('qname') or name
    if re.search(r'(?<![a-z])ds(?![a-z])', norm(_dsname)) and re.search(r'3p', norm(_dsname)):
        amp=re.search(r'(\d+)a(?![a-z])', norm(_dsname))
        if amp:
            a=amp.group(1)
            c3=next((c['code'] for c in cands if '3P' in c['name'] and f'{a}A' in c['name'] and '標準' in c['name']),'')
            if c3:
                return R(c3,'○',f'DS 3極{a}A(標準)で確定(×3は三相表記のため1台)')
    # 高圧DS(断路器): 単極用と3極用がある。極数明記が無い場合、
    # 「単極をN個」か「3極を1個」かの判断が要るため△で両論を注記する。
    if re.search(r'(?<![a-z])ds(?![a-z])', norm(_dsname)) and not re.search(r'[123]p', norm(_dsname)):
        amp=re.search(r'(\d+)a(?![a-z])', norm(_dsname))
        if amp:
            a=amp.group(1)
            c1=next((c['code'] for c in cands if '1P' in c['name'] and f'{a}A' in c['name'] and '標準' in c['name']),'')
            c3=next((c['code'] for c in cands if '3P' in c['name'] and f'{a}A' in c['name'] and '標準' in c['name']),'')
            if c1 or c3:
                return R(c3 or c1,'△',f'DS極数要確認: 単極{a}A({c1})×N個 or 3極{a}A({c3})×1個')
    # LGR/LG-RY: 図面に「2段警報」明記が無ければ ZCT付(46401)を標準採用。
    if re.search(r'(?<![a-z])(lgr|lg-ry)(?![a-z])', norm(name)):
        if '2段' not in str(name) and '警報' not in str(name):
            if '46401' in byCode: return R('46401','◎','LGR ZCT付(標準)で確定')
        else:
            if '46403' in byCode: return R('46403','○','LGR 2段警報で確定')
    # --- 候補数で信頼度を決定 ---
    if not cands:
        return R('','△','該当コードなし・要確認')
    # VMC(真空電磁接触器)→VCS(43系)。御社の実見積書で確定した規則で変種を決める(PF/引出=43103, 素=43101)。◎。
    v=_vmc_code(name)
    if v:
        _rv=R(v[0],v[1],v[2])
        if len(v)>3 and v[3]: _rv['candidates']=v[3]
        return _rv
    if len(cands)==1:
        return R(cands[0]['code'],'◎','属性一致(単一候補)')
    # 複数候補の最終判定
    top=cands[0]
    dvals=meta.get('dvals') or set()
    opts=meta.get('opts') or []
    tv=attr_value_set(top['name'], top.get('volt',''))
    matched=dvals & tv
    missing=dvals - tv
    # CT範囲を考慮（変流比のmissは範囲内なら解消）
    ratio_miss=[v for v in missing if v.endswith('/5A')]
    if ratio_miss and _ct_range_match(dvals, top['name']):
        missing=missing-set(ratio_miss)
    # 極数(P)のmissは、DB品名に極数表記が無い機器(PAS/VT/DGR等)では不問にする
    if any(m.endswith('P') for m in missing) and not re.search(r'[234]P', top['name']):
        missing={m for m in missing if not m.endswith('P')}
    # 電圧の具体値(100V/200V等)は、TR等で結線仕様にすぎない場合 missから除外
    # (相数φとKVAが一致していれば確定とみなす)
    if any(m.endswith('φ') for m in matched) and any(m.endswith('KVA') for m in matched):
        missing={m for m in missing if not (m.endswith('V'))}
    # オプション語がトップに一致しているか
    opt_hit = any(norm(o) and norm(o) in norm(top['name']) for o in opts)
    # 単一候補→◎
    if len(cands)==1:
        return R(top['code'],'◎','単一候補で確定')
    # 2位とのスコア差が大きい/オプション一致/属性過不足なし → ○
    if not missing and matched:
        return R(top['code'],'○',f'属性値一致({len(matched)}項)')
    if opt_hit and not missing:
        return R(top['code'],'○','オプション語一致で絞込')
    # 既定型(静止型/標準/手動)がトップで、図面に特別仕様の記述が無ければ○
    if not missing and any(k in top['name'] for k in ['静止型','標準','(手動)']):
        return R(top['code'],'○','標準/既定型として確定')
    # 属性不足/複数候補でも、候補(同一機器系統)がある以上は行き止まり△でなく○(最善候補＋候補提示)にし、
    # 確認ゲートで変種を確定できるようにする(確定率100%方針・茂泉様確定)。◎でないので誤答リスクなし。
    _cl=[{'code':c['code'],'name':byCode.get(c['code'],{}).get('name',''),'volt':c.get('volt','')} for c in cands[:6] if c.get('code')]
    _note=(f'属性不足{sorted(missing)}・確認ゲートで変種確定' if missing else f'候補{len(cands)}件・確認ゲートで変種確定')
    _r=R(top['code'],'○',_note)
    if _cl: _r['candidates']=_cl
    return _r

def _get_amp(n):
    import re as _re
    m=_re.search(r'(\d+)a', n); return m.group(1) if m else ''

# リモコン設備(65系): 数値属性が無く gen_candidates で落ちるため、名称で直接引く。
# 照合前に NFKC 正規化(半角カナ→全角: ﾀｲﾏ→タイマ 等)するのでパターンは全角で書く。
_REMOCON_MAP=[
 (r'伝送ユニット','65022'),
 (r'(年間)?プログラムタイマ|年間タイマ|タイマユニット|ﾀｲﾏﾕﾆｯﾄ','65019'),
 (r'EE.{0,4}連動|自動点滅.{0,4}連動|点滅器?連動','65018'),
 (r'自動点滅器(?!.{0,4}連動)','74082'),
 (r'(パターン|タイプ).{0,6}設定','65050'),
 (r'蛍光灯調光|調光.{0,3}[TＴ]/?U','65037'),
 (r'接点入力','65040'),
 (r'ノイズフィルタトランス','65016'),
 (r'信号線?用?.{0,3}サージ|サージ保護.{0,3}(信号|ユニット)','65017'),
 (r'リモコン(ト)?ランス|R[.\s]?TR','65000'),
]
def _remocon_code(name):
    s=unicodedata.normalize('NFKC', str(name))   # 半角カナ→全角に正規化(ﾀｲﾏ→タイマ)
    if ('リモコン' in s and 'リレー' in s) or re.search(r'R[.\s]?RY|リモコンリレ', s, re.I):
        # 既定2P(65002)。実見積書(城山)でR.RYは2P。明示1P時のみ65001。
        return ('65001' if re.search(r'1\s*P(?![.\d])', s) else '65002'),'○','リモコンリレー(65系・既定2P)'
    for pat,code in _REMOCON_MAP:
        if re.search(pat, s, re.I) and code in byCode:
            return code,'○','リモコン設備(65系)'
    return None

# 社内プロのフィードバック(△の実選定)由来の名称→コード直引き。実案件で人が確定したもの。
# (東部/八戸/木村/尼崎の見積レビュー△回答。◯/◎は未回答のため対象外。)
_PRO_MAP=[
 (r'計器盤|指示計器盤',                         '42081','計器盤=マルチ指示計'),
 (r'(^|\s)DA(\s|$|\d|[0-9〜~])',                '42083','DA(デマンド計)=マルチ指示計'),
 (r'(?<![0-9])27R(?![A-Za-z])',                 '73000','27R=AUX-RY(プロ確定)'),
 (r'(^|\s)UV(\s|$)|不足電圧要素',               '46011','UV=UVR(静止型)'),
 # 電力計=IWM / 力率計=PFM(DBは略記名)。普通角/広角は案件で分かれる→○(要確認)で確認ゲートへ。
 # 「多機能電力計」(マルチ)と衝突しないよう能の直後は除外。「電力量計」(WHM)は別(量で不一致)。
 (r'(?<!能)電力計|(?<![A-Za-z])IWM(?![A-Za-z])', '42020','電力計=IWM(普通角/広角42030は要確認)'),
 (r'力率計|(?<![A-Za-z])PFM(?![A-Za-z])',        '42022','力率計=PFM(普通角/広角42032は要確認)'),
 (r'計器\s*[\(（]\s*VM\s*/?\s*AM',               '42001','計器(VM/AM)=VM(普通角)・AM等は要確認'),
 (r'(?<![0-9])84\s*電圧|電圧継電器',              '46011','電圧継電器=UVR(静止型)・変種要確認'),
 # RPR(逆電力継電器): 図面「RPR」からは変種(単独46385/太陽光用複合OVGR･RPR46361/協調)が確定できない。
 # 唯一コードで◎になっていたが実案件で単独46385は不使用(六本木は太陽光用46361)→○(要確認)で確認ゲートへ。
 (r'(?<![A-Za-z])RPR(?![A-Za-z])|逆電力継電',    '46385','RPR(逆電力)・変種(単独/太陽光用複合46361/協調)は要確認'),
 (r'EL\s*漏電継電器|漏電継電器(?!.*ZCT)',        '46401','EL漏電継電器=LG-RY'),
 (r'接地端子(?!盤)',                            '62901','接地端子=接地端子盤1P'),
 (r'24H停電補償|停電補償付.*TM|TM[×xX]\d.*停電',  '73202','24H停電補償=TM-SW'),
 (r'中央監視盤|機械設備中央監視',                 '56900','中央監視盤=総合盤BOX'),
 (r'(?<![A-Za-z])VT([×xX]\d+)?(?![A-Za-z])',    '44011','VT=6KV 50VA(コイルモールド)'),
 # IGR絶縁監視(実見積書で全低圧盤に計上): 電圧発生器→46210 / 絶縁監視装置(1回路)→46311。
 # ※電圧発生器を先に判定(装置パターンに吸われないように)。ZCTは別(ZCT処理へ委ねる)。
 (r'絶縁監視.{0,6}電圧発生|(絶縁監視|IGR).{0,6}発生器|電圧発生器',   '46210','IGR絶縁監視電圧発生器'),
 (r'絶縁監視装置|(^|\s|絶縁監視)IGR[-\s]?500(?!.*ZCT)',            '46311','IGR絶縁監視装置(1回路用)'),
 # 分電盤のリモコン系部品(実見積書で計上)。略称は誤検出防止のため完全一致で。変種は要確認○。
 (r'^\s*(RT|R[.\-]?TR|リモコントランス|ﾘﾓｺﾝﾄﾗﾝｽ)\s*$|リモコン\s*T[rR]|ﾘﾓｺﾝ\s*T[rR]',  '65000','RT=リモコントランス'),
 (r'^\s*(Ry|RY|R[.\-]?RY|リモコンリレー|ﾘﾓｺﾝﾘﾚ-?)\s*(2P)?\s*$',   '65002','Ry=リモコンリレー2P'),
 (r'接点入力\s*T\s*/?\s*U',                                     '65040','接点入力T/U'),
 (r'リレー制御用\s*T\s*/?\s*U|ﾘﾚ-?制御用T/?U|^\s*T\s*/?\s*U\s*(\(?\d回路\)?)?\s*$', '65025','T/U=リレー制御用(4回路)'),
 (r'プログラムタイマ|ﾌﾟﾛｸﾞﾗﾑﾀｲﾏ|ﾌﾟﾛｸﾞﾗﾑﾕﾆｯﾄ',                 '65019','プログラムタイマユニット'),
 # THサーマルリレー(電動機過負荷)→TH-RY 73300(実見積書 六本木/城山で一致)。個別部品方式の制御盤前提。
 (r'(TH[-\s]?RY|THｻｰﾏﾙ|TH\s*ｻｰﾏﾙ|TH\s*サーマル|ｻｰﾏﾙﾘﾚ|サーマルリレー|熱動継電器)', '73300','THサーマルリレー'),
 # ELR漏電警報器(集合形・回路数不明は5回路を既定で要確認)。
 (r'(?<![A-Za-z])ELR(?![A-Za-z])|漏電警報器',                    '46410','ELR漏電警報器(集合形5回路・回路数要確認)'),
 (r'(?<![A-Za-z])TC(?![A-Za-z]).{0,6}トリップコイル|トリップコイル|ﾄﾘｯﾌﾟｺｲﾙ', '43390','TC=LBSトリップコイル'),
 # 盤付属品(負荷でなく盤内の標準小物)。名称が一意=負荷の換気扇(FE/FS 3φkW)等と衝突しないもののみ。
 # サーモスタット=盤内温度調節/警報用電源装置=警報回路電源。実見積書で全案件に計上され、従来は△行き止まりだった。
 (r'サ[ー-]?モスタット|ｻｰﾓｽﾀｯﾄ',                                  '74107','サーモスタット(盤付属)'),
 (r'警報用電源装置|警報用\s*電源|警報電源装置',                       '74079','警報用電源装置(CP付)'),
]
def _vmc_code(name):
    """VMC(真空電磁接触器)→VCS(43系)。御社の実見積書で確定した規則:
       ・VMC は常に VCS(43系)で計上(VMC→VCS は社内慣行で確定=◎)。
       ・PF/引出(E)/万能ヒューズ/カウンター付 → 電磁引出PF付(200A=43103/400A=43113)。
       ・変種語なし(素の電磁) → 電磁(200A=43101/400A=43111)。
       実績: 尼崎/西新宿=43101, 表参道/六本木/城山=43103, 木村/尼崎(万能ヒューズ・カウンター)=人手43103。"""
    s=unicodedata.normalize('NFKC',str(name))
    _vmc = bool(re.search(r'(?<![A-Za-z])VMC(?![A-Za-z])', s, re.I))
    _vcs = bool(re.search(r'(?<![A-Za-z])VCS(?![A-Za-z])', s, re.I))
    # PF/引出/万能ヒューズ/カウンター/(電磁,E) を PF付引出の指標とする
    pf = bool(re.search(r'PF|万能ヒュ|ｶｳﾝﾀ|カウンタ|引出|引き出し|電磁[,，、]\s*[EＥ]', s))
    # VMCは常にVCS解釈。素のVCSは通常照合に任せるが、VCS＋PF指標は変種確定のためここで拾う。
    if not (_vmc or (_vcs and pf)): return None
    amp400 = bool(re.search(r'400\s*A', s))
    if pf:
        code = '43113' if amp400 else '43103'
    else:
        code = '43111' if amp400 else '43101'
    if code not in byCode:  # DB欠番時は200A電磁へフォールバック
        code = '43101' if '43101' in byCode else None
    if not code: return None
    _pfx = 'VMC→VCS' if _vmc else 'VCS変種確定'
    # VCS素(43101)/電磁引出PF付(43103)の別は図面から確実に読めない(実測:西新宿は"PF(G)20A"表記でも
    # 正解43101=素、城山/表参道/六本木は素表記でも正解43103=引出PF付)。PF/引出の文字は回路のヒューズ定格を
    # 指すことがあり変種の確定根拠にならない。誤ると◎誤答なので常に○(要確認)＋両候補提示で確認ゲートへ委ねる。
    _base = '43111' if amp400 else '43101'
    _var  = '43113' if amp400 else '43103'
    _cands=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in (_base,_var) if c in byCode]
    return code,'○','%s(変種=素/PF付引出を確認ゲートで確定・既定=%s)'%(_pfx, byCode.get(code,{}).get('name','')[:16]), _cands
def _whm_code(name):
    """WHM(電力量計・電子式コンパクト)→70系。相(1φ2W/Nφ3W)×容量(30A/120A/N-5A/250A)×検定(検付/未検/
    通信付)で選定。N/5A(CT動作型)はコード表p13/41の注記でCTを別途拾う(呼出側でco-selection)。
    取引用/貸与の電力量計は別処理で対象外除外済のためここは計量用のWHMのみ。"""
    s=unicodedata.normalize('NFKC',str(name))
    # 遮断器行(MCCB/MCB/ELB…)に付く「Wh」は計量注記でありWHM本体ではない→除外(誤爆防止)。
    if re.search(r'MCCB|MCB|ELB|ELCB|LBS|VCB|VCS|(?<![A-Za-z])AF(?![A-Za-z])|\d+\s*AF|\d+\s*AT', s, re.I): return None
    # WHM本体は名称先頭がWHM/電力量計/Wh(単独)であること。数値注記中のwhには反応しない。
    if not (re.match(r'\s*(WHM|電力量計)', s, re.I) or re.search(r'(^|\s)Wh(\s|$)', s) or '電力量計' in s): return None
    if re.search(r'取引|貸与|課金', s): return None  # 貸与/取引は対象外
    is_n5=bool(re.search(r'N\s*/\s*5A|(?<!\d)/\s*5A', s))     # N/5A(CT動作)
    phase = 'N' if re.search(r'Nφ3W|N相|3\s*W|三相|3φ', s) else '1'  # Nφ3W or 1φ2W
    if is_n5: unit = 6 if phase=='N' else 3
    elif re.search(r'250\s*A', s): unit=7
    elif re.search(r'120\s*A', s): unit=5 if phase=='N' else 2
    elif re.search(r'30\s*A', s): unit=4 if phase=='N' else 1
    else: unit = 4 if phase=='N' else 1                       # 容量不明→30A既定(要確認)
    tens = 3 if re.search(r'通信.{0,3}(未検|無検)', s) else 2 if re.search(r'通信', s) \
           else 1 if re.search(r'未検|無検', s) else 0        # 既定=検付
    code='703%d%d'%(tens,unit)
    if code not in byCode:
        # 欠番(例:Nφ3W N/5A検付=70306)は未検→SP(7039x)→30A既定の順でフォールバック
        for alt in ['703%d%d'%(1,unit),'7039%d'%unit,'70301']:
            if alt in byCode: code=alt; break
    if code not in byCode: return None
    note='WHM %s %s%s'%('Nφ3W' if phase=='N' else '1φ2W','N/5A' if is_n5 else ('容量要確認' if unit in(1,4) and not re.search(r'30\s*A',s) else ''),'・検定/容量要確認')
    return code,'○',note.strip()
def _spd_code(name):
    """SPD本体: クラスI→74131(1P25KA)/クラスII→74134(3P20KA)標準を既定。4P明記→74136。
    SPD用分離器(74113/74123)・SPD用MCCBは別処理なので除外。67108(2種耐熱)等の特殊は明記時のみ。
    数量規則: クラスI(1Pのみ)=極数分, クラスII(3P/4P)=1個(呼出側/レビューで数量調整)。"""
    n=norm(name)
    if 'spd' not in n and '避雷' not in n: return None
    if re.search(r'用ﾋｭ|用ヒュ', n): return None  # SPD用ヒューズは対象外
    is_c1 = bool(re.search(r'ｸﾗｽ?\s*[i](?![i])|クラス\s*[iⅠ](?![iⅠ])|class[-\s]*1|(?<![a-z0-9])c[-\s]*1(?![0-9])|1種', n))
    # SPD用分離器(セパレータ)の単独行 → 分離器コードを直接返す(クラス別3P)。
    # ※本体+分離器が同一行(例「SPD CLASS-1 MCCB 3P225」)の場合は本体を返し、分離器はselect側で自動ペア計上。
    if re.search(r'分離|ｾﾊﾟﾚ|separat', n):
        if is_c1 and '74113' in byCode: return ('74113','○','SPD用分離器(クラスI)3P')
        if '74123' in byCode: return ('74123','○','SPD用分離器(クラスII)3P')
        return None
    if '2種耐熱' in n and '67108' in byCode: return ('67108','○','SPD(クラスII 2種耐熱)')
    if is_c1 and '74131' in byCode: return ('74131','○','SPD本体 クラスI(1P25KA・数量=極数分)')
    if re.search(r'4\s*p', n) and '74136' in byCode: return ('74136','○','SPD本体 クラスII(4P20KA)')
    if '74134' in byCode: return ('74134','○','SPD本体 クラスII(3P20KA)標準・既定')
    return None

def _pro_map(name):
    s=unicodedata.normalize('NFKC',str(name))
    v=_vmc_code(name)
    if v: return v
    sc=_scott_code(name)
    if sc: return sc
    wh=_whm_code(name)
    if wh: return wh
    # 高圧避雷器LA: 8.4kV 3P。標準/引出型/断路型/10KAの変種は図面から読めないことが多く、
    # 放電電流(2.5kA等)は変種選定に使えない→標準(43461)を○(変種要確認)で確定し△を解消。
    if re.search(r'(?<![A-Za-z])LA(?![A-Za-z])|避雷器', s) and re.search(r'\d\.?\d*\s*kv', s, re.I):
        if re.search(r'10\s*kA', s, re.I) and '43464' in byCode: return '43464','○','LA避雷器8.4kV(10KA)'
        if '43461' in byCode: return '43461','○','LA避雷器8.4kV(標準)・変種(引出/断路/10KA)は要確認'
    sp=_spd_code(name)
    if sp: return sp
    # マルチ指示計(マルチメータ): 型式(42075-88)は単線図から確定困難→既定を○で出し型式は確認ゲート。
    # 行き止まり△を無くす(確定率向上)。候補はgen_candidatesが42xxxを提示。誤ると◎になるので○止め。
    if re.search(r'マルチ\s*(指示計|メ|ﾒ)', s) and not re.search(r't/d|ﾏﾙﾁt|マルチt', s, re.I):
        if '42083' in byCode: return '42083','○','マルチ指示計(型式を確認ゲートで最終確定・既定ME110GF系)'
    for pat,code,note in _PRO_MAP:
        if re.search(pat, s) and code in byCode:
            return code,'○',note+'(プロ確定)'
    return None

# 高圧LBS(43320系): 3P200A枠のみDB実在。PFヒューズ定格→G感度バンド(75A以下/100A/200A)＋
# オプション(PF無/AL/TC/電動/エネセーバ)でコード確定。バンドが読めればPF=30/50/75Aは同一(75以下)で◎。
# 「励磁突入電流抑制機能」=エネセーバ(省エネ)機能→エネセーバ系。定格が読めなければ安全側△(呼出側の従来処理へ)。
_LBS_MAP={
 ('75','PF'):'43320',('75','PFなし'):'43324',('75','AL'):'43321',('75','TC'):'43325',('75','電動'):'43326',('75','エネセーバ'):'43327',
 ('100','PF'):'43330',('100','PFなし'):'43334',('100','AL'):'43331',('100','エネセーバ'):'43337',
 ('200','エネセーバ'):'43347',
}
def _lbs_code(name, panel=''):
    n=unicodedata.normalize('NFKC',str(name))
    if not re.search(r'(?<![A-Za-z])LBS(?![A-Za-z])', n, re.I): return None
    # 3P200A枠以外(例:400A)はDBに無い→従来処理へ委ねる
    if re.search(r'(\d{3,4})\s*A', n) and not re.search(r'200\s*A', n): return None
    # G感度バンドは【明示的なG表記(G100A/感度100A/(G)100A)がある時のみ】数値確定。
    # PFヒューズ定格(PF T87A 等)は G感度の代理にならない(実物件城山: PF T87A でも G75A以下=43321)。
    # ∴ 明示Gが無ければ band は御社標準の「75A以下」を既定にする(八戸/城山とも真値75以下)。
    #   Noneにするとエネセーバ等の変種が拾えず取りこぼす(八戸43327)ため、必ず75既定へ。
    mg=re.search(r'(?:感度|[GＧ])\s*[=＝(（]?\s*(\d+)\s*A', n, re.I)
    gexplicit=bool(mg)
    pf=int(mg.group(1)) if mg else None
    if pf is None: band='75'          # 明示G無し→御社標準の75A以下を既定
    elif pf<=75: band='75'
    elif pf<=100: band='100'
    else: band='200'
    # オプション(優先度高い順)。variant_explicit=図面に変種を示す語があったか(無ければPF/ALの区別不能)。
    variant_explicit=True
    if re.search(r'エネセ[ー-]?バ|エネルギ[ー-]?セ[ー-]?バ|励磁突入|突入電流抑制|インラッシュ', n): variant,vnote='エネセーバ','エネセーバ(励磁突入抑制)'
    elif re.search(r'電動', n): variant,vnote='電動','電動操作'
    elif re.search(r'(?<![A-Za-z])TC(?![A-Za-z])|トリップコイル', n, re.I): variant,vnote='TC','トリップコイル'
    elif re.search(r'(?<![A-Za-z])AL(?![A-Za-z])|アラ[ー-]?ム|溶断接点|溶断表示|ヒュ[ー-]?ズ溶断|溶断ﾋｭ', n, re.I): variant,vnote='AL','アラーム接点(溶断表示)'
    elif re.search(r'PF\s*無|ヒュ[ー-]?ズ\s*無|無ヒュ|ﾋｭ-ｽﾞ無', n): variant,vnote='PFなし','PF無し'
    else: variant,vnote,variant_explicit='PF','PF付',False
    # 変種語が図面に無い場合、PF付/AL付は盤種で決まる(実見積書8面照合で確立):
    #   低圧/一般 電灯盤・動力盤(TR二次側配電盤) → AL付(43321系)が定番【八戸/東部/城山/表参道/西新宿=5/5】
    #   高圧コンデンサ盤・饋電盤 → 素PF(43320系)【八戸コンデンサ/船引饋電=2/2】
    #   受電盤・変圧器盤・その他 → PF/AL判別不能→△(迷ったら安全側)
    pn=norm(panel)
    ambiguous=False
    if not variant_explicit:
        if any(k in pn for k in ['低圧','一般']) and ('電灯' in pn or '動力' in pn):
            variant,vnote='AL','アラーム接点(低圧配電盤の定番)'
        elif 'コンデンサ' in pn or '饋電' in pn or 'き電' in pn:
            variant,vnote='PF','PF付(饋電/コンデンサ側)'
        else:
            # 受電盤/変圧器盤等のTR側LBSは、変種語が無くてもAL付(43321)が御社定番(実測:西新宿の
            # 受電盤・変圧器盤LBS正解=43321)。素PF(43320)はコンデンサ/饋電側のみ。ambiguousで○止め+PF候補提示。
            variant,vnote,ambiguous='AL','アラーム接点(受電/変圧器盤=TR側LBSの定番・最善推定)',True
    if band is None:
        # G感度バンド不明。盤種でAL/PFが決まる場合は基本形を△で提示、そうでなければ従来処理へ。
        base='43321' if (variant=='AL' and '43321' in byCode) else ('43320' if '43320' in byCode else None)
        if base and not ambiguous:
            return base,'○','高圧LBS 3P200A %s(G感度定格は要確認)'%vnote
        if variant=='PF' and '43320' in byCode:
            return '43320','○','高圧LBS 3P200A PF付(最善推定・G感度/オプション要確認)'
        return None   # 変種明示ありでバンド不明→従来処理
    code=_LBS_MAP.get((band,variant)) or _LBS_MAP.get((band,'PF'))
    if not code or code not in byCode: return None
    gtxt='75A以下' if band=='75' else band+'A'
    # PF/AL判別不能(受電盤等・変種語なし)→最善推定の基本形を○で提示(行き止まり△を無くす・要確認)。
    if ambiguous:
        _alt=_LBS_MAP.get((band,'PF')) if variant=='AL' else _LBS_MAP.get((band,'AL'))
        _cl=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in (code,_alt) if c and c in byCode]
        return code,'○','高圧LBS 3P200A G%s %s(AL/PF/TC/電動等のオプションは確認ゲート)'%(gtxt,vnote), _cl
    # ◎は【明示的なG感度表記があり】かつ変種も確定した時のみ。G感度が既定(75)の場合や
    # エネセーバ(名称解釈)は○(要確認)。既定75は多数派だが誤ると◎誤答になるため安全側。
    if not gexplicit:
        return code,'○','高圧LBS 3P200A G%s(既定・G感度要確認) %s'%(gtxt, vnote)
    conf='○' if variant=='エネセーバ' else '◎'
    return code,conf,'高圧LBS 3P200A G%s %s'%(gtxt, vnote)

# AC/DCリアクトル(52系): INV(インバータ)分岐の付随品(支給品が多い)。容量kW→最近傍上位。
# 電圧は明記が無ければ既定200V(小容量INVは200Vが通例)。半角カナ(ﾘｱｸﾄﾙ)はNFKC正規化で吸収。
def _reactor_code(name):
    s=unicodedata.normalize('NFKC',str(name))
    m=re.search(r'(AC|DC)\s*リアクトル', s, re.I)
    if not m: return None
    kind=m.group(1).upper()
    volt='400' if re.search(r'400\s*V|3[φΦ]?\s*400', s) else '200'
    mk=re.search(r'(\d+\.?\d*)\s*KW', s, re.I)
    if not mk: return None
    want=float(mk.group(1))
    pool=[]
    for d in DB:
        mm=re.match(r'(AC|DC)リアクトル\s*(\d+)V\s*(\d+\.?\d*)KW', d['name'])
        if not mm or mm.group(1)!=kind or mm.group(2)!=volt: continue
        pool.append((float(mm.group(3)), d['code']))
    ge=sorted([(v,c) for v,c in pool if v>=want-1e-6])
    if not ge: return None
    v0,c0=ge[0]
    exact=abs(v0-want)<1e-6
    return c0,('◎' if exact else '○'),'%sリアクトル %sV %gKW%s'%(kind,volt,v0,'' if exact else '(仕様%g→%gKW繰上)'%(want,v0))

# スコットトランス(45050系・支給品): 3φ→1φ変換TR。kVA最近傍上位。高圧TR同様に購入品コードなし=支給品。
_CTRL_PART=[
 (r'FL-?10W.*(ドアSW|ﾄﾞｱ)', '71091'),(r'FL-?10W', '71093'),
 (r'(^|\s)PL\s*[\(（]?\s*R\s*[,、]\s*G\s*[,、]\s*0', '71057'),
 (r'(^|\s)PL\s*[\(（]?\s*R\s*[,、]\s*G', '71056'),
 (r'(^|\s)PL\s*[\(（]?\s*R(?![a-z])', '71054'),
 (r'(^|\s)PL\s*[\(（]?\s*0(?![a-z])', '71053'),
 (r'(^|\s)PL(?![a-z])|表示灯|ﾊﾟｲﾛｯﾄ|パイロット', '71051'),
 (r'(^|\s)BZ(?![a-z])|ブザ|ﾌﾞｻﾞ', '74001'),
 (r'(^|\s)COS(?![a-z])', '71021'),(r'(^|\s)CS(?![a-z])', '71010'),
 (r'AUX-?RY.*高級', '73001'),(r'AUX-?RY', '73000'),
 (r'(^|\s)T-?RY(?![a-z])', '73220'),(r'(^|\s)27X(?![a-z])', '73020'),
 (r'TM-?SW.*年間', '73205'),(r'TM-?SW.*週間', '73203'),(r'TM-?SW', '73201'),
 (r'(^|\s)PBS.*(ON|OFF)', '71042'),(r'(^|\s)PBS.*PL付', '71043'),(r'(^|\s)PBS(?![a-z])', '71041'),
]
def _ctrl_part_code(name):
    """盤内の制御小物(PL/BZ/COS/CS/AUX-RY/T-RY/27X/TM-SW/FL-10W/PBS)を個別コードで拾う。
    (見積システムへの入力コード。板金/端子/配線のみ自動計算で対象外・茂泉様)。
    ※セット盤では set 側が内包→呼出側(select_from_extracted)で抑制する。"""
    s=unicodedata.normalize('NFKC',str(name))
    if re.search(r'MCB|MCCB|ELB|ELCB|LBS|VCB|VCS|MCTT|盤$|BOX|函|変圧器|(?<![A-Za-z])TR(?![A-Za-z])|kVA', s, re.I): return None
    # COSφ(力率)・各種計器は制御小物でない→除外(COS切替スイッチと誤マッチ防止)
    if re.search(r'cos\s*[φΦ]|力率|指示計|計器', s, re.I): return None
    for pat,code in _CTRL_PART:
        if re.search(pat, s, re.I) and code in byCode:
            note=byCode[code].get('name','')
            if code=='73000': return code,'○',note+'(仕様レベル要確認: 高級=73001)'
            return code,'○',note
    return None
def _scott_code(name):
    s=unicodedata.normalize('NFKC',str(name))
    if not re.search(r'スコット|ｽｺｯﾄ|scott|sctt', s, re.I): return None
    mk=re.search(r'(\d+\.?\d*)\s*KVA', s, re.I)
    if not mk: return None
    want=float(mk.group(1))
    pool=[(float(mm.group(1)), d['code']) for d in DB
          if (mm:=re.match(r'スコットTR\(支給品\)\s*(\d+\.?\d*)KVA', d['name']))]
    ge=sorted([(v,c) for v,c in pool if v>=want-1e-6])
    if not ge: return None
    v0,c0=ge[0]; ex=abs(v0-want)<1e-6
    return c0,('◎' if ex else '○'),'スコットTR(支給品) %gKVA%s'%(v0,'' if ex else '(仕様%g→%gKVA繰上)'%(want,v0))

# 限流ヒューズPF単体(43622-43631): G定格(G40A等)またはPF NNN A→最近傍上位。
# (LBS内蔵PFは_lbs_codeで処理済。ここは単体PF行。力率PF/PFM計器と誤検出しないよう接頭PFに限定)
def _pf_code(name):
    s=unicodedata.normalize('NFKC',str(name)).upper()
    if not (re.match(r'\s*PF(\s|G|\d|$)', s) or 'パワーヒューズ' in s or '限流ヒューズ' in s): return None
    if re.search(r'PFM|力率', s): return None
    m=re.search(r'G?\s*(\d+)\s*A(?![A-Za-z])', s)
    if not m: return None
    want=int(m.group(1))
    pool=[(int(mm.group(1)), d['code']) for d in DB if (mm:=re.match(r'PF\s*(\d+)A$', d['name']))]
    ge=sorted([(v,c) for v,c in pool if v>=want])
    if not ge: return None
    v0,c0=ge[0]
    # 単体PFは常に○(要確認)。図面のPF行はSC(コンデンサ)直列ヒューズ等で支給品SC組込=別計上しない
    # ケースが多く(実測:表参道 PF×3 G10A→SC2系統は見積書に単体計上なし)、◎にすると◎誤答となる。
    # 単体計上要否と定格は確認ゲートで確定。候補として当該PFコードを提示。
    _cd={'code':c0,'name':byCode.get(c0,{}).get('name',''),'volt':''}
    return c0,'○','限流ヒューズPF %dA%s(単体計上要否は確認ゲート)'%(v0,'' if v0==want else '(仕様%d→%dA繰上)'%(want,v0)), [_cd]

# 直列リアクトルSRの容量は、同一盤のコンデンサSCとペアで決まる(SR kvar = L% × SC kvar)。
# SR単独では容量表記が無く△になるため、盤内SCのkvarから算定して45系SRコードを確定する。
def _sr_from_sc(name, sc_kvar):
    s=str(name)
    if not re.search(r'(?<![A-Za-z])SR(?![A-Za-z])|直列ﾘｱｸﾄﾙ|直列リアクトル', s): return None
    ml=re.search(r'L\s*[=＝]?\s*(\d+)\s*%', s) or re.search(r'(?<![\d.])(\d+)\s*%', s)  # 「6%」表記(Lなし)も拾う
    if not ml or not sc_kvar: return None
    Lpct=int(ml.group(1)); want=Lpct/100.0*sc_kvar
    pool=[(float(mm.group(2)), d['code']) for d in DB
          if (mm:=re.match(r'SR\(支給品\)\s*L=(\d+)%\s*([\d.]+)KVAR', d['name'])) and int(mm.group(1))==Lpct]
    ge=sorted([(v,c) for v,c in pool if v>=want-1e-6])
    if not ge: return None
    v0,c0=ge[0]
    return c0,'○','SR(支給品) L=%d%% %gkvar(SC %gkvar×%d%%)'%(Lpct,v0,sc_kvar,Lpct)

# 受変電の保護継電器(名称で一意に決まるもの)の直引き。候補多で△に落ちるのを救済。
# 容量変種があるもの(EL-RY等)は名称だけで断定しないので対象にしない(generic選定へ委ねる)。
def _relay_code(name, panel=''):
    n=unicodedata.normalize('NFKC',str(name)).upper(); s=str(name); pn=str(panel)
    引出='引出' in s or 'ﾋｷﾀﾞｼ' in s
    # DGR 地絡方向継電器
    if re.search(r'(?<![A-Z])DGR(?![A-Z])', n) or '地絡方向' in s:
        if 引出 and '46040' in byCode: return '46040','○','DGR(引出型)'
        if '46041' in byCode: return '46041','○','DGR(方向性)'
    # OCR 過電流継電器(静止型が既定・引出型は46000)。OCR51は46387。
    if re.search(r'(?<![A-Z])OCR(?![A-Z])', n) or '過電流継電' in s:
        if re.search(r'(?<![0-9])51(?![0-9])', n) and '46387' in byCode: return '46387','○','OCR 51'
        if 引出 and '46000' in byCode: return '46000','○','OCR(引出型)'
        if '46001' in byCode: return '46001','○','OCR(静止型)'
    # UVR/27 不足電圧継電器(静止型が既定・引出型は46010・27Hは46382)
    if re.search(r'(?<![A-Z])UVR(?![A-Z])|(?<![0-9])27H(?![0-9])', n) or '不足電圧継電' in s \
       or re.search(r'(?<![0-9])27(?![0-9RA-Z])', n):
        if '27H' in n and '46382' in byCode: return '46382','○','UVR 27H'
        if 引出 and '46010' in byCode: return '46010','○','UVR(引出型)'
        if '46011' in byCode: return '46011','○','UVR(静止型)'
    # OVGR 地絡過電圧(RPR併記は46361, 64Gは46381, 単独46360)
    if re.search(r'(?<![A-Z])OVGR(?![A-Z])|(?<![0-9])64G(?![0-9])', n) or '地絡過電圧' in s:
        if re.search(r'(?<![A-Z])RPR(?![A-Z])|逆電力', n) and '46361' in byCode: return '46361','○','OVGR・RPR'
        if '64G' in n and '46381' in byCode: return '46381','○','OVGR 64G'
        if '46360' in byCode: return '46360','○','OVGR(地絡過電圧)'
    # RPR 逆電力継電器(67P)
    if re.search(r'(?<![A-Z])RPR(?![A-Z])|逆電力継電|(?<![0-9])67P(?![0-9])', n):
        if '46385' in byCode: return '46385','◎','RPR(逆電力 67P)'
    # GR 地絡継電器(無方向・ZCT/LG/DGRでない)。
    if re.search(r'(?<![A-Z])GR(?![A-Z])', n) and not re.search(r'DGR|LGR|LG-?RY|ZCT|IGR|OVGR', n) \
       and not re.search(r'方向', s):
        # 高圧受電/饋電盤は高圧GR無方向=46031。低圧の変圧器盤/電灯盤/動力盤ではLG-RY(46401)相当。
        if re.search(r'受電|饋電|き電|高圧', pn) and '46031' in byCode: return '46031','○','GR(無方向性)'
        if re.search(r'変圧器盤|電灯盤|動力盤|低圧|スコット|ｽｺｯﾄ', pn) and '46401' in byCode: return '46401','○','GR=LG-RY(低圧地絡・ZCT付)'
        if '46031' in byCode: return '46031','○','GR(無方向性)'
    # LGR/LG-RY 地絡継電器(ZCT付)。ZCT併記かつ低圧アンペア指定が無いものだけ(46401=汎用)。
    if (re.search(r'(?<![A-Z])LGR(?![A-Z])|LG-?RY', n) and re.search(r'ZCT', n)) \
       or (re.search(r'地絡継電', s) and '方向' not in s and 'ZCT' in n):
        if not re.search(r'\d{2,4}\s*A(?![A-Za-z])', n) and '46401' in byCode:
            return '46401','○','LG-RY(ZCT付)'
    return None

# 統合: 1機器を選定
def select_one(name, panel='', prev_is_main=False, volt='', symbol='', kw='', group='', legend=None, breaker=''):
    # 「E3P50/20」等の "E"＋極数 は ELB(漏電遮断器)の略記(動力制御盤の負荷分岐に多い)。
    # ELB略記を明示化して B)MCB と誤認しないようにする(手本: 六本木E3P50→B)ELB 51536)。
    if isinstance(name,str):
        name=re.sub(r'(?<![A-Za-zＡ-Ｚ])[EＥ](?=\s*[２-４2-4]\s*[PＰ])', 'ELB', name)
    # リモコン設備(65系)は名称直引き(数値属性が無く候補生成に乗らないため)
    _rc=_remocon_code(name)
    if _rc: return R(_rc[0],_rc[1],_rc[2])
    # 社内プロのフィードバック由来の直引き(計器盤/DA/27R/UV/EL/接地端子/TM/監視盤/VMC万能ヒューズ)
    _pm=_pro_map(name)
    if _pm:
        _r=R(_pm[0],_pm[1],_pm[2])
        if len(_pm)>3 and _pm[3]: _r['candidates']=_pm[3]
        return _r
    # 盤内制御小物(PL/BZ/COS/CS/AUX-RY/T-RY/27X/TM-SW/FL-10W/PBS)を個別コードで拾う。
    # (見積システムへの入力コード。セット盤ではset側が内包→呼出側で抑制)
    _cp=_ctrl_part_code(name)
    if _cp: return R(_cp[0],_cp[1],_cp[2])
    # 受変電の保護継電器(DGR方向性/LG-RY ZCT付/OCR/UVR/OVGR/GR)の直引き
    _ry=_relay_code(name, panel)
    if _ry: return R(_ry[0],_ry[1],_ry[2])
    _ns=unicodedata.normalize('NFKC',str(name))
    # MCTT(電源切替器): 実見積書で盤種により系統が分かれる(茂泉様確定):
    #  ①配電盤(非常・保安の動力盤/電灯盤=受変電低圧配電盤) → 47系(3P-DT:47000系 / 3P-ST:47020系)。
    #     47系内は「動力盤の最大容量1台=DT(主電源切替)、他=ST」→ ST既定で出し、盤内最大をpost-passでDT昇格。
    #  ②制御盤(M/P番号)・分電盤(L番号) → 64系(交流3P:64001系)。
    #  ③ラッチ/LT明記 → 64系(交流ラッチ3P:64101系)。容量は最近傍上位(切上)、極数/容量不明は△。
    if re.search(r'MCTT|MC-?DT|MCDT', _ns) or (re.search(r'手動.{0,3}切替', _ns) and re.search(r'(?<![A-Za-z])DT(?![A-Za-z])', _ns)):
        is_latch=bool(re.search(r'ラッチ|ﾗｯﾁ|(?<![A-Za-z])LT(?![A-Za-z])|latch', _ns, re.I))
        mp=re.search(r'([234])\s*P', _ns); ma=re.search(r'(\d+)\s*A[TF]?(?![A-Za-z])', _ns)
        pole=mp.group(1) if mp else '3'; amp=int(ma.group(1)) if ma else None
        _kd=_mctt_kind(panel)
        if amp:
            if is_latch:  # LT型=64系 交流ラッチ3P
                for a,c in [(20,'64101'),(30,'64102'),(50,'64104'),(60,'64105'),(80,'64106'),(100,'64107'),(150,'64108'),(200,'64109'),(300,'64110'),(400,'64111'),(600,'64112')]:
                    if amp<=a and c in byCode: return R(c,('◎' if amp==a else '○'),'MCTT 交流ラッチ(LT)3P %dA%s'%(a,'' if amp==a else '(容量繰上)'))
            elif _kd in ('ctrl','bunden'):  # 制御盤/分電盤 → 64系 交流3P
                for a,c in [(20,'64001'),(30,'64002'),(50,'64004'),(60,'64005'),(80,'64006'),(100,'64007'),(150,'64008'),(200,'64009'),(300,'64010'),(400,'64011'),(600,'64012')]:
                    if amp<=a and c in byCode: return R(c,('◎' if amp==a else '○'),'MCTT 交流3P %dA%s'%(a,'' if amp==a else '(容量繰上)'))
            elif pole=='4':  # 配電盤の4P(DT)
                for a,c in [(100,'47042'),(200,'47043'),(400,'47045'),(600,'47046'),(800,'47047')]:
                    if amp<=a and c in byCode: return R(c,('◎' if amp==a else '○'),'MCTT 4P-DT %dA%s'%(a,'' if amp==a else '(容量繰上)'))
            else:  # 配電盤(haiden) 3P。名称にDT明記(MC-DT等)→直接DT(47系ST-20)。無印は既定ST型(盤内最大をpost-passでDT昇格)。
                _st=[(30,'47020'),(60,'47021'),(100,'47022'),(200,'47023'),(300,'47024'),(400,'47025'),(600,'47026'),(800,'47027'),(1000,'47028'),(1200,'47029'),(1600,'47030'),(2000,'47031')]
                _explicit_dt=bool(re.search(r'(?<![a-z])dt(?![a-z])', _ns, re.I))  # MC-DT/○○DT等の明記
                for a,c in _st:
                    if amp<=a and c in byCode:
                        if _explicit_dt:
                            dtc=str(int(c)-20)  # DT=47系ST-20(47025→47005)
                            if dtc in byCode: return R(dtc,('◎' if amp==a else '○'),'MCTT 3P-DT %dA(DT明記)%s'%(a,'' if amp==a else '(容量繰上)'))
                        r=R(c,('◎' if amp==a else '○'),'MCTT 3P-ST %dA%s'%(a,'' if amp==a else '(容量繰上)'))
                        r['_mctt']={'amp':a,'st':c,'dt':str(int(c)-20)}
                        return r
        return R('47005','○','MCTT(電源切替器)・極数/容量は確認ゲートで確定(既定3P)')
    # 手動電源切替器DT(68系): 極数×容量。極数/容量が読めれば確定、読めなければ△(既定3P60A提示)。
    if re.search(r'手動.{0,2}切替|手動電源切替|切替器.*DT|DT.*切替|(?<![A-Za-z])DT(?![A-Za-z]).{0,6}(\d+\s*P|\d+\s*A)', _ns):
        mp=re.search(r'([234])\s*P', _ns); ma=re.search(r'(\d+)\s*A(?![A-Za-z])', _ns)
        pole=mp.group(1) if mp else None; amp=int(ma.group(1)) if ma else None
        _DT={('2','60'):'68721',('2','100'):'68722',('3','60'):'68731',('3','100'):'68732',('3','200'):'68733',('3','400'):'68734',
             ('4','60'):'68751',('4','100'):'68752',('4','200'):'68753',('4','400'):'68754'}
        if pole and amp:
            steps=[60,100,200,400]; astd=next((str(s) for s in steps if amp<=s), '400')
            c=_DT.get((pole,astd))
            if c: return R(c,('◎' if amp in (60,100,200,400) else '○'),'手動電源切替器DT %sP %sA'%(pole,astd))
        return R('68731','○','手動電源切替器DT・極数/容量は確認ゲートで確定(既定3P60A)')
    # THR: サーマルリレー(TH-RY)が主(実見積書で73300)。最善推定○(COSの可能性もあり要確認)。
    if re.fullmatch(r'\s*THR\s*', str(name), re.I):
        return R('73300','○','THR=サーマルリレー最善推定(COSの可能性あり要確認)')
    # スコットトランス(支給品45050系)
    _sk=_scott_code(name)
    if _sk: return R(_sk[0],_sk[1],_sk[2])
    # 限流ヒューズPF単体(43622-43631)
    _pf=_pf_code(name)
    if _pf:
        _rpf=R(_pf[0],_pf[1],_pf[2])
        if len(_pf)>3 and _pf[3]: _rpf['candidates']=_pf[3]
        return _rpf
    # 警報盤の函体(56000 BOX)。警報点(◯◯異常/接点)は別処理で除外するのでここは盤本体のみ。
    if re.fullmatch(r'\s*警報盤\s*', str(name)) and '56000' in byCode:
        # 警報盤の函体は寸法でコードが分かれる(実見積書=東部は56011/56012等のサイズ別BOX)。
        # 汎用56000で◎にすると◎誤答になるため○(サイズは確認ゲート)。サイズ別BOX候補を提示。
        _bx=[c for c in ('56000','56011','56012','56013','56014') if c in byCode]
        _r=R('56000','○','警報盤 函体(BOX)・寸法別コードは確認ゲート(既定=汎用56000)')
        _r['candidates']=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in _bx]
        return _r
    # AC/DCリアクトル(52系): INV分岐の付随品。容量kW→最近傍上位(電圧既定200V)
    _rk=_reactor_code(name)
    if _rk: return R(_rk[0],_rk[1],_rk[2])
    # 高圧LBS(43320系): PFヒューズ定格→G感度バンドで確定(PF=30/50/75Aは同一枠)
    _lb=_lbs_code(name, panel)
    if _lb:
        _rl=R(_lb[0],_lb[1],_lb[2])
        if len(_lb)>3 and _lb[3]: _rl['candidates']=_lb[3]
        return _rl
    # 動力盤: 主回路記号があれば記号方式を最優先。
    # ただし名称に明示のMCCB分岐仕様(NP＋NNNAF/MMAT)があれば、記号方式より分岐遮断器選定を優先
    # (記号A/Cが主回路パターンでなく負荷分類タグの図面があり、記号方式では解けないため)。
    # 明示ブレーカー仕様(NP＋NNNAF、または NP NNN/MMM=枠/トリップ)を持つ負荷は、記号が始動回路でも
    # 「その遮断器で個別に拾う」(当社セット=MMCB+MC+2E等 と 図面=単なる遮断器 が違う→個別。茂泉様確定)。
    # 負荷側がパッケージ機器(自前で制御)の場合、盤はその遮断器だけを持つ。
    _has_afat=bool(re.search(r'[1-4]\s*[pP].{0,5}\d+\s*AF', str(name)) or re.search(r'[1-4]\s*[pPＰ]\s*\d{2,4}\s*/\s*\d{2,4}', str(name)))
    # 記号付き(動力制御盤の分岐負荷)で明示ブレーカー付きの場合のみ、機器ID(PAC-01-04等の数字)が
    # ブレーカー解析(極数/容量)を汚染するので、ブレーカー仕様部分だけで個別選定する
    # (手本: 六本木のPAC室外機E3P225→B)ELB 3P225)。主幹/個別盤の名前は削らない(記号なしは対象外)。
    if _has_afat and symbol:
        _bkm=re.search(r'(?:ELB|ELCB|MCCB|MCB)\s*[234]\s*[PＰ]\s*\d{2,4}\s*(?:/\s*\d{2,4}|AF\s*/?\s*\d*|AF)', str(name), re.I)
        if _bkm: name=_bkm.group(0)
    if symbol and not _has_afat:
        shikyu=('支給' in str(name))
        kwv = kw or (_get_kw(norm(name)) or '')
        parts=select_power_symbol(symbol, kwv, volt or '200V', shikyu, legend, breaker)
        # 主部品(1つ目)を主選定とし、残りは候補/内訳として保持
        first=parts[0]
        code=first[0]; note=first[2]; qty=first[1]
        if code:   # 記号方式で解けた場合のみ確定。空なら通常選定へフォールバック
            # 記号→凡例→回路種別で22-29系セットに変換したが、これは凡例を信頼した変換で、当社セットの機器構成
            # (MMCB+MC+ACL+2E等)と図面に実描画された部品リストの一致を検証したわけではない。分岐をセット(22-29)で
            # 拾うか個別遮断器で拾うかは「セット構成 vs 図面構成」の比較で決まる確認ゲート事項(茂泉様確定)。
            # 実測:六本木は全分岐が個別(22系なし)だが凡例①"直入起動(INV)"→22043を◎誤答していた。よって○止め。
            conf = '○'
            sel=R(code,conf,f'[動力記号{symbol}] '+note+' / 分岐をセット(22-29)/個別のどちらで拾うかは図面部品構成と照合し確認ゲートで確定')
            sel['parts']=[{'code':c,'qty':q,'note':nt,'name':byCode.get(c,{}).get('name','') if c else ''} for c,q,nt in parts]
            sel['set_qty']=qty
            return sel
    meta,cands=gen_candidates(name,volt,panel)
    # 高圧コンデンサの略号「C」(C3φ/C 3φ/SC C3φ)は SC の別名だが device検出が崩れ main=None になる。
    # kvar表記＋(コンデンサ盤 or C略号)なら SC として扱い、45系(支給品)選定へ載せる。
    if not meta.get('main') and re.search(r'kvar', str(name), re.I) \
       and (re.search(r'コンデンサ', str(panel)+str(name)) or re.search(r'(^|\s)C\s?\d?\s*[φΦ]', str(name))):
        meta['main']='SC'
    sel=refine(meta,cands,name,panel,prev_is_main,volt)
    sel['candidates']=[{'code':c['code'],'name':c['name'],'volt':c['volt']} for c in cands[:5]]
    # SC/SR容量の丸め(切下/切上)は確認ゲート: 前後の公称値を候補に提示(既定=切下)。
    if sel.get('_round_gate'):
        sel['candidates']=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in sel['_round_gate']]
        sel.pop('_round_gate',None)
    # TR(支給品・45系高圧変圧器)は◎にしない→○(要確認)。支給品は客先支給の変圧器で、単線図の容量表記は
    # 変圧器盤と電灯盤で食い違うことがあり(実測:六本木 T-L1=変圧器盤100kVA vs 電灯盤300kVA、正解300kVA/45008)、
    # 容量誤読が◎誤答に直結する。容量・支給/購入の別は確認ゲートで確定。前後容量ステップを候補提示。
    # ※SC/SR/スコットTRは容量がkvar/明示で安定して読めるため対象外(実測で正解一致)。
    _scd=sel.get('code','')
    if sel.get('conf')=='◎' and _scd and byCode.get(_scd,{}).get('name','').startswith('TR(支給品)'):
        _tm=re.match(r'TR\(支給品\)\s*(\S+)\s*([\d.]+)KVA', byCode[_scd]['name'])
        if _tm:
            _ph,_kv=_tm.group(1),float(_tm.group(2))
            _steps=sorted({float(m.group(2)) for d in DB
                           if (m:=re.match(r'TR\(支給品\)\s*(\S+)\s*([\d.]+)KVA', d['name'])) and m.group(1)==_ph})
            _i=_steps.index(_kv) if _kv in _steps else -1
            _near=[_kv]+([_steps[_i+1]] if 0<=_i<len(_steps)-1 else [])+([_steps[_i-1]] if _i>0 else [])
            _tc=[c for v in _near for d in DB if (m:=re.match(r'TR\(支給品\)\s*(\S+)\s*([\d.]+)KVA', d['name'])) and m.group(1)==_ph and float(m.group(2))==v for c in [d['code']]]
            sel['conf']='○'
            sel['note']='TR支給品(容量%s %gkVA)は確認ゲートで容量・支給区分を確定(単線図で容量表記が食い違う例あり) / %s'%(_ph,_kv,sel.get('note',''))
            sel['candidates']=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in _tc] or sel.get('candidates')
    # 動力/電灯/制御盤の3P分岐: AX付(補助接点付)は図面特記が無ければ盤種単位で要確認(既定=非AX)。
    _ag=_ax_gate(sel.get('code',''), name, panel)
    if _ag:
        axc, ptype = _ag
        if sel.get('conf')=='◎': sel['conf']='○'
        sel['note']='AX付の有無を盤種【%s】単位で要確認(図面特記なし・既定=非AX) / %s'%(ptype, sel.get('note',''))
        _cl=sel.get('candidates') or []
        _codes=[c.get('code') for c in _cl]
        _new=[{'code':sel['code'],'name':byCode.get(sel['code'],{}).get('name',''),'volt':''},
              {'code':axc,'name':byCode.get(axc,{}).get('name',''),'volt':''}]
        sel['candidates']=_new+[c for c in _cl if c.get('code') not in (sel['code'],axc)]
    # 保護セット: 付属(group有でリレー本体以外)は、親リレーが特定された文脈で
    # コードが一意に決まれば確定度を上げる（ZCT/CT/VT等が単独で△に落ちるのを救う）
    if group and cands:
        my=norm(meta.get('main_kw') or name)
        # 親リレー名と自分が異なる=付属。候補が容量等で1つに絞れていれば○へ
        if norm(group) not in my and sel['conf']=='△' and sel['code']:
            sel['conf']='○'; sel['note']=f'保護セット[{group}]の付属として確定: '+sel['note']
        sel['group']=group
    # --- 属性駆動チェック: 部品の必要属性が図面表記に揃っているか検査 ---
    # 不足があれば missing に記録。確信度◎なら○へ下げ要確認の材料にする。
    try:
        ml=(meta.get('main') or '').upper()
        spec=ATTR_TABLE.get(ml) or ATTR_TABLE.get(meta.get('main',''))
        if spec and sel.get('code'):
            nm_in=str(name)
            pats=spec.get('patterns',{})
            # 必須属性: 不足なら確信度を下げる。任意属性: 記録のみ。
            miss_req=[a for a in spec.get('required',[]) if pats.get(a) and not re.search(pats[a],nm_in,re.I)]
            miss_opt=[a for a in spec.get('optional',[]) if pats.get(a) and not re.search(pats[a],nm_in,re.I)]
            if miss_req:
                sel['missing_attrs']=miss_req
                if sel['conf']=='◎':
                    sel['conf']='○'
                    sel['note']=f"必須属性不足({'/'.join(miss_req)})・要確認: "+sel.get('note','')
            if miss_opt:
                sel['missing_opt']=miss_opt
    except Exception:
        pass
    return sel

def _amp_af(n):
    """AF(フレーム)とAT(トリップ)を抽出。AFを優先的に枠として使う。
    対応表記:
      ・「225af/200at」 → af=225, at=200 (明示)
      ・「mcb3p225/150」「3p225/150」 → 極数表記直後の 数字/数字 を AF/AT とみなす
      ・「mcb3p225」 → AF=225 のみ
    重要: norm()で空白が消えるため「50/30 1.5kW」→「50/3015kw」のように容量が
    連結する。AF/ATは規格アンペア値しか取らないので、トリップ側は標準AT値に
    厳密一致する場合のみ採用し、容量・幹線番号の混入(3015,5044,1502等)を排除する。"""
    import re as _re
    # 標準アンペア値(AF/ATが取りうる値)。これ以外は容量/幹線番号の混入とみなす
    STD={'15','20','30','40','50','60','75','100','125','150','175','200','225',
         '250','300','350','400','500','600','700','800','1000','1200','1600',
         '2000','2500','3200'}
    def _clean_at(v):
        if not v: return None
        if v in STD: return v
        # 連結ゴミ: 先頭から標準値に一致する最長の接頭辞を試す(例 3015→30, 1502→150)
        for L in (4,3,2):
            if len(v)>=L and v[:L] in STD: return v[:L]
        return None
    # 1) 明示AF/ATが最優先
    af=_re.search(r'(\d+)af', n); at=_re.search(r'(\d+)at', n)
    af_v=af.group(1) if af else None
    at_v=at.group(1) if at else None
    if af_v:
        return af_v, _clean_at(at_v)
    # 2) 極数表記(2p/3p/4p)の直後の「AF(/AT)」。AFは標準値、ATも標準値のみ採用
    m=_re.search(r'[234]p\s*(\d{2,4})(?:\s*/\s*(\d{2,4}))?', n)
    if m:
        af_c=m.group(1)
        at_c=_clean_at(m.group(2)) or _clean_at(at_v)
        return af_c, at_c
    return af_v, _clean_at(at_v)

def _pole(n):
    import re as _re
    m=_re.search(r'(\d)p', n)
    return m.group(1) if m else '3'

# AF枠→コード末尾2桁前(代表AF: 50/100/225/400/600/800/1000...)
_AF_KEY={'50':'5','100':'1','225':'2','200':'2','400':'4','600':'6','800':'8','1000':'70','1200':'73','1600':'76'}

_WET_ELB=re.compile(r'ｺﾝｾﾝﾄ|コンセント|便座|洗浄|浴室|給湯|食洗|温水|水栓|ｳｫｼｭ|ウォシュ|ﾃﾞｨｽﾎﾟｰｻﾞ|ディスポーザ|洗濯|乾燥機|屋外|ﾍﾞﾗﾝﾀﾞ|ベランダ|ﾙｰﾌ|屋上', re.I)
def _compact_branch(name, panel):
    """分電盤(60系)の2P分岐は、実見積書4案件(西新宿/表参道/城山/六本木)で
    **全て コンパクト**(B)MCB(コンパクト)2P50AF=60012 / B)ELB(コンパクト)=60014 /
    スペース=60028)。非コンパクトの60522/60525は4案件とも0件。
    ∴ 分電盤の2P分岐(MCB/ELB明示)はコンパクトで確定する。枠は50AF(実見積書は全て2P50AF、
    100AF以上の2P分岐は無い。抽出の数値は容量VA/電圧の混入が多く枠表記として信頼しない)。
    真の空きスロット(遮断器種別が付かない=本関数に到達しない)のみ通常選定へ委ねる。"""
    n=norm(name); pn=norm(panel)
    if 'mcb' not in n and 'elb' not in n and 'mccb' not in n and 'elcb' not in n: return None  # 遮断器種別が必要
    if _pole(n) not in ('1','2',''):    # 単相分岐(1P/2P)が対象。3P/4P主幹・動力分岐は通常選定へ
        return None
    # 盤種: 制御盤(50系端子台)・受変電低圧配電盤(40系)以外=分電系(60系コンパクト)。
    #  ※norm はハイフン/括弧を除去するためL番号検出が外れる盤名(AC-GC(LG-201)/共用盤/専用盤等)がある。
    #    そこで「制御/受変電でなければ分電系」と広く判定(実見積書4案件で2P50AF分岐は全てコンパクト)。
    # 盤種は_panel_kindで統一判定(制御50/配電40はコンパクト対象外・分電60のみコンパクト)。
    # ※「自立型」は筐体種別でありL/P番号を優先(5L-1(自立型)は分電盤)。
    if _panel_kind(panel)!='bunden': return None
    # 3P/4P明示や大枠(100AF以上)は通常分岐(60系)へ委ねる。50AFのみコンパクト。
    if re.search(r'3\s*p|4\s*p', n): return None
    # VA(負荷容量)の数値はAF枠ではない。明示のAF枠(NNNAF)または枠/トリップ対(NNN/NNN)のみを枠とみなし、
    # 100AF以上の明示枠だけコンパクト対象外にする。「NNNVA MCB1P」等の負荷表記はコンパクト分岐(50AF)。
    _has_frame = bool(re.search(r'\d{2,4}\s*af', n) or re.search(r'\d{2,4}\s*/\s*\d{2,4}(?![\d.]|\s*va)', n))
    if _has_frame:
        af,_at=_amp_af(n)
        if af and af not in ('50','30','20'):   # 100AF以上の明示枠はコンパクト対象外
            try:
                if int(af) >= 100: return None
            except: pass
    # コンセント/温水洗浄便座/給湯/浴室等の湿式・接触注意負荷は漏電遮断器(ELB)が電気規定上必須。
    # 単線図の分岐開閉器列(1P/2P/ELB)は列が近接しVisionが2P⇔ELBを誤読しやすいため、負荷種別で補正する
    # (実見積書でも客室のコンセント・温水洗浄便座は全てELB=60014)。
    is_elb = ('elb' in n) or ('elcb' in n) or ('漏電' in n) or bool(_WET_ELB.search(name))
    # 空きスペース(遮断器なしの空きスロット)＝スペースコード。「予備実装」等の実装済スペアは通常コンパクト。
    is_space = bool(re.search(r'スペース|ｽﾍﾟｰｽ|空き|空棒|空回路', name)) and '実装' not in name
    if is_space:
        code = '60029' if is_elb else '60028'
        return (code,'○','コンパクト空きスペース(2P50AF・%s)要確認'%('ELB' if is_elb else 'MCB')) if code in byCode else None
    code = '60014' if is_elb else '60012'
    if code not in byCode: return None
    return code, '◎', 'コンパクト分岐(2P50AF・%s)実見積書4案件で確定'%('ELB' if is_elb else 'MCB')

# 図面種別ヒント(配電盤図=haiden/分電盤図=bunden/制御盤図=ctrl)。図面は種別ごとに分かれて
# 描かれる(茂泉様確定)ので「どの図面に載るか」が盤種の最確実な信号。select_from_extractedが
# 盤ごとにセットし、_panel_kindが名前判定より優先する。contextvarでスレッド安全に通す。
_DRAWING_KIND = contextvars.ContextVar('drawing_kind', default=None)

def _panel_kind(panel):
    """盤種別→端子台系統: 制御盤=ctrl(50系)/分電盤=bunden(60系)/受変電低圧配電盤=haiden(40系)。
    図面種別ヒント(_DRAWING_KIND)があればそれを最優先(図面は種別ごとに分かれて描かれるため確実)。
    無ければ盤名から判定。※_mcb_code/_lug_code で共用(挙動を一致させるため単一定義)。"""
    _hint=_DRAWING_KIND.get()
    _nb=_panel_kind_byname(panel, mark_default=True)
    # 名前に明確な信号(制御/分電/高圧/L・P番号/電灯盤等)があればそれが最優先。
    # 図面は必ずしも種別ごとに綺麗に分かれておらず(実例:西新宿の分電盤図に制御盤P/LPが混在し
    # 手本では制御系51系が正)、明確な名前信号を図面種別で上書きすると誤るため。
    # 図面種別ヒントは、名前で判定できない盤(_default)のフォールバックにのみ使う。
    if _nb=='_default':
        return _hint if _hint in ('haiden','bunden','ctrl') else 'bunden'
    return _nb

def _panel_kind_byname(panel, mark_default=False):
    """盤名だけから盤種別を判定。mark_default=Trueなら、名前で判定できず既定に落ちる場合に
    'bunden'でなく'_default'を返す(図面種別ヒントのフォールバック対象を識別するため)。
      ① 制御盤(制御/自立/M・P番号・動力制御盤) = 端子台付き50系
      ② 分電盤(分電/照明分電/L・J・S番号)     = 端子台なし60系
      ③ 無印の電灯盤・動力盤(一般/低圧/非常/保安) = 上流の低圧配電盤40系"""
    pn=norm(panel)
    # 「GL＋数字」は地下階の位置コード(G=地下/Ground Level, GL2=地下2階)であり盤種でない。
    # 分類前に除去して、"GL2"の"L"を分電盤のL番号と誤読しないようにする(茂泉様確定)。
    pn=re.sub(r'gl\d+', '', pn)
    # 型文字(P/M=制御, L/J/S=分電)の後に「-T1」等の端子台付き接尾辞(文字+数字)が付く形も許容する。
    _mp=re.search(r'(^|[^a-zａ-ｚ])[a-zａ-ｚ]?\d*[mｍpｐ][ｰ\-－]?(?:[a-zａ-ｚ][ｰ\-－]?)?\d', pn)   # 1M-1,1P-1,1P-T1,M1(制御盤)
    _lj=re.search(r'(^|[^a-zａ-ｚ])[a-zａ-ｚ]?\d*[lｌjｊsｓ][ｰ\-－]?(?:[a-zａ-ｚ][ｰ\-－]?)?\d', pn)  # 1L-1,1L-T1,1S-1(分電盤)
    if '制御' in pn: return 'ctrl'
    if '分電' in pn: return 'bunden'
    # 「LP」(LP+数字)=Lighting Panel。"P"は「Panel」で制御のPではない(茂泉様確定)。LP盤は
    # 分電盤+制御盤の一体盤で、区画(制御50/分電60)は負荷内容で決まる(照明→分電/ポンプ等モーター→制御)。
    # ∴ LP名は曖昧(_default)とし、select側が負荷内容で判定したヒントに委ねる。ヒント無しは分電を既定。
    if re.search(r'(^|[^a-zａ-ｚ])[a-zａ-ｚ]?\d*lp[ｰ\-－]?\d', pn) and '制御' not in pn:
        return '_default' if mark_default else 'bunden'
    # 高圧・受変電系は型文字判定(_mp/_lj)より先にhaidenへ。「高圧コンデンサ盤 SC-1」の"SC-1"を
    # 分電盤のS番号と誤読して60系(M)LUG誤付与)になるのを防ぐ。コンデンサ/SC/SR盤も含める。
    # 「電灯盤/動力盤」(複合語)は受変電低圧配電盤(haiden)。盤名の(GL2)等のL番号で_ljが誤発火する前に確定
    # (例「保安電灯盤(GL2)」の"L2"で分電盤誤分類を防ぐ)。分電盤(1L-1等)は"電灯盤/動力盤"を含まない。
    # 変圧器盤(受変電・配電盤図に載る)は低圧配電盤カテゴリー=haiden。"T-L1/T-M1"は
    # 「L1電灯/M1動力用の変圧器」の意で、分電盤/制御盤のL/M番号ではない(茂泉様確定)。
    if any(k in pn for k in ['配電','受電','受変電','高圧','コンデンサ','ｺﾝﾃﾞﾝｻ','ｷｭ-ﾋﾞｸﾙ','キュービクル','饋電','き電','スコット','ｽｺｯﾄ','電灯盤','動力盤','変圧器盤']): return 'haiden'
    if _mp and not ('電灯' in pn or '照明' in pn): return 'ctrl'
    if _lj: return 'bunden'
    # 「自立(型)」は筐体種別(自立/壁掛)であり盤種でない。L/P番号や制御/分電の後に判定し、
    # 番号も種別語も無い「自立盤」のみ制御盤とみなす(分電盤も自立型があるため優先しない)。
    if '自立' in pn: return 'ctrl'
    if '電灯' in pn or '動力' in pn or '照明' in pn: return 'haiden'
    return '_default' if mark_default else 'bunden'

def _infer_drawing_kind(panels):
    """1図面(=1ファイル)の盤リストから図面種別を推定。図面は種別ごとに分かれて描かれるので、
    載っている盤の名前判定の多数決で図面種別が決まる。参照シート(標準図/凡例/一覧)は除外。
    明確な多数派(2件以上かつ過半数)のみ採用し、混在や不明瞭ならNone(=盤ごとの名前判定に委ねる)。"""
    from collections import Counter
    c=Counter()
    for p in panels:
        pn=p.get('panel','') if isinstance(p,dict) else str(p)
        if not pn: continue
        if re.search(r'標準|パターン|凡例|一覧|参考|負荷|設備配線|ダクト|標準結線|接地端子', norm(pn)): continue
        c[_panel_kind_byname(pn)]+=1
    if not c: return None
    top,n=c.most_common(1)[0]
    total=sum(c.values())
    return top if (n>=2 and n/total>=0.6) else None

def _bunden_load_branch(name, panel):
    """分電盤(bunden)は「1負荷=1分岐回路」(社内方式・実見積書で確認)。遮断器記号が無い裸のVA負荷でも
    分電盤ならコンパクト分岐で確定。湿式・接触注意負荷(コンセント/洗濯/給湯/浴室/温水等)=ELB(60014)、
    乾式(電灯等)=MCB(60012)。3P/3φ明示や遮断器枠(AF)明示・盤見出し/合計/機器名は対象外(通常選定へ)。"""
    if _panel_kind(panel)!='bunden': return None
    n=norm(name)
    if not re.search(r'\d+\s*va', n): return None                          # VA負荷であること
    if re.search(r'合計|盤$|受電|変圧器|コンデンサ|ｺﾝﾃﾞﾝｻ', n): return None      # 盤見出し/機器でない
    if re.search(r'3\s*p|3φ|\d{2,4}\s*af', n): return None                 # 3P/枠明示は通常分岐へ
    if re.match(r'^\s*(予備|スペース|ｽﾍﾟｰｽ|sp)\b', n): return None            # 予備/空きは別処理
    # 非負荷(計器/継電器/SPD/変成器/盤付属品)は負荷回路でない→対象外(VA=計器負担等の誤読を防ぐ)
    if re.search(r'ﾏﾙﾁ|マルチ|計器|ﾒｰﾀ|メータ|指示計|継電器|ﾘﾚｰ|リレー|r[.]?ry|spd|避雷|換気扇|ﾋｰﾀ|ヒータ|ﾌﾞｻﾞ|ブザ|表示灯|(?<![a-z])pl(?![a-z])|(?<![a-z])bz(?![a-z])|(?<![a-z])aux|(?<![a-z])[cv]t(?![a-z])|zct', n): return None
    code='60014' if _WET_ELB.search(name) else '60012'
    return code if code in byCode else None

def _mcb_code(name, panel, meta):
    n=norm(name); pn=norm(panel)
    is_main = n.startswith('m)') or ('主幹' in name) or ('主開閉器' in name) or (meta.get('main','').startswith('M)'))
    is_elb = ('elb' in n) or (meta.get('main')=='ELB' or meta.get('main')=='M)ELB' or meta.get('main')=='B)ELB')
    af,at=_amp_af(n)
    if not af:
        import re as _re
        # まず「数字/数字」ペア(AF/AT)を探す。無ければ極数除去後の最初の数値群。
        mp=_re.search(r'(\d{2,4})\s*/\s*(\d{2,4})', n)
        if mp:
            af=mp.group(1); at=at or mp.group(2)
        else:
            m=_re.search(r'(\d{2,4})', n.replace('3p','').replace('2p','').replace('4p',''))
            af=m.group(1) if m else None
    if not af: return ''
    # AF枠の標準化: 標準枠(50/100/225/400/600/800)以外なら、ATを収容する最小枠へ
    STD_AF=[50,100,225,400,600,800,1000,1200,1600,2000,2500,3200]
    try:
        afi=int(af)
        if afi not in STD_AF:
            # ATは「AF以下」かつ「標準枠上限以内」の妥当値のみ採用。
            # 容量混入等で AT>AF や AT>3200 の異常値はAFを基準にする(暴走防止)。
            at_i=int(at) if at else None
            base_val = at_i if (at_i and at_i<=afi and at_i<=STD_AF[-1]) else afi
            af=str(next((s for s in STD_AF if s>=base_val), STD_AF[-1]))
    except: pass
    pole=_pole(n)
    # 盤種別判定: 制御盤=50系 / 分電盤=60系 / 配電盤(受変電の低圧電灯・動力盤)=40系
    # ※「電灯/動力」は配電盤(低圧○○盤)と分電盤(電灯分電盤/1L-1)の両方に出るので区別する:
    #   低圧+電灯/動力 or 配電/受電/高圧/キュービクル → 配電盤40系
    #   分電 or 1L-1/2L-2 形式 or 単独の電灯 → 分電盤60系  / 制御 → 制御盤50系
    # 盤種判定ルール(茂泉様確定・実見積書4案件で全一致):
    #  判別語は「分電/制御があるか」。無地の「電灯盤/動力盤」は配電盤スケルトン上の
    #  低圧配電盤(上流)＝40系。「分電盤/制御盤」名やL/M/P番号は下流の個別盤。
    #  ① 制御盤(制御/自立/M・P番号・動力制御盤) = 端子台付き50系
    #  ② 分電盤(分電/照明分電/L・J・S番号)     = 端子台なし60系
    #  ③ 無印の電灯盤・動力盤(一般/低圧/非常/保安) = 上流の低圧配電盤40系(AX付=41系)
    kind=_panel_kind(panel)
    # 主幹MCB 3P
    afmap_main={
      'ctrl':{'50':'50503','100':'50103','225':'50203','200':'50203','400':'50403','600':'50603','800':'50803'},
      # 分電盤主幹は欠相保護無(50系)を既定とする（会社確認: 欠相保護有なら60系）
      # 分電盤=60系(50系は制御盤・茂泉様確定)。M)MCB(ELB)3P50〜600Aは中性線欠相保護付が既定(コード表p33)。
      'bunden':{'50':'60503','100':'60103','225':'60203','200':'60203','400':'60403','600':'60603','800':'60803'},
      'haiden':{'50':'40503','100':'40103','225':'40203','200':'40203','400':'40403','600':'40603','800':'40803'},
    }
    # AX付(欠相・中性線欠相・補助接点付)は系統を+1(40→41/50→51/60→61)。図面に明記がある時のみ。
    # AX付＝欠相・中性線欠相・補助接点付。「警報付」も警報出力用の補助接点を要するためAX扱い
    # (実見積書=東部動力制御盤で図面『警報付』⟺見積書『AX付』が一致)。制御盤(ctrl)のみ適用し
    # 配電盤/分電盤の警報は別要素なので誤適用を避ける。
    is_ax = bool(re.search(r'AX付?|(?<![A-Za-z])AX(?![A-Za-z])|中欠|欠相|補助接点', name)) \
            or (kind=='ctrl' and bool(re.search(r'警報付|警報出力', name)))
    def _ax(code):
        if is_ax and len(code)==5 and code[1]=='0':
            axc=code[0]+'1'+code[2:]
            if axc in byCode: return axc
        return code
    if is_main and not is_elb and pole=='3':
        code=afmap_main.get(kind,{}).get(af,'')
        if code in byCode: return _ax(code)
    # 分岐MCB/ELB(B)系): 極数×AF×盤種別で実在コードを探す
    if not is_main:
        for cand in _branch_candidates(pole, af, is_elb, kind):
            if cand in byCode: return _ax(cand)
    return ''

# --- M)LUG(主幹端子)コード: DB実在コードから (系統,極数)→[(容量,コード)] を構築 ---
# LUG容量枠は連続でない(50/100/225/400/600/800/1000…)ため、名称を切上げでDB枠に丸める。
# コード体系(例 40609=M)LUG 3P600A)は系統2桁+容量2桁+極尾で不規則→DBを正として直引き。
_LUG_MAP={}
_KIND_SERIES={'ctrl':'50','bunden':'60','haiden':'40'}
def _build_lug_map():
    for c,v in byCode.items():
        m=re.match(r'\s*M\)LUG\s+(\d)P\s+(\d+)A', str(v.get('name','')))
        if m and len(c)==5:
            _LUG_MAP.setdefault((c[:2],m.group(1)),[]).append((int(m.group(2)),c))
    for k in _LUG_MAP: _LUG_MAP[k].sort()
_build_lug_map()

def _lug_code(name, panel, meta):
    """M)LUG(SPD等の分岐頭・断路スイッチのみでMCCB表記なし=端子受け)→系統(端子台)×容量(切上)。
    系統は_panel_kind(制御50/分電60/配電40)。容量は名称のA値をDBのLUG枠へ最近傍上位で丸める
    (例『M)LUG 3P200A』は200A枠が無く225A枠→xx209)。読めなければNone(安全側で呼出側が△)。"""
    n=norm(name)
    if not ('lug' in n or str(meta.get('main',''))=='M)LUG'): return None
    pole=_pole(n) or '3'
    af,at=_amp_af(n)
    amp=None
    for v in (af,at):
        if v:
            try: amp=int(v); break
            except: pass
    if not amp: return None
    series=_KIND_SERIES.get(_panel_kind(panel),'60')
    lst=_LUG_MAP.get((series,pole)) or _LUG_MAP.get((series,'3'))
    if not lst: return None
    for a,code in lst:
        if a>=amp: return code
    return lst[-1][1]

def _tr_lug_from_names(panel, names):
    """配電盤(受変電低圧40系・個別化時)のTR二次主幹端子M)LUG。TR二次は端子(LUG)受け。
    容量=TR二次電流(210V)切上: 1φ=kVA×1000/210, 3φ=÷√3, Scott=半容量1φ扱い, 発電機G=3φ扱い。
    手本(城山)で40609/40770/40409/40209を再現。返り値 code or None。names=盤内の元抽出品名。"""
    import math
    if _panel_kind(panel)!='haiden': return None
    pn=norm(panel)
    if not re.search(r'電灯|動力|低圧|変圧|scott|ｽｺｯﾄ|発電', pn+' '+' '.join(names).lower()): return None
    best=None
    for raw in names:
        n=norm(str(raw))
        if re.search(r'kvar|ﾘｱｸﾄﾙ|リアクトル|ｺﾝﾃﾞﾝｻ|コンデンサ|(^|[^a-z])s[rc](?![a-z])', n): continue
        mk=re.search(r'(\d+)\s*kva', n)
        if not mk: continue
        is_scott='scott' in n or 'ｽｺｯﾄ' in str(raw) or 'スコット' in str(raw)
        is_gen=bool(re.search(r'発電|generator', n)) or bool(re.search(r'(^|[^a-z])g\b', n))
        if not (re.search(r'(^|[^a-z])t\b|t\s*[:：]', n) or '変圧器' in str(raw) or is_scott or is_gen): continue
        kva=int(mk.group(1))
        mp=re.search(r'([13])\s*[φΦ相]', str(raw)) or re.search(r'([13])φ', n)
        phase='1' if is_scott else (mp.group(1) if mp else '3')
        k=kva/2 if is_scott else kva
        I=k*1000/210 if phase=='1' else k*1000/(math.sqrt(3)*210)
        code=_lug_code('M)LUG 3P %dA'%math.ceil(I), panel, {'main':'M)LUG'})
        if code and (best is None or kva>best[1]): best=(code, kva)
    return best[0] if best else None

def _main_elb_code(name, panel, meta):
    """主幹ELB(『ELB付 主幹』等)の最善推定コード。M)ELB=base+AF桁+極数尾(3P=06/4P=07/2P=05)。
    実見積書(表参道)ではこの主幹が M)MCB(EL・中欠)(AX)=61系01 になる例があるが、EL/中欠/AXは
    図面に明記されない社内標準のため◎にせず、素のM)ELBを○(最善推定・変種要確認)で返す。"""
    n=norm(name); pn=norm(panel)
    is_main = n.startswith('m)') or ('主幹' in name) or ('主開閉器' in name) or (str(meta.get('main','')).startswith('M)'))
    is_elb = ('elb' in n) or ('elcb' in n) or ('漏電' in n)
    if not (is_main and is_elb): return None
    af,_at=_amp_af(n)
    if not af:
        mp=re.search(r'(\d{2,4})\s*/\s*(\d{2,4})', n)
        af=mp.group(1) if mp else None
    if not af: return None
    STD_AF=[50,100,225,400,600,800]
    try:
        afi=int(af); af=str(next((s for s in STD_AF if s>=afi), 800))
    except: return None
    afdig={'50':'5','100':'1','225':'2','400':'4','600':'6','800':'8'}.get(af,'')
    if not afdig: return None
    pole=_pole(n) or '3'
    tail='07' if pole=='4' else ('05' if pole=='2' else '06')
    if '制御' in pn or '自立' in pn: bases=['50','60']
    elif any(k in pn for k in ['配電','受電','高圧','ｷｭ-ﾋﾞｸﾙ','キュービクル','饋電']): bases=['40','60','50']
    else: bases=['60','50']
    for base in bases:
        c=base+afdig+tail
        if c in byCode: return c
    return None

def _branch_candidates(pole, af, is_elb, kind):
    """極数・AF・盤種別から分岐コード候補を実在順に返す。
    末尾規則: 2P MCB=22/ELB=25, 3P MCB=33/ELB=36, 4P ELB=46。
    AF桁: 50→5,100→1,225→2,400→4,600→6,800→8。盤: 配電40/制御50/分電60。"""
    afdig={'50':'5','100':'1','225':'2','200':'2','400':'4','600':'6','800':'8','30':'3'}.get(af,'')
    if not afdig: return []
    # 盤種別の基番号を優先順で（分電盤=60系優先, 配電盤=40系優先, 制御盤=50系優先）
    if kind=='ctrl': bases=['50','40','60']
    elif kind=='haiden': bases=['40','60','50']
    else: bases=['60','40','50']   # bunden(分電盤)は60系優先
    out=[]
    for base in bases:
        if pole=='2':
            if is_elb: out.append(base+afdig+'25')   # 2P ELB 例:50525/60525
            else: out.append(base+afdig+'22')        # 2P MCB 例:40522/60522
        elif pole=='4':
            out.append(base+afdig+'46')              # 4P ELB
        else:  # 3P
            if is_elb: out.append(base+afdig+'36')   # 3P ELB 例:40536
            else: out.append(base+afdig+'33')        # 3P MCB 例:40533
    return out

def _mcb_note(name, panel):
    n=norm(name)
    role='主幹' if (n.startswith('m)') or '主幹' in name) else '分岐'
    typ='ELB' if 'elb' in n else 'MCB'
    return f'{role}{typ}(盤種別・AF枠で選定)'

def _mctt_kind(panel):
    """MCTTの盤種判定: 制御盤(M/P番号・制御/自立)/分電盤(L番号・分電)→'ctrl'/'bunden'(→64系)、
    それ以外の配電盤(非常・保安の動力盤/電灯盤=受変電低圧配電盤)→'haiden'(→47系)。"""
    pn=norm(panel)
    if '制御' in pn or '自立' in pn: return 'ctrl'
    if '分電' in pn: return 'bunden'
    # 無印/非常/保安/一般/低圧の「動力盤・電灯盤」は受変電低圧配電盤=haiden(47系)を優先。
    # (盤名に変圧器参照のM4等が括弧付きで混じっても、動力盤/電灯盤の語があれば配電盤とする)
    if re.search(r'動力盤|電灯盤', pn): return 'haiden'
    _mp=re.search(r'(^|[^a-zａ-ｚ])[a-zａ-ｚ]?\d*[mｍpｐ][ｰ\-－]?\d', pn)
    _lj=re.search(r'(^|[^a-zａ-ｚ])[a-zａ-ｚ]?\d*[lｌjｊsｓ][ｰ\-－]?\d', pn)
    if _mp and not ('電灯' in pn or '照明' in pn): return 'ctrl'
    if _lj: return 'bunden'
    return 'haiden'

def _ax_gate(code, name, panel):
    """AX付(補助接点付)は図面に特記が無いと確定できない社内標準仕様。実見積書(城山/表参道/六本木)では
    受変電低圧動力盤/電灯盤(40系)・動力制御盤(50系)・分電盤(60系)の3P主幹・分岐が一律AX付(41/51/61系)
    だが図面に明記なし(コード表p33: AX付は上2桁を61)。∴ 40/50/60系の3P主幹(M)・分岐(B)で図面にAX特記が
    無ければ、AX付は【盤種単位の確認ゲート】項目とし、非AXを既定表示しつつ○(要確認)＋AX付変種を候補提示する。
    戻り値: (AX付コード, 盤種ラベル) / None。"""
    if not code or len(code)!=5 or code[:2] not in ('40','50','60'): return None
    nm=byCode.get(code,{}).get('name','')
    if not re.search(r'^[BM]\)', nm) or not re.search(r'[23]\s*[PＰ]', nm): return None  # 2P/3P 主幹(M)/分岐(B)
    if code[1]!='0': return None                                   # 既にAX(41/51/61)なら対象外
    if re.search(r'AX|欠相|中欠|補助接点', str(name)): return None    # 図面にAX特記あり→確定
    axc=code[0]+'1'+code[2:]
    if axc not in byCode: return None
    pn=norm(panel)
    if code[0]=='6': ptype='分電盤'
    elif code[0]=='5': ptype='動力制御盤'
    elif '電灯' in pn or '照明' in pn: ptype='電灯盤'
    else: ptype='動力盤'
    return axc, ptype

# ===== 動力盤：主回路記号(A〜L)方式 =====
_LS_SUF={'2.2':'000','3.7':'001','5.5':'011','7.5':'021','11':'031','15':'041','18.5':'061','22':'071','30':'081','37':'091','45':'111','55':'121','75':'131'}
_SD_SUF={'7.5':'022','11':'032','15':'042','18.5':'062','22':'072','30':'082','37':'092','45':'112','55':'122','75':'132'}
_INV_SUF={'0.75':'993','1.5':'983','2.2':'053','3.7':'003','5.5':'013','7.5':'023','11':'033','15':'043','18.5':'063','22':'073','30':'083','37':'093','45':'113','55':'123','75':'133','90':'143'}

def _round_cap(kw, suf):
    caps=sorted([float(k) for k in suf], key=float)
    try: v=float(kw)
    except: return None
    for c in caps:
        if v<=c: return ('%g'%c)
    return ('%g'%caps[-1])

def select_power_symbol(symbol, kw, volt='200V', shikyu=False, legend=None, breaker=''):
    """主回路記号＋容量＋電圧 → [(code, qty, note),...]
    legend: 図面の主回路パターン凡例 {記号:回路種別}。あれば記号→種別のたすき掛けを最優先。
    breaker: ●(=MCCB)/○(=ELB) の別。ELBならAX付き変種を狙う。"""
    pfx='26' if volt=='400V' else '22'
    sym=(symbol or '').upper().strip()
    # --- 図面凡例があれば記号→回路種別で直接たすき掛け(図面ごとの割付差に対応) ---
    if legend and sym in {str(k).upper().strip() for k in legend}:
        raw=next(v for k,v in legend.items() if str(k).upper().strip()==sym)
        typ=_normalize_type(raw)
        if typ in ('L-S(AM付)','スターデルタ','INV','INV(スターデルタ)'):
            try: kwf=float(re.sub(r'[^\d.]','',str(kw)) or 0)
            except: kwf=0
            if kwf>0:
                dev=_dev_from_breaker(breaker)
                code,conf,pick=_bunki_find(typ,kwf,dev,volt)
                if code: return [(code,1,f'凡例{sym}→{typ} {pick}kW {volt}')]
                return [('',1,f'凡例{sym}→{typ} {kw}kW容量外→確認')]
        # typ=='MCCB'(電源/コンセント直接)や不明種別は下のハードコード/通常処理へ委ねる
    def ls():
        kk=_round_cap(kw,_LS_SUF); c=pfx+_LS_SUF.get(kk,'') if kk else ''
        return (c,kk) if c in byCode else ('',kk)
    def sd():
        kk=_round_cap(kw,_SD_SUF); c=pfx+_SD_SUF.get(kk,'') if kk else ''
        return (c,kk) if c in byCode else ('',kk)
    def inv():
        kk=_round_cap(kw,_INV_SUF)
        if not kk: return ('',None)
        s=_INV_SUF.get(kk,'')
        if shikyu and s: s=s[:-1]+'5'
        c=pfx+s
        return (c,kk) if c in byCode else ('',kk)
    if sym in ('A','B'):
        return [('',1,f'記号{sym}:電源供給のみ(モーター制御部なし)')]
    if sym=='C':
        c,kk=ls(); return [(c,1,f'直入L-S {kw}→{kk}kW枠')] if c else [('',1,f'L-S {kw}kW要確認')]
    if sym=='D':
        c,kk=sd(); return [(c,1,f'Y-Δ {kw}→{kk}kW枠')] if c else [('',1,f'Y-Δ {kw}kW要確認')]
    if sym in ('E','G'):
        c,kk=ls(); return [(c,2,f'直入L-S {kw}→{kk}kW ×2(記号{sym})')] if c else [('',2,f'L-S×2要確認')]
    if sym in ('F','H'):
        c,kk=sd(); return [(c,2,f'Y-Δ {kw}→{kk}kW ×2(記号{sym})')] if c else [('',2,f'Y-Δ×2要確認')]
    if sym=='I':
        c,kk=inv(); return [(c,1,f'INV {kw}→{kk}kW')] if c else [('',1,f'INV {kw}kW要確認')]
    if sym=='L':
        c,kk=inv(); return [(c,2,f'INV {kw}→{kk}kW ×2(二重化)')] if c else [('',2,f'INV×2要確認')]
    if sym in ('J','K'):
        ic,ik=inv(); bc,bk=(ls() if sym=='J' else sd())
        out=[]
        out.append((ic,1,f'INV {ik}kW') if ic else ('',1,'INV要確認'))
        out.append((bc,1,f'{"直入" if sym=="J" else "Y-Δ"}バイパス {bk}kW') if bc else ('',1,'バイパス要確認'))
        return out
    return [('',1,f'記号{sym}不明・要確認')]

# ===== 値ベース属性照合（値が種類を語る） =====
def attr_value_set(text, dbvolt=''):
    """テキストから属性値の集合を作る。値そのものが種類(極数/電圧/容量/AF/変流比)を表す。"""
    import re as _re
    n=unicodedata.normalize('NFKC',str(text))
    vals=set()
    for m in _re.findall(r'(\d)P(?![a-zA-Z0-9])', n): vals.add(m+'P')        # 極数(1桁)
    for m in _re.findall(r'(\d{2,3})P(?![a-zA-Z])', n): vals.add(m+'P')       # ポート数(2-3桁:20P/50P/100P)
    for m in _re.findall(r'(\d+)\s*AF', n): vals.add(m+'AF')               # フレーム
    for m in _re.findall(r'(\d+)\s*AT', n): vals.add(m+'AT')               # トリップ
    if _re.search(r'(MCB|ELB|LUG|MCCB)', n):                               # MCB系のA枠
        for m in _re.findall(r'(\d+)\s*A(?![a-zA-Z]|F|T)', n): vals.add(m+'AF')
    if _re.search(r'7\.2kV|6\.6kV|6kV', n, _re.I): vals.add('HV')
    if '415' in n or '440' in n: vals.add('400V')
    if _re.search(r'210|220', n): vals.add('200V')
    if _re.search(r'105|100V', n): vals.add('100V')
    if dbvolt:
        if 'kv' in dbvolt.lower(): vals.add('HV')
        for x in ('400','200','100'):
            if x in dbvolt: vals.add(x+'V')
    for m in _re.findall(r'(\d+\.?\d*)\s*kW', n, _re.I): vals.add(m+'kW')  # 容量
    for m in _re.findall(r'(\d+\.?\d*)\s*kvar', n, _re.I): vals.add(m+'kvar')
    for m in _re.findall(r'(\d+\.?\d*)\s*kVA', n): vals.add(m+'kVA')
    # 変流比（具体値）。範囲(100/5A~200/5A)はDB側で別途判定
    for m in _re.findall(r'(\d+)/5A', n): vals.add(m+'/5A')
    # 高圧機器の定格電流(A)：VCB/DS/LBS/PAS等。kV/kvar/VA/ATと紛れないものだけ
    for m in _re.findall(r'(?<![/\d])(\d{2,4})A(?![A-Za-z]|F|T)', n):
        vals.add(m+'A')
    # kA(遮断容量)・kV
    for m in _re.findall(r'(\d+\.?\d*)kA', n, _re.I): vals.add(m.lower().replace('.0','')+'kA')
    # 弱電の型式・分配数: 2D(2分配)/AMP/MIX 等
    for m in _re.findall(r'(\d)D(?![a-z])', n): vals.add(m+'D')
    if 'amp' in n.lower(): vals.add('AMP')
    if 'mix' in n.lower(): vals.add('MIX')
    # T/U(リモコン制御ユニット)の回路数・型式
    for m in _re.findall(r'(\d+)回路', n): vals.add(m+'回路')
    if '片切' in n: vals.add('片切')
    if '両切' in n: vals.add('両切')
    if '6a' in n.lower(): vals.add('6A')
    if '調光' in n: vals.add('調光')
    if '接点入力' in n: vals.add('接点入力')
    # 変圧器: 相数(1φ/3φ)とKVA容量
    for m in _re.findall(r'([13])[φΦ]', n): vals.add(m+'φ')
    for m in _re.findall(r'(\d+)\s*KVA', n, _re.I): vals.add(m+'KVA')
    # LBSのPFヒューズ定格: PF=30A / PF30A 等
    for m in _re.findall(r'PF[=]?(\d+)A', n, _re.I): vals.add('PF'+m+'A')
    # SPDのクラス: クラスI/II、保護レベルI/II(=クラス)
    if re.search(r'クラスi(?!i)|保護レベルi(?!i)|classi(?!i)', n, _re.I): vals.add('クラスi')
    if re.search(r'クラスii|保護レベルii|classii', n, _re.I): vals.add('クラスii')
    # KA容量(SPD): 25KA/20KA等は既にkAで拾うが大文字KA表記も
    for m in _re.findall(r'(\d+)KA', n): vals.add(m.lower()+'ka')
    return vals

def _ct_range_match(draw_vals, dbname):
    """高圧CTの範囲(20/5A~40/5A)に、図面の具体値(150/5A)が入るか"""
    import re as _re
    mr=_re.search(r'(\d+)/5A\s*[~〜\-]\s*(\d+)/5A', dbname)
    if not mr: return False
    lo,hi=int(mr.group(1)),int(mr.group(2))
    for v in draw_vals:
        m=_re.match(r'(\d+)/5A', v)
        if m and lo<=int(m.group(1))<=hi: return True
    return False

def value_score(draw_vals, d):
    """図面の属性値集合と、DBコードの属性値集合を突き合わせてスコア化"""
    cv=attr_value_set(d['name'], d.get('volt',''))
    if not draw_vals: return 0, cv
    inter=draw_vals & cv
    miss=draw_vals - cv
    score=len(inter)*10
    # CT範囲の特別扱い：図面の変流比がDB範囲に入れば加点（missから変流比を除外）
    ratio_miss=[v for v in miss if v.endswith('/5A')]
    if ratio_miss and _ct_range_match(draw_vals, d['name']):
        score+=10; miss=miss-set(ratio_miss)
    score-=len(miss)*8
    return score, cv

def option_bonus(opts, dbname, draw_name=''):
    """オプション語がDB品名に含まれれば加点。図面に無い特別仕様(引出/SP等)は減点し標準を優先。"""
    n=norm(dbname); dn=norm(draw_name); b=0
    for o in opts:
        on=norm(o)
        if on and on in n: b+=5
    # 特別仕様語: 図面に無いのにDB側が該当なら減点（標準/既定を優先）
    SPECIAL=['引出','スペース','(sp)','ｓｐ','電動','コージェネ','高性能','可逆']
    for sp in SPECIAL:
        if norm(sp) in n and norm(sp) not in dn:
            b-=4
    # 標準系: 図面に特記が無いとき、標準/静止型を優先的に加点
    DEFAULTS=['静止型','標準','(手動)']
    has_special_in_draw = any(norm(s) in dn for s in ['引出','電動','可逆','スペース'])
    if not has_special_in_draw:
        for df in DEFAULTS:
            if norm(df) in n: b+=4
    return b

# ===== フェーズ3: AI絞り込み（DB候補内のみ・生成禁止）=====
def ai_pick(cli, raw):
    n=norm(raw); toks=[w for w in re.split(r'[\s\u3000()（）]',raw) if len(w)>=2]
    cand=[d for d in DB if any(norm(tk) in norm(d['name']) for tk in toks)][:25]
    if not cand: return dict(code='',name='',conf='△',note='DB候補なし・初見/要確認')
    cand_txt="\n".join(f"{c['code']}\t{c['name']}\t{c['kind']}" for c in cand)
    # 確信度も返させる（HIGH=ほぼ確実 / LOW=自信なし）。形式: 「コード,確信度」
    prompt=f"""図面の機器表記「{raw}」に最も合致する積算コードを下記候補から1つ選び、確信度も答えてください。
候補に適切なものが無ければ NONE。候補外コードの創作は禁止。
回答形式は「コード,HIGH」または「コード,LOW」または「NONE」のみ。説明不要。
HIGH=表記と候補が明確に一致し確信できる / LOW=候補はあるが断定できない
候補:
{cand_txt}"""
    try:
        msg=cli.messages.create(model=MODEL,max_tokens=20,messages=[{"role":"user","content":prompt}])
        out="".join(b.text for b in msg.content if b.type=="text").strip()
        m=re.match(r'\s*(\d+|none|NONE)\s*,?\s*(HIGH|LOW)?', out, re.I)
        if m:
            ans=m.group(1); conf_lv=(m.group(2) or 'LOW').upper()
            if ans in byCode:
                # 確信度HIGH → ○に格上げ。LOW → △のまま（人確認）
                if conf_lv=='HIGH':
                    return dict(code=ans,name=byCode[ans]['name'],conf='○',note='AI選定(確信度高)')
                return dict(code=ans,name=byCode[ans]['name'],conf='△',note='AI候補内選定(要人確認)')
    except Exception: pass
    return dict(code='',name='',conf='△',note='要確認')

# ===== 統合処理 =====
_EXTRACT_CACHE={}  # 入力ハッシュ→抽出結果。同じ図面は同じ結果を返す(再現性の保険)
def extract_panels(fname, data_bytes):
    import hashlib
    key=hashlib.sha256(data_bytes).hexdigest()
    if key in _EXTRACT_CACHE:
        return _EXTRACT_CACHE[key]
    cli=client()
    res=extract_input(cli, fname, data_bytes)
    # 図面種別を推定して各盤にタグ付け(1ファイル=1図面種別)。選定時の盤種判定の第一信号にする。
    _dk=_infer_drawing_kind(res.get('panels',[]))
    if _dk:
        for p in res.get('panels',[]):
            p.setdefault('_drawing_kind', _dk)
    _EXTRACT_CACHE[key]=res
    return res

# ===== 付属品の親吸収（二重計上の防止） =====
# 同一盤内で、親機器が付属検出器を内蔵している場合、単独計上された付属行を
# 「計上対象外(—/△)」に落として二重計上を防ぐ。消さずに残し理由を明記(トレーサビリティ)。
def _is_standalone_zct(row):
    """行が『単独のZCT/ZCTT』か。ZCT内蔵型の継電器(LG-RY/EL-RY等)や、
    名称にZCT以外の主機器を含むものは対象外(誤吸収防止)。"""
    nm=norm(row.get('raw',''))
    if not re.search(r'(?<![a-z])zct{1,2}(?![a-z])', nm):  # zct / zctt
        return False
    # 主機器名を伴う場合は付属でなく本体扱い → 吸収しない
    if re.search(r'(lgr|lg-ry|el-ry|igr|dgr|ry|mcb|elb|lbs|vcb|tr|変圧器)', nm):
        return False
    return True

def _has_zct_builtin_relay(rows):
    """盤内に ZCT内蔵型の地絡継電器(LG-RY/EL-RY 等、DB spec 'ZCT付'/'ZCT含')が
    選定されている行があれば True。"""
    for r in rows:
        code=r.get('code','')
        d=byCode.get(code,{})
        nm=d.get('name','')+d.get('spec','')
        if code and ('ZCT付' in nm or 'ZCT含' in nm or 'LG-RY' in d.get('name','')):
            return True
    return False

def absorb_accessories(rows):
    """盤内の付属吸収を適用。現状はB案件: LGR(ZCT付)内蔵時の単独ZCTを非計上にする。"""
    if _has_zct_builtin_relay(rows):
        for r in rows:
            if _is_standalone_zct(r):
                r['code']=''
                r['conf']='—'
                r['note']='LGR(ZCT付)に内蔵のため計上対象外(二重計上防止)'
                r['absorbed']=True
    return rows

# ===== 配電盤セットコード選定（決定的・確認ゲート付き / CLAUDE.md 第6章）=====
# 選定はDBの settype 付きレコード(11/16/17系)を属性完全一致＋容量切上で1つに決める。AI推測なし。
# 読み切れない仕様は推測せず確認ゲート(コンボボックス)へ。計器種別・端子盤極数は常に確認。
SET_CODES=[d for d in DB if d.get('settype')]
SC_DEFAULTS={'meter':'普通角'}
SC_OPTIONS={'meter':['普通角','広角','マルチ'],'phase':['1φ3W','3φ3W','スコット'],
            'role':['受電盤','饋電盤','一段積','二段積','三段積','母線連絡','母線連絡+一段積'],
            'vcb':['8KA','12.5KA'],'op':['手動','電動','電動引出','電磁','電磁引出PF'],'cap':[]}
SC_REQ={'低圧':['meter','phase','cap'],'高圧':['role','meter','vcb','op'],
        '段積':['role','meter','vcb'],'段積VCS':['role','op']}
SC_ALWAYS_CONFIRM={'meter','op'}   # 計器種別・VCB操作方式(手動/電動/引出)は単線図で誤読しやすく
                                   # 誤るとセットコードが変わる(◎誤答)→常に人が確認。実見積書照合で
                                   # 八戸受電盤が op=手動 と誤抽出(正解=電動)だった実例に基づく。

# セットが発火した盤で「セット内包の計器・変成器」を型で抑制する判定。
# 受電/低圧セット(11/16/17系)は計器一式(VM/AM/VS/AS/W/Wh/力率/マルチ指示計)と
# 計器用変成器(VT/CT/ZCT)をセット内に含む→積算ソフトがセットコードから展開するので個別計上しない。
# 個別に別計上する品(主変圧器TR・LBS・PF・分岐MCB・函体・SPD等)は抑制しない(TR/LBS/PFは別コード)。
_SET_METER_RE=re.compile(
    r'電圧計|電流計|電力計|電力量計|力率計|無効電力計?|周波数計|指示計|計器盤|指示計器盤|マルチ(指示計|メ[ータ]|)|'
    r'計器用変[成流圧]|変流器|表示灯|ﾊﾟｲﾛｯﾄ|ブザ|ﾌﾞｻﾞ|補助継電|'
    r'(?<![a-z0-9])(vm|am|vs|as|wh|whm|pfm|iwm|fm|cos|cs|vt|ct|zct|pl|bz|t-?ry|aux-?ry|27x|fl-?10w|pbs|tm-?sw)(?![a-z])', re.I)

def _is_set_internal_meter(name):
    """セット内包の計器/変成器か。主変圧器(変圧器/TR/kVA)や開閉器(LBS/PF/VCB)は除外。"""
    n=norm(name)
    # 計器盤(複合計器盤 V/A/DA/KW/KWh/Pf/var)はセット内包メータの集合→抑制(Pf=力率でPF除外に誤当たりするため先に判定)
    if re.search(r'計器盤|指示計器盤', n): return True
    if re.search(r'変圧器|(?<![a-z])tr(?![a-z])|kva|lbs|pas|vcb|vcs|(?<![a-z])pf(?![a-z])|mccb|mcb|elb|elcb|端子|函|盤$', n):
        return False
    return bool(_SET_METER_RE.search(n))

def _sc_capval(s):
    m=re.search(r'(\d+)',str(s or '')); return int(m.group(1)) if m else None

def _sc_meter_key(m):
    """計器種別の表記ゆれを吸収(DBは高圧/段積='普通'・低圧='普通角'、UI既定='普通角')。"""
    m=str(m or '')
    if 'マルチ' in m: return 'マルチ'
    if '広角' in m or m=='広': return '広角'
    if '普通' in m: return '普通'
    return m

def sc_classify(panel_name):
    n=str(panel_name or '')
    if re.search(r'受電',n): return {'settype':'高圧','role':'受電盤'}
    if re.search(r'饋電|き電',n): return {'settype':'高圧','role':'饋電盤'}
    # 分電盤(電灯分電盤/動力分電盤/1L-1等)は60系個別選定=配電盤セット対象外。制御盤も対象外。
    if re.search(r'分電|制御|操作',n): return {'settype':None}
    if re.search(r'スコット|ｽｺｯﾄ',n): return {'settype':'低圧','phase':'スコット'}
    # 17系低圧セットは受変電のTR二次側(低圧電灯盤/低圧動力盤/変圧器盤・TR盤)。'低圧'または変圧器盤の明示が要る。
    if re.search(r'低圧',n) and re.search(r'電灯',n): return {'settype':'低圧','phase':'1φ3W'}
    if re.search(r'低圧',n) and re.search(r'動力',n): return {'settype':'低圧','phase':'3φ3W'}
    if re.search(r'変圧器盤|ﾄﾗﾝｽ盤|TR盤',n): return {'settype':'低圧'}
    if re.search(r'コンデンサ|ｺﾝﾃﾞﾝｻ',n): return {'settype':None}
    return {'settype':None}

def sc_needs_confirm(attrs):
    return [k for k in SC_REQ.get(attrs.get('settype'),[]) if not attrs.get(k)]

def _sc_valid_options(st, spec, attrs):
    """settypeと既知属性から、実在するセットコードに存在するspec値だけを選択肢に返す。
    存在しない組合せ(例:受電盤に'電動引出')を選んで△の行き止まりになるのを防ぐ。"""
    if spec=='cap': return []                       # 容量は数値入力(選択肢なし)
    pool=[c for c in SET_CODES if c.get('settype')==st]
    # 確定済の他属性(role/phase/vcb等)で母集団を絞る。ALWAYS_CONFIRM(計器/操作)自身では絞らない。
    for k in SC_REQ.get(st,[]):
        if k==spec: continue
        v=attrs.get(k)
        if v and k not in SC_ALWAYS_CONFIRM:
            if k=='meter': pool=[c for c in pool if _sc_meter_key(c.get('meter'))==_sc_meter_key(v)]
            else: pool=[c for c in pool if c.get(k)==v]
    if spec=='meter':
        keys=set(_sc_meter_key(c.get('meter')) for c in pool)
        ui=[o for o in SC_OPTIONS.get('meter',[]) if _sc_meter_key(o) in keys]
        return ui or SC_OPTIONS.get('meter',[])
    vals=[c.get(spec) for c in pool if c.get(spec)]
    ordered=[o for o in SC_OPTIONS.get(spec,[]) if o in vals]
    for v in vals:                                   # 表示順にないDB実在値も末尾に補完
        if v not in ordered: ordered.append(v)
    return ordered or SC_OPTIONS.get(spec,[])

def _sc_default(spec, options, attrs):
    """初期選択値: 抽出値が選択肢にあればそれ、無ければ近い値/既定/先頭を返す。"""
    v=attrs.get(spec)
    if v and v in options: return v
    if spec=='op' and v:                             # 電動引出/電磁引出等→電動系にフォールバック
        if '電動' in v and '電動' in options: return '電動'
        if '手動' in v and '手動' in options: return '手動'
    if spec=='meter' and v:                          # 普通/広/等の表記ゆれを吸収
        for o in options:
            if _sc_meter_key(o)==_sc_meter_key(v): return o
    d=SC_DEFAULTS.get(spec,'')
    if d in options: return d
    return options[0] if options else ''

def sc_confirm_form(attrs):
    st=attrs.get('settype'); fields=list(sc_needs_confirm(attrs))
    for k in SC_REQ.get(st,[]):
        if k in SC_ALWAYS_CONFIRM and k not in fields: fields.append(k)
    out=[]
    for k in fields:
        opts=_sc_valid_options(st, k, attrs)
        out.append({'spec':k,'options':opts,'default':_sc_default(k, opts, attrs)})
    return out

def sc_apply_defaults(attrs):
    a=dict(attrs)
    for k in sc_needs_confirm(a):
        if k in SC_DEFAULTS: a[k]=SC_DEFAULTS[k]
    return a

def sc_select(attrs):
    st=attrs.get('settype'); pool=[c for c in SET_CODES if c.get('settype')==st]
    if st=='低圧':
        cs=[c for c in pool if _sc_meter_key(c.get('meter'))==_sc_meter_key(attrs.get('meter')) and c.get('phase')==attrs.get('phase')]
        kva=_sc_capval(attrs.get('cap'))
        if kva is None: return '','△','容量不明→確認'
        cs=[c for c in cs if _sc_capval(c.get('cap'))>=kva]
        if not cs: return '','△','該当容量なし→確認(コード表に無い小容量は追加せず人へ)'
        best=min(cs,key=lambda c:_sc_capval(c.get('cap')))
        if _sc_capval(best.get('cap'))==kva: return best['code'],'◎',''
        return best['code'],'○','容量切上 %d→%dKVA'%(kva,_sc_capval(best.get('cap')))
    if st=='高圧':
        cs=[c for c in pool if c.get('role')==attrs.get('role') and _sc_meter_key(c.get('meter'))==_sc_meter_key(attrs.get('meter'))
            and c.get('vcb')==attrs.get('vcb') and c.get('op')==attrs.get('op')]
        return (cs[0]['code'],'◎','') if len(cs)==1 else ('','△','高圧セット属性不一致→確認')
    if st=='段積':
        cs=[c for c in pool if c.get('role')==attrs.get('role') and _sc_meter_key(c.get('meter'))==_sc_meter_key(attrs.get('meter')) and c.get('vcb')==attrs.get('vcb')]
        return (cs[0]['code'],'◎','') if len(cs)==1 else ('','△','段積セット属性不一致→確認')
    if st=='段積VCS':
        cs=[c for c in pool if c.get('role')==attrs.get('role') and c.get('op')==attrs.get('op')]
        return (cs[0]['code'],'◎','') if len(cs)==1 else ('','△','VCS段積属性不一致→確認')
    return '','△','盤種セット対象外'

def sc_resolve(panel_name, set_attrs=None):
    attrs=dict(sc_classify(panel_name))
    if set_attrs:
        for k,v in set_attrs.items():
            if v and k!='settype': attrs[k]=v
        if set_attrs.get('settype'): attrs['settype']=set_attrs['settype']
    if not attrs.get('settype'): return None
    form=sc_confirm_form(attrs)
    code,conf,note=sc_select(sc_apply_defaults(attrs))
    # 確認ゲート項目(計器種別/操作方式=誤読でコードが変わる)が未確定のうちは◎にしない。
    # ゲートで人が確定(set_attrs['_confirmed']=True)して初めて◎。既定/抽出値のままは○(要ゲート確定)。
    # 実見積書照合で八戸受電盤が既定=普通角/抽出=手動で誤◎だった反省に基づく。
    if code and conf=='◎' and form and not (set_attrs or {}).get('_confirmed'):
        specs='・'.join({'meter':'計器種別','op':'操作方式','vcb':'VCB容量','role':'役割','cap':'容量'}.get(f['spec'],f['spec']) for f in form)
        conf='○'; note=('確認ゲートで%sを最終確定してください(既定/抽出値のまま)'%specs)+((' '+note) if note else '')
    return {'settype':attrs['settype'],'attrs':attrs,'code':code,'conf':conf,'note':note,'confirm':form}

# ---- 弱電端子盤(62系)選定: 極数P＋端子指標で最近傍上位。極数が読めなければ確認(呼出側) ----
_TERM_EXCLUDE=re.compile(r'MCB|ELB|MCCB|ELCB|LBS|VCB|VCS|DS|TR|SC|SR|CT|VT|MGS|MCTT|ﾌﾞ-ｽﾀ|ｾﾊﾟﾚ|ｺﾝｾﾝﾄ|コンセント|ﾋｰﾀ|換気|FL|PL|RY|BZ|COS|PBS')
_TERM_POS=re.compile(r'端子|MDF|主配線|保安器|安定器|E1|E2|E3|ﾄｸE|トクE|接地|T付|(^|\s)TB(\s|$)|電話|放送|情報|通信|LAN|露出盤')
def _term_poles(name):
    m=re.search(r'(\d+)\s*[PＰ]',str(name)); return int(m.group(1)) if m else None
def terminal_select(name):
    n=re.sub(r'[（）]',lambda m:'(' if m.group()=='（' else ')',str(name))
    if _TERM_EXCLUDE.search(n) and not re.search(r'端子|MDF|保安器',n): return None
    p=_term_poles(n)
    p_explicit = p is not None   # 端子数(P)が名称に明示されているか。既定値なら◎にしない(端子数要確認)。
    # 極数表記が無くても「端子盤/保安器/接地端子」と分かれば既定極数で選定(プロは10P端子無で計上)。
    if p is None:
        if not _TERM_POS.search(n): return None
        p=5 if '保安器' in n else 10   # 保安器函は5P、端子盤は10Pを既定
    elif not _TERM_POS.search(n): return None
    if re.search(r'E1|E2|E3|ﾄｸE|トクE|接地',n): fam='接地端子盤'
    elif '保安器' in n: fam='保安器函'
    elif re.search(r'MDF|主配線',n) and '保安器' not in n: fam='MDF(主配線盤)'
    elif '安定器' in n: fam='安定器'
    else: fam='端子盤'
    # 端子台の既定は「端子無」(プロ確定: 弱電端子盤T-xは62001端子無)。端子付は明記時のみ。
    tsuki = bool(re.search(r'T付|端子付',n))
    def pool(pred):
        out=[]
        for c,row in byCode.items():
            pp=_term_poles(row['name'])
            if pp and pred(row['name']): out.append((pp,c))
        return sorted(out)
    if fam=='端子盤':
        cand=pool(lambda nm:'端子盤' in nm and ('端子付' if tsuki else '端子無') in nm)
    else:
        cand=pool(lambda nm:fam.split('(')[0] in nm)
    pick=next((c for pp,c in cand if pp>=p),(cand[-1][1] if cand else None))
    if not pick: return None
    gp=_term_poles(byCode[pick]['name'])
    # 端子数が明示され、かつ在庫極数と一致した時のみ◎。既定極数(未明示)や切上は○(端子数要確認)。
    if p_explicit and gp==p: return pick,'◎',''
    _note='端子数(P)が図面から未確定→既定%dP・要確認'%p if not p_explicit else '極数切上%dP→%dP'%(p,gp)
    return pick,'○',_note

def select_from_extracted(data):
    out=[]
    # 受変電部で上段に出たマルチ指示計のコードを記憶し、下段のV/電流計に継承する。
    multi_meter_code=None
    # 動力制御盤の「標準図/パターン集」の凡例をマスター凡例として抽出。各盤の局所凡例より権威
    # (茂泉様確定): 標準図が記号A〜Lの正しい定義。局所凡例の誤取得(例: 六本木でA=電源送りを
    # A=直入L-Sと誤読)を是正し、記号の機器構成(電源送りMCCB か L-S始動 か)を正しく判定する。
    _master_legend={}
    for _p in data.get('panels',[]):
        if re.search(r'標準図|パターン集|パターン図|標準回路|回路図集', str(_p.get('panel',''))) and _p.get('legend'):
            for _k,_v in _p['legend'].items():
                _master_legend[str(_k).upper().strip()]=_v
    for p in data.get('panels',[]):
        # 図面種別ヒントを盤ごとにセット(_panel_kindが名前判定できない盤のフォールバック)。
        _dk=p.get('_drawing_kind')
        # LP等の一体盤(名前が曖昧)は、負荷内容で区画を判定してヒントにする:
        #   ポンプ/MC/インバータ/制御/警報/電動 等のモーター制御負荷 → 制御区画(50系)
        #   照明/誘導灯/コンセント/看板/自動ドア 等 → 分電区画(60系)。
        _pn_norm=norm(p.get('panel',''))
        if re.search(r'(^|[^a-zａ-ｚ])[a-zａ-ｚ]?\d*lp[ｰ\-－]?\d', _pn_norm) and '制御' not in _pn_norm and '分電' not in _pn_norm:
            _itxt=' '.join(norm(str(it.get('name',''))) for it in p.get('items',[]))
            _dk='ctrl' if re.search(r'ポンプ|ﾎﾟﾝﾌﾟ|ｲﾝﾊﾞ|インバ|制御|警報|薬注|ﾌﾞｰｽﾀ|ブースタ|(?<![a-z])mc(?![a-z])|電動', _itxt) else 'bunden'
        _DRAWING_KIND.set(_dk if _dk in ('haiden','bunden','ctrl') else None)
        rows=[]
        prev_is_main=False
        panel_nm=p.get('panel','')
        # 有効凡例=局所凡例にマスター凡例(標準図)を上書き。標準図の記号定義を権威とし、
        # 局所凡例の誤取得を是正(記号A=電源送りMCCB等を正しく解決→電源送りは個別遮断器で拾う)。
        _eff_legend=dict(p.get('legend') or {})
        for _mk,_mv in _master_legend.items():
            _eff_legend[_mk]=_mv
        # 動力制御盤の「標準図/パターン集」は各負荷盤が参照する主回路パターンの凡例定義シート。
        # それ自体は積算対象でない(構成部品MC/2E/MMCB等は各分岐回路コードに内包)。各負荷盤は
        # 自前のlegendを持つので、この参照シートは盤ごとスキップする(誤△の大量発生を防ぐ)。
        if re.search(r'標準図|パターン集|パターン図|標準回路図?|回路図集|結線図', panel_nm):
            continue
        # 電力会社供給品(高圧キャビネット/UAS/PAS/区分開閉器等)は電力会社の資産=当社積算対象外→盤ごとスキップ。
        if re.search(r'電力会社供給|電力会社支給|(供給|支給)品\)?\s*$|電力会社.{0,4}(キャビネット|ｷｬﾋﾞﾈｯﾄ)', panel_nm):
            continue
        is_jushaden = ('受電' in panel_nm or '受変電' in panel_nm or '高圧' in panel_nm)
        # 制御盤の判定: 盤名に「制御/動力」が無くても、主回路パターン凡例(legend)や主回路記号(A-L)を
        # 持つ盤は動力制御盤→分岐MCBは端子台付き(50系)。盤名がM-1A/P-x等コードのみの制御盤を救済。
        _panel_ctrl = bool(p.get('legend')) or any(it.get('symbol') for it in p.get('items',[])) \
                      or ('自立' in panel_nm and not re.search(r'低圧|受電|配電|電灯盤|分電', panel_nm))
        _panel_nm_for_sel = panel_nm
        # 記号/凡例で_panel_ctrlがTrueでも、L/S番号(分電盤)の盤には「制御盤」を付加しない
        # (5L-1(自立型)等が制御盤50系扱いになり分岐がコンパクト化されず△になるのを防ぐ)。
        _is_Lpanel = bool(re.search(r'(^|[^a-zａ-ｚ])[a-zａ-ｚ]?\d*[lｌjｊsｓ][ｰ\-－]?(?:[a-zａ-ｚ][ｰ\-－]?)?\d', norm(panel_nm)))
        if _panel_ctrl and not re.search(r'制御|動力|分電', panel_nm) and not _is_Lpanel:
            _panel_nm_for_sel = panel_nm + ' 制御盤'   # _mcb_code等の盤種判定をctrl(50系)へ寄せる
        # 盤内コンデンサSCのkvarを先に把握(同一盤の直列リアクトルSRの容量算定に使う)
        _panel_sc_kvar=None
        for _it in p.get('items',[]):
            _nm=str(_it.get('name',''))
            if re.search(r'(?<![A-Za-z])SC(?![A-Za-z])|ｺﾝﾃﾞﾝｻ|コンデンサ', _nm):
                _mk=re.search(r'(\d+\.?\d*)\s*kvar', _nm, re.I)
                if _mk: _panel_sc_kvar=float(_mk.group(1))
        # --- 配電盤セット判定(後方互換: set_attrs が無ければ従来通り全item個別選定) ---
        # set_attrs があればセットコードを1行出力。セット内包品(計器/TR/LBS等)は個別計上せず抑制。
        # セットが確定した(code有)場合のみ内包品を抑制。未確定(vcb/op等が未確認)でも確認ゲート行は出す。
        _set_expand=set(); _set_meter=''; _set_row_ref=None
        # 配電盤(受変電: 低圧17系/高圧11系/段積16系)は全て個別で拾う(プロ助言・茂泉様確定)。
        # セットだと過不足の差し引きが大変。個別なら段積みか単独かの判別も不要になる。
        # ∴ settype付き(低圧/高圧/段積)はセットパスを通さず個別選定に回す。制御盤の分岐は22-29系(別経路)。
        _use_set = False
        if _use_set:
            _sc=sc_resolve(panel_nm, p['set_attrs'])
            if _sc:
                if _sc.get('code'):
                    _srow=byCode.get(_sc['code'],{})
                    _set_expand={x.split('x')[0] for x in (_srow.get('expand','') or '').split(';') if x}
                    _set_meter=_sc_meter_key(_srow.get('meter') or (_sc.get('attrs') or {}).get('meter'))
                    _set_row_ref=dict(code=_sc['code'],name=_srow.get('name',''),conf=_sc['conf'],
                                     note=_sc['note'],raw=panel_nm,qty='1',is_setcode=True,
                                     set_confirm=_sc['confirm'],set_attrs=dict(_sc.get('attrs') or {}),
                                     load_detail=False,feed='',deviations=[])
                    rows.append(_set_row_ref)
                elif _sc.get('confirm') and sc_classify(panel_nm).get('settype'):
                    # 未確定セット: 確認ゲート行を出す(コード空・△)。ユーザがコンボボックスで確定→再選定で発火。
                    # ただし盤名が配電盤セット(受電/饋電/低圧電灯・動力/スコット/変圧器盤)と分かる場合のみ。
                    # (分電盤/制御盤/端子盤が抽出時に誤って settype 付与されても、余計なゲート行を出さない)
                    rows.append(dict(code='',name='(セット未確定:計器種別/VCB等を確認)',conf='△',
                                     note=_sc.get('note','セット属性の確認が必要'),raw=panel_nm,qty='1',
                                     is_setcode=True,set_confirm=_sc['confirm'],set_attrs=dict(_sc.get('attrs') or {}),
                                     load_detail=False,feed=''))
        for it in p.get('items',[]):
            nm=it.get('name','')
            # 負荷明細行(親分岐MCBの負荷内訳)は、直前の機器行(親)に集約してリストから外す。
            # (ケーブル判定より先に処理。負荷名称末尾にケーブルサイズが付くため)
            if it.get('load_detail'):
                load_nm=re.split(r'\s', nm.strip())[0] if nm.strip() else ''
                if rows:
                    rows[-1].setdefault('loads',[]).append(load_nm or nm.strip())
                continue
            # 予備スペース(遮断器仕様なしの「予備」「スペース」)は空きスロット=機器なし→計上対象外。
            # (動力制御盤 適用表の control_apply でも予備行はスキップ。仕様付き「予備 MCCB2P 50/20AT」等は残す)
            if re.match(r'^\s*(予備|スペース|ｽﾍﾟｰｽ|SP)\s*(\(|（|[0-9A-Za-z\-]{0,6}\)?|）)?\s*$', nm) \
               and not re.search(r'MCCB|MCB|ELB|ELCB|LBS|\d\s*AF|\d\s*AT|[1-4]\s*P|\d+\s*A(?![A-Za-z])', nm, re.I):
                continue
            # 操作スイッチ(運転/停止SW)・CP(サーキットプロテクタ)はDBに個別コードが無い盤内小物→対象外。
            # ※押ボタンPBS(71041)・表示灯PL(71051)・ブザーBZ(74001)はコード有り=見積システム入力コード
            #   なので除外しない(_ctrl_part_codeで拾う。セット盤では set 側が内包→set内包抑制で除外)。
            if re.search(r'操作\s*(SW|スイッチ)|運転\s*(SW|スイッチ)|(^|\s)CP(\s|$)|サーキットプロテクタ|ｻｰｷｯﾄﾌﾟﾛﾃｸﾀ', nm) \
               and not re.search(r'切換|VS|AS|計器', nm):
                continue
            # 警報点・監視信号(◯◯異常/故障/満水/減水/外部接点)は警報盤の内部入力点であり
            # 個別計上対象外(警報盤の函体56000で計上)。盤名でなく信号ラベル(括弧付き含む)。
            if re.search(r'外部接点', nm):
                continue
            if re.search(r'異常|故障|満水|減水', nm) and re.search(r'警報|発電機|ポンプ|受水槽|ボイラー|受変電|外部', nm) \
               and not re.search(r'盤$|BOX|函|継電器|ﾘﾚｰ|リレー|RY', nm):
                continue
            # 「一括警報」「一括(MCCB)トリップ」「一括MCCB分岐」(各盤の信号/分岐を集約した表記)は
            # 機器単体でなく集約表記→個別計上対象外(実見積書にも該当コードなし)。
            if re.search(r'一括警報|一括.{0,4}トリップ|トリップ.{0,4}\(?一括|一括.{0,4}MCCB|一括.{0,4}分岐', nm) and not re.search(r'盤$|BOX|函|継電器|ﾘﾚｰ|リレー|RY', nm):
                continue
            # 操作電源(制御電源)回路は制御盤の内部制御機器→分岐回路コード/盤製作費に内包(実見積書でも個別計上なし)。
            # ただし独立した操作用変圧器(kVA明記)は別計上のため除外しない。
            if re.search(r'操作電源|制御電源', nm) and (_panel_ctrl or re.search(r'制御|動力', panel_nm)) \
               and not re.search(r'\d+\s*k?va|変圧器|ﾄﾗﾝｽ|トランス', nm):
                continue
            # 制御回路の計装小物(電極棒/電動弁MV/温度センサー/フロート/電磁弁/インターロックリレー等)は
            # DBに個別コードが無く制御一式・盤製作費に内包(実見積書に単体計上なし・東部の給水タンク制御フロー等)。
            # 遮断器/容量(AF/AT/kW)を伴わない計装小物のみ対象→対象外(行き止まり△を残さない・確定率100%)。
            # ※サーモスタット/温度調節器は盤付属品コード(74107)があるので除外しない(_pro_mapで計上)。
            if re.search(r'電極棒|電動弁|(?<![A-Za-z])MV(?![A-Za-z])|温度セ[ンﾝ]ｻ?[ーｰ]?|フロ[ーｰ]ト|ﾌﾛｰﾄ|電磁弁|(?<![A-Za-z])リレー\s*[\(（]?\s*(インターロック|ｲﾝﾀｰﾛｯｸ)|センサ[ー-]\s*[TＴ]?\s*[\(（]', nm) \
               and (_panel_ctrl or re.search(r'制御|動力|フロー', panel_nm)) \
               and not re.search(r'\d+\s*(AF|AT|kVA|kW)|MCCB|MCB|ELB|ELCB|継電器\d', nm, re.I):
                continue
            # 盤本体の函体(屋内自立型函体/盤製作の筐体そのもの)は盤製作費で、選定コード対象外
            # (小型の警報盤BOX 56系は「警報盤」で別途計上)。「屋内自立型函体」等の筐体行を除外。
            if re.search(r'自立型?函体|盤\s*函体|筐体|ｷｮｳﾀｲ', nm) and not re.fullmatch(r'\s*警報盤\s*', nm):
                continue
            # 電力会社の検針用メーター(取引用・貸与品)は当社積算対象外。DB非実在の「84リレー」(単独)も
            # コード化できない機器なので計上対象から除外(行き止まり△を残さない・確定率100%方針)。
            if re.search(r'電力会社.{0,4}(検針|メ-?タ|ﾒ-?ﾀ)|検針用\s*メ-?タ|取引用.{0,4}メ-?タ', nm) \
               or (re.search(r'電力契約メ-?タ|契約用メ-?タ', nm) and re.search(r'電力会社|取付|貸与', nm)):
                continue
            if re.fullmatch(r'\s*84\s*(リレー|ﾘﾚｰ|継電器)?\s*', nm) and not re.search(r'電圧|不足|過電圧|地絡', nm):
                continue
            # 盤リストの要約行(負荷名＋盤参照P-nのみで、電気的仕様が全く無い)は積算対象の機器行でなく
            # 盤の参照→除外(詳細機器は別明細で計上/未抽出なら再抽出=品質ゲート)。行き止まり△を残さない。
            # 仕様(kW/A/AF/V/VA/遮断器/計器/継電器/端子/函/PL等)を一切含まない負荷名のみが対象。
            if not re.search(r'\d+\s*(kw|kva|kvar|va|af|at|[av]\b|φ|kv)|MCB|MCCB|ELB|ELCB|LBS|VCB|VCS|LUG|(?<![A-Za-z])TB|継電器|ﾘﾚｰ|リレー|(?<![A-Za-z])RY|計器|SPD|(?<![A-Za-z])CT|(?<![A-Za-z])VT|(?<![A-Za-z])PF|端子|BOX|函|(?<![A-Za-z])PL|(?<![A-Za-z])BZ|COS|kg|セット|一式', nm, re.I) \
               and re.search(r'(制御盤|電源盤|動力盤)\s*[A-ZＡ-Ｚ]?P?[-ｰ]?\d', panel_nm) \
               and re.search(r'[A-ZＡ-Ｚ]?P[-ｰ]?\d\s*$|空調|ﾎﾟﾝﾌﾟ|ポンプ|ﾌｧﾝ|ファン|ボイラー|洗浄|炊飯|煮炊|検収|コンテナ', nm):
                continue
            # 盤見出しの重複行(item名＝盤名、または盤名+重量(kg)注記のみ・機器仕様なし)は盤自身の
            # 見出しが機器行として抽出されたもの→対象外(盤内機器・受電盤等のセットは別行で計上済)。
            _nm_np = re.sub(r'\s*[\(（]\s*\d+\s*(kg|ｋｇ|t|ﾄﾝ)\s*[\)）]\s*$', '', str(nm)).strip()
            if _nm_np and norm(_nm_np)==norm(panel_nm) \
               and re.search(r'盤$|受電|電灯盤|動力盤|分電盤|制御盤|配電盤|キュービクル|ｷｭ-ﾋﾞｸﾙ', str(nm)) \
               and not re.search(r'MCB|MCCB|ELB|ELCB|LBS|VCB|VCS|(?<![A-Za-z])TR|(?<![A-Za-z])CT|(?<![A-Za-z])VT|(?<![A-Za-z])SC|(?<![A-Za-z])SR|(?<![A-Za-z])PF|\d+\s*(AF|AT|kVA|kW|kvar)|[1-4]\s*[PＰ]\b|継電器|ﾘﾚｰ|リレー|計器|指示計', str(nm)):
                continue
            # 盤/設備の見出し行(機器詳細記載なし・設備名のみ)は個別機器でない→計上対象外
            # (盤内の機器/セットコードは別行で計上済)。
            if re.search(r'機器詳細記載なし|詳細記載なし|明細記載なし|機器記載なし', nm):
                continue
            # 制御盤内のモータ/回路参照タグ(M1-102, MG-101, P3-05, EM-2 等・機器仕様なし)は、
            # その回路自体は別行(分岐回路コード/遮断器)で計上済の参照ラベル→対象外(実見積書にもタグ計上なし)。
            if (_panel_ctrl or re.search(r'制御|動力|(^|[^a-z])[MP]\d*-?\d', panel_nm)) \
               and re.fullmatch(r'\s*(MG|MC|EM|M|P)\d*[-ｰ]\d+[A-Za-z]?\s*(\([^)]*\))?\s*', nm) \
               and not re.search(r'MCB|MCCB|ELB|ELCB|LBS|kW|VA|AF|kVA|[0-9]+\s*A(?![a-z])|回路|盤', nm):
                continue
            # 「高圧受電設備/受変電設備 6kV」等の設備一式見出し(電圧表記のみ・機器仕様なし)も対象外。
            if re.search(r'(高圧)?受(変)?電設備|ｷｭ-ﾋﾞｸﾙ設備|キュービクル設備', nm) \
               and not re.search(r'\d+\s*(AF|kW|kVA|P\b)|MCB|MCCB|ELB|LBS|VCB|VCS|TR|SC|SR|CT|VT|PF', nm):
                continue
            # 柱上装柱材・外構(玉碍子/腕金/支線/根かせ/引込柱/装柱/マスト)・照明器具/灯具は
            # 盤でなく別業者スコープ(柱上・外構・照明設備)→計上対象外。
            if re.search(r'碍子|腕金|アームタイ|ｱｰﾑﾀｲ|支線|根かせ|根枷|引込柱|装柱|管端|止水|(^|\s)マスト|照明器具|灯具|ダウンライト|ﾀﾞｳﾝﾗｲﾄ', nm):
                continue
            # 電力会社貸与品(取引用電力量計)・防災設備(自動火災報知・受信機)・弱電(電話/情報)・
            # 配線材(プルボックス)は本システム(盤コード)の対象外→計上しない。
            if re.search(r'電力量計.{0,4}(貸与|電力会社)|(貸与|電力会社).{0,4}電力量計|自動火災報知|火災報知|受信機|(^|\s)電話・?情報|プルボックス|ﾌﾟﾙﾎﾞｯｸｽ|(^|\s)PB[-\d]', nm):
                continue
            # 弱電通信設備の端子(機械警備/拡声/構内交換/誘導支援/非常放送/ｲﾝﾀｰﾎﾝ/共聴/監視カメラ等)は
            # 別業者スコープ(弱電端子盤)→計上対象外。ただし分電盤の負荷(VA表記)や盤本体・継電器は除外しない。
            if re.search(r'機械警備|拡声|構内交換|誘導支援|ｲﾝﾀｰﾎﾝ|インターホン|テレビ共聴|共聴|監視カメラ|ITV|放送設備|自火報|非常放送|防排煙|ﾅ-ｽｺｰﾙ|ナースコール', nm) \
               and not re.search(r'VA|盤$|BOX|函|継電器|ﾘﾚｰ|リレー|RY', nm):
                continue
            # 高圧ケーブル端末処理・ケーブルヘッド(CH)はケーブル付属→計上対象外(ケーブル類は対象外)。
            if re.search(r'ケーブルヘッド|ｹｰﾌﾞﾙﾍｯﾄﾞ|ケーブル端末|ｹｰﾌﾞﾙ端末|端末処理', nm):
                continue
            # 電力会社引込の高圧開閉器(UGS/UAS/PGS/地中線用気中・ガス開閉器)はDBに購入品コードが無い
            # =電力会社供給/支給の引込設備→当社積算対象外。地絡検出器(ZVD/PDS)も検出器単体は支給/内蔵→対象外。
            if re.search(r'(?<![A-Za-z])(UGS|UAS|PGS|PAS)(?![A-Za-z])|地中線用.{0,6}(気中|ｶﾞｽ|ガス)?.{0,4}開閉器|(?<![A-Za-z])(ZVD|PDS)(?![A-Za-z])', nm) \
               and not re.search(r'盤$|BOX|函|制御', nm):
                continue
            # テナント工事(端末開閉器以降)・別スコープの見出しは当社範囲外→対象外。
            if re.search(r'テナント工事|以降\s*テナント|別途工事|別途スコープ', nm):
                continue
            # 照明器具(誘導灯/非常照明の器具本体)・設備機器・貸与計器は別業者/対象外(分岐遮断器は別行で計上)。
            # ※負荷名としての「誘導灯 XXVA」は分電盤分岐で扱うが、器具単体行(コード化対象外)はここで除外。
            # 接頭の回路記号(A/a等)が付く器具単体行も対象(遮断器記号が無い＝器具本体)。
            if re.search(r'(^|\s)誘導灯|(^|\s)非常照明|(^|\s)非常灯|ドックレベラー|ﾄﾞｯｸﾚﾍﾞﾗｰ|(^|\s)DL(\s|$)|受水槽(?!.*(盤|警報|ポンプ制御))|SCG制御箱|(^|\s)TW-\d|私設メーター|ﾀﾞｲﾔﾙ温度計|ダイヤル温度計|(^|\s)SOG制御器|自動応答装置', nm) \
               and not re.search(r'MCB|MCCB|ELB|AF|\d+\s*A(?![a-z])|盤$|BOX|函', nm):
                continue
            # DCC(放電コイル)は高圧コンデンサに内蔵→計上対象外(コンデンサ側で計上)。
            if re.search(r'(?<![A-Za-z])DCC(?![A-Za-z])|放電コイル|ﾃﾞｨｽﾁｬｰｼﾞｺｲﾙ|ディスチャージコイル', nm) \
               and not re.search(r'盤$|BOX|函', nm):
                continue
            # 所用電源(所内電源)・警報表示用等の補機信号は盤製作/回路に内包→計上対象外。
            if re.fullmatch(r'\s*(所用電源|所内電源|警報表示用?|表示用|信号用|運転表示用?)\s*', nm):
                continue
            # 換気扇・TV/弱電収納盤・太陽光PCS(支給)・ケーブル引込口(板金付属)・空きスペース・取引/課金用電力量計(貸与)は対象外。
            if re.search(r'換気扇(?!.*(盤$|制御))|TV機器収納|TV収納|太陽光発電|ﾊﾟﾜｰｺﾝ|パワーコンディショナ|ケーブル引込口|引込口|(取引|課金|計量)用?.{0,4}電力量計|電力量計.{0,4}(取引|課金|W:課金)', nm) \
               and not re.search(r'MCB|MCCB|ELB|AF|\d+\s*A(?![a-z])|継電器|ﾘﾚｰ', nm):
                continue
            # 空きスペース(「13 スペース」「予備実装」等・機器なし)は空きスロット→対象外。
            if re.fullmatch(r'\s*(\d+\s*)?(スペース|ｽﾍﾟｰｽ|予備実装|空き?|SP)\s*', nm) or re.fullmatch(r'\s*[\(（]?回路?\d+[\)）]?\s*予備実装\s*', nm):
                continue
            # 予備・分岐の回路参照(予備(107)/分岐(G05)等・機器仕様なし)は空き回路の参照→計上対象外。
            if re.fullmatch(r'\s*(予備|分岐|スペース|ｽﾍﾟｰｽ)\s*[\(（][0-9A-Za-z\-]{1,6}[\)）]\s*', nm):
                continue
            # === 社内プロのフィードバック(△回答)由来の積算対象外ルール ===
            # 盤外設置: デマンド計/デマンド検出器/需要率計・サージインジケータ/雷電流記録カード。
            if re.search(r'デマンド計|ﾃﾞﾏﾝﾄﾞ計|デマンド検出|ﾃﾞﾏﾝﾄﾞ検出|需要率計|サージインジケータ|ｻ-ｼﾞｲﾝｼﾞｹ-ﾀ|雷電流記録|雷電流ｶ-ﾄﾞ|雷電流カード', nm):
                continue
            # 計器用PF(VTに含む): 定格の無い素の「PF」は計器用ヒューズ扱い→VTに含む。
            # (LBS内蔵PF・限流ヒューズPF(定格明記/コンデンサ40A級)は別扱いで残す)
            if re.fullmatch(r'\s*(PF|ＰＦ)\s*', nm) or (re.match(r'^\s*PF\s*$', nm)):
                continue
            # UP-OVG(地絡過電圧の付属)は継電器に含む→計上対象外。
            if re.search(r'UP-?OVG|OVG付属', nm):
                continue
            # 個別のアナログ計器(電圧計V/電流計A/切替スイッチVS・AS)は範囲外(客先/別途)。
            # ただし計器盤(複合)・DA(デマンド計器)・マルチ指示計は_pro_mapで計上するので除外しない。
            if re.search(r'電圧計|電流計|(^|\s)VS\s*電圧切替|(^|\s)AS\s*電流切替|電圧切替スイッチ|電流切替スイッチ', nm) \
               and not re.search(r'計器盤|指示計器盤|マルチ|ﾏﾙﾁ|(^|\s)DA(\s|$)', nm):
                continue
            # 発電機本体・エンジン・UPS(無停電電源)は支給品(客先支給・別途)でDB購入コードなし→計上対象外。
            # ただし「発電機充電用/ヒーター MCCB…」等の分岐遮断器や発電機「盤」は計上対象なので除外しない
            # (遮断器仕様AF/AT/MCCBや盤の語がある行は残す)。
            if re.search(r'発電機(?!盤)|ﾃﾞｨｰｾﾞﾙ|ディーゼル|(^|\s)D\.?ENG|ｴﾝｼﾞﾝ|(?<![ァ-ヶ])エンジン|(?<![A-Za-z])UPS(?![A-Za-z])|無停電電源', nm) \
               and not re.search(r'MCCB|MCB|ELB|ELCB|\d\s*AF|\d\s*AT|盤|充電用|ヒータ', nm, re.I):
                continue
            # 対象外の弱電機器: 本システムの弱電スコープは端子盤(端子/MDF/保安器)のみ。
            # TVアンテナ設備(アンテナ/マスト/増幅器/分配器/混合器/ブースター)・LAN機器(HUB)は
            # 別スコープ(弱電業者)→計上対象外(ユーザ方針: 弱電は端子のみ)。
            if re.search(r'アンテナ|ｱﾝﾃﾅ|マスト|ﾏｽﾄ|増幅器|ﾌﾞｰｽﾀ|ブースタ|分配器|分岐器|混合器|(?<![A-Za-z])HUB|ハブ|ﾊﾌﾞ', nm) \
               and not re.search(r'端子|MDF|保安器', nm):
                continue
            # 弱電/一般コンセント(遮断器仕様なし)は配線器具→対象外。
            # ただし動力コンセント分岐(例「フォーク用コンセント 3P 100AF/60AT」=遮断器付)は残す。
            if re.search(r'コンセント|ｺﾝｾﾝﾄ', nm) and not re.search(r'\d+\s*A[FT]|MCCB|MCB|ELB|ELCB', nm, re.I):
                continue
            # ケーブル・電線類は機器でないので計上対象外(リストから除外)
            # FP/FPT/CVT/CV/VVF/KIV/HIV等の電線。EV(エレベータ負荷)等と誤判定しないよう限定。
            if it.get('cable') or re.match(r'^\s*(\d+k?v\s+)?(fpt|cvt|cvv|vvf|kiv|hiv|fp|cv)\s*\d', norm(nm)) \
               or re.search(r'(fpt|cvt|vvf)\s*\d', norm(nm)):
                continue
            # 計器ヒューズ(F / F×N)は安価な付随小物なので計上対象外。
            # 「F×3」等はFが3個の意味。PF(パワーヒューズ=限流ヒューズ)は計上対象なので残す。
            _n=norm(nm)
            if not _n.startswith('pf'):  # PF(パワーヒューズ)は除外しない
                # 先頭がF(後ろが×N/数字/空白/末尾)、または「計器ヒューズ」表記
                if re.match(r'^f([x×]\s*\d+)?(\s|$|\d)', _n) or '計器ヒューズ' in nm:
                    continue
            # 計器ヒューズ(F/Fx/計器ヒューズ)は安価な付随小物なので計上対象外。
            # 限流ヒューズPF(高圧コンデンサ盤の40A級等)とは区別する。
            # norm後は空白除去・長音消失("計器ヒューズ"→"計器ヒュズ", "Fx2 3A"→"fx23a")。
            nmn=norm(nm)
            if re.fullmatch(r'f|fx\d*|fx\d*\d+a|計器ヒュ?ズ|計器用ヒュ?ズ', nmn) \
               or (nmn.startswith('fx') and 'pf' not in nmn):
                continue
            # 計器ヒューズ(F/Fx/計器ヒューズ)は安価な付随小物なので計上対象外。
            # ただし限流ヒューズPF(高圧40A級など)は計上対象なので除外しない。
            nn_chk=norm(nm)
            if not re.search(r'(?<![a-z])pf(?![a-z])', nn_chk):  # PFを含まないこと
                if re.match(r'^\s*f(x|\d|\b)', nn_chk) or '計器ヒューズ' in nm or re.match(r'^\s*f付', nn_chk):
                    continue
            # 数量サフィックス(x2等)を分離し、品名はクリーン版で選定
            cleaned, qsuf = split_qty_suffix(nm)
            # 品名に「×N」がある場合はそれを数量の正とする(抽出側の数量より優先)
            if qsuf: it['qty']=qsuf
            # 動力制御盤(記号/凡例あり)の主幹認識(茂泉様確定): 記号を持たない3P遮断器(AF枠付・
            # 操作電源/予備でない)は盤の頭=主幹(主開閉器)。「主幹」を冠して M)MCB/M)ELB として拾い、
            # M)LUG・制御一式(最低必要部品セット)を同時計上する。主幹は盤に1つ(既出なら対象外)。
            _base_nm = cleaned if qsuf else nm
            _nn=norm(nm)
            if _panel_ctrl and not it.get('symbol') and not re.match(r'\s*(主幹|主開閉|主)', _base_nm) \
               and re.search(r'(mccb|mcb|elcb|elb)', _nn) and re.search(r'\d{2,4}\s*af', _nn) \
               and re.search(r'3\s*[pφ]', _nn) and not re.search(r'操作|制御電源|予備|ｽﾍﾟｰｽ|スペース', _nn):
                if not any(str(byCode.get(str(r.get('code','')),{}).get('name','')).startswith('M)') for r in rows if not r.get('load_detail')):
                    _base_nm='主幹 '+_base_nm
            sel=select_one(_base_nm, _panel_nm_for_sel, prev_is_main, it.get('volt',''), it.get('symbol',''), it.get('kw',''), it.get('group',''),
                           legend=_eff_legend, breaker=it.get('breaker',''))
            # 弱電端子盤の救済: 未選定/△、または明らかに端子盤系(端子盤/保安器/接地端子/MDF)なら
            # terminal_select(極数P＋端子指標、既定=端子無)を優先。遮断器/コンセントは除外済。
            # (プロ確定: 端子盤T-x=62001端子無、保安器収納盤=62895保安器函)
            _is_term = bool(re.search(r'端子盤|保安器|接地端子|(電話|放送|情報|LAN|通信).{0,4}\d+\s*[PＰ]', nm)) \
                       and not re.search(r'MCB|ELB|MCCB|ELCB|LBS|VCB|VCS', nm, re.I)
            if (not sel.get('code')) or sel.get('conf')=='△' or _is_term:
                _tr=terminal_select(nm)
                if _tr and _tr[0]:
                    nmp=byCode.get(_tr[0],{}).get('name','')
                    sel=dict(code=_tr[0],name=nmp,conf=_tr[1],note='端子盤 極数選定'+(('・'+_tr[2]) if _tr[2] else ''))
            # 直列リアクトルSRの容量継承: SR単独で容量不明(△)なら、同一盤SCのkvarから算定(SR=L%×SC)。
            if ((not sel.get('code')) or sel.get('conf')=='△') and _panel_sc_kvar:
                _sr=_sr_from_sc(nm, _panel_sc_kvar)
                if _sr: sel=dict(code=_sr[0],conf=_sr[1],note=_sr[2],candidates=[])
            # 動力制御盤のモータ負荷救済: 主回路パターンが図面に無く未選定でも、
            # 3φ＋kWの明確なモータ負荷は「直入L-S(標準)」を仮定して分岐回路コードを△で提示。
            # (小容量モータは直入がほぼ標準。回路種別は要確認なので△のまま。空△より実用的)
            if (not sel.get('code')) and (re.search(r'制御|動力', str(p.get('panel',''))) or _panel_ctrl):
                _mm=re.search(r'([\d.]+)\s*KW', str(nm), re.I)
                # kWがあり1φ明記でなければ3φモーター負荷と仮定(PAC室外機等は「200V NNkW」で3φ表記が無い)。
                if _mm and not re.search(r'1\s*[φΦ]', str(nm)):
                    try: _kwf=float(_mm.group(1))
                    except: _kwf=0
                    if _kwf>0:
                        _cc,_cf,_pk=_bunki_find('L-S(AM付)',_kwf,'', (it.get('volt') if it.get('volt') in ('200V','400V') else '200V'))
                        if _cc:
                            # 枠不明のモーター負荷はL-S回路コードで計上し○(方式/回路種別は確認ゲートで確定)。
                            sel=dict(code=_cc,conf='○',note=f'回路種別=直入L-S(標準)仮定・確認ゲートで確定({_pk}kW枠)',candidates=[])
            # セット内包品(計器/TR/LBS/LG-RY等)はセットコードから積算ソフトが展開→個別計上しない
            if _set_expand and sel.get('code') in _set_expand:
                continue
            # セット発火盤の計器・変成器(VM/AM/VS/AS/W/Wh/力率/マルチ指示計・VT/CT/ZCT)は、
            # 選定コードがexpand表に無くてもセット内包→個別計上しない(主変圧器TR/LBS/PFは除外・別計上)。
            if _set_expand and _is_set_internal_meter(nm):
                # 抑制する前にMDA(マルチ指示計)なら型を記憶(下流の非セット盤の計器統一に使う)。
                if sel.get('code','')[:3]=='420' and 'マルチ指示計' in byCode.get(sel.get('code',''),{}).get('name','') \
                   or 'マルチ指示計' in nm or 'mda' in norm(nm):
                    if sel.get('code','')[:3]=='420': multi_meter_code=sel['code']
                # === セット構成差分チェック(茂泉様の原則) ===
                # 標準構成(expand)に無い『追加計器』は金額差を生むので抑制せず個別計上+セット標準外フラグ。
                # 基本計器(VM/AM/VS/AS/切換)とマルチ計器セット(マルチメータが全計器を覆う)は従来どおり内包。
                _nn=norm(nm)
                _is_basic=bool(re.search(r'(?<![a-z])(vm|am|vs|as|v|a)(?![a-z])|電圧計|電流計|切換', _nn)) \
                          and not re.search(r'iwm|無効|力率|pfm|周波|(?<![a-z])fm(?![a-z])|(?<![a-z])wh?(?![a-z])|kwh|電力量|電力計|需要|デマンド', _nn)
                _is_extra_meter=bool(re.search(r'iwm|無効電力|(?<![a-z])pfm(?![a-z])|力率計|周波数計|(?<![a-z])fm(?![a-z])|(?<![a-z])wh(?![a-z])|電力量計|(?<![a-z])w(?![a-z])電力|電力計', _nn))
                if _set_meter!='マルチ' and sel.get('code') and sel.get('code') not in _set_expand and not _is_basic and _is_extra_meter:
                    sel['conf']='○'
                    sel['note']='セット標準外の追加計器(金額差・個別計上): '+byCode.get(sel.get('code',''),{}).get('name','')[:16]
                    if _set_row_ref is not None: _set_row_ref['deviations'].append(sel.get('code'))
                    # 抑制せず下へ流して個別計上する
                else:
                    continue
            nn=norm(nm)

            # --- マルチ指示計(MDA)の型継承(案件全体・盤またぎ) ---
            # 受電盤等でマルチ指示計(42075-42088)が出たら型を記憶。型式は図面で確定困難なため△。
            if sel.get('code','')[:3]=='420' and 'マルチ指示計' in byCode.get(sel['code'],{}).get('name',''):
                multi_meter_code=sel['code']
                # 型式は図面から確定困難だが、既定型式を出し確認ゲートで最終確定→○(行き止まり△を無くす)。
                sel['conf']='○'
                sel['note']='マルチ指示計・型式を確認ゲートで最終確定(既定' + byCode.get(sel['code'],{}).get('name','')[:16] + ')'
                sel['candidates']=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in ('42081','42082','42083','42084','42086','42087','42088') if c in byCode]
            # 上位にMDA(マルチ指示計)がある場合、以降のどの盤の計器(電圧計V/電流計A/
            # 切換スイッチTHR=VS/AS)もMDAに統一する(盤またぎ・型式要確認△)。
            # ただしTR・変圧器・SC・SR等の明確な機器は継承対象にしない(誤統一を防ぐ)。
            elif multi_meter_code and not re.search(r'(tr|変圧器|sc|sr|mcb|elb|lbs|vcb|vcs|vmc|pas|ds|lgr|zct|ct|vt|pf|kva|kvar|kw)', nn) and (
                    sel.get('code','') in ('71001','71002','98802','98803','42001','42002','42003','42014','74275','74276')
                    or re.fullmatch(r'(thr|vs|as|v/s|a/s|切換スイッチ|切換sw|vm|am|vm/am|v|a|電圧計|電流計|計器\(?[va/]+\)?)', nn)):
                mm=byCode.get(multi_meter_code,{}).get('name','')
                sel=dict(code=multi_meter_code, name=mm, conf='○',
                         note=f'上位MDA(マルチ指示計)に統一・型式を確認ゲートで最終確定({mm[:16]})',candidates=[])

            prev_is_main = bool(re.search(r'm\)?(mcb|lug)', nn)) and ('tb' not in nn[:3])
            if it.get('unclear'):
                # 読取不明瞭でもコードが出ていれば行き止まり△にせず○(確認ゲートで確定)。◎のみ○へ格下げ
                # (◎誤答ゼロ)。コードが無い時のみ△(真に特定不能)。確定率100%方針。
                if sel.get('code'):
                    if sel['conf']=='◎': sel['conf']='○'
                    if '要確認' not in str(sel.get('note','')) and 'ゲート' not in str(sel.get('note','')):
                        sel['note']='図面読取要確認・'+str(sel.get('note',''))
                elif sel['conf']!='△':
                    sel['conf']='△'; sel['note']='図面読取不明瞭・'+str(sel.get('note',''))
            row=dict(sel)
            row['raw']=it.get('name','')
            row['qty']=it.get('qty','')
            row['load_detail']=False
            # 幹線番号(1L1,2M1,1EM2等)を抽出し、表示用に〈〉でくくる
            mfeed=re.match(r'\s*(\d+[A-Za-z]+\d*)', it.get('name','') or '')
            row['feed']=mfeed.group(1) if mfeed else ''
            rows.append(row)
            # 相間バリア(43392): 標準高圧LBS(43321=AL付)には必ず1台オプションで付随(実見積書4案件で
            # 43392件数=43321件数と厳密一致。エネセーバ43347には付かない=六本木で43392=43321数のみ)。
            # 決定的な同時計上(取りこぼし解消)。エネセーバ/素PF等の他変種は付随実績が無いので対象外。
            if row.get('code')=='43321' and '43392' in byCode:
                rows.append(dict(code='43392',name=byCode['43392'].get('name',''),conf='○',
                                 note='高圧LBS(43321)に付随の相間バリア(標準オプション・同時計上)',
                                 raw=it.get('name',''),qty=it.get('qty',''),load_detail=False,feed=row.get('feed','')))
        # 付属品の親吸収(二重計上防止): 盤内で確定後にまとめて適用
        rows=absorb_accessories(rows)
        # 各行の表示名「部品名＋仕様 〈幹線〉(負荷名称)」を組み立てる
        for r in rows:
            base=r.get('raw','')
            feed=r.get('feed','')
            # 幹線番号を raw 先頭から除いた本体を部品名＋仕様とする
            body=base
            if feed and base.startswith(feed):
                body=base[len(feed):].strip()
            disp=body
            if feed: disp=f'{body} 〈{feed}〉'
            loads=r.get('loads',[])
            if loads: disp=f'{disp}({", ".join(loads)})'
            r['display']=disp.strip()
        # MCTT 47系: 動力配電盤(非常・保安動力盤)は盤内の最大容量1台をDT(主電源切替)へ昇格、他はST維持。
        # 電灯盤(動力の語なし)は全てST維持(実見積書で電灯盤=全ST)。
        _mrows=[r for r in rows if r.get('_mctt')]
        if _mrows and '動力' in norm(p.get('panel','')) and _mctt_kind(p.get('panel',''))=='haiden':
            _mx=max(_mrows, key=lambda r: r['_mctt']['amp'])
            _dt=_mx['_mctt']['dt']
            if _dt in byCode:
                _up='容量繰上' in str(_mx.get('note',''))
                _mx['code']=_dt; _mx['name']=byCode[_dt].get('name','')
                _mx['conf']='○' if _up else '◎'
                _mx['note']='MCTT 3P-DT %dA(盤内最大=主電源切替)%s'%(_mx['_mctt']['amp'],'(容量繰上)' if _up else '')
                # コード表p13: DTの時は必ず 18-100(電源切替制御一式:COS×1,PL×2,T-Ry×2,AUX-Ry×2)を同時計上。
                # AUX-Ryの型(標準73000/高級73001)は案件の仕様レベルで変わる(森ビル等の高仕様=高級)。
                # 仕様レベルは確認ゲートで確定: 既定=標準(18100)、高級(18101)を候補提示し○(要確認)。
                if '18100' in byCode:
                    _cands=[{'code':'18100','name':byCode.get('18100',{}).get('name',''),'volt':''}]
                    if '18101' in byCode: _cands.append({'code':'18101','name':byCode.get('18101',{}).get('name',''),'volt':''})
                    rows.append(dict(code='18100',name=byCode['18100'].get('name',''),conf='○',
                                     note='DT付随(コード表p13)。仕様レベルを確認ゲートで選択: 標準=18100/高級=18101(AUX-Ry高級)',
                                     raw='(MCTT DT付随 18-100)',qty='1',load_detail=False,feed='',
                                     candidates=_cands,spec_gate=True))
        for r in rows: r.pop('_mctt',None)   # 内部タグ除去
        # コード表p41: N/5A(CT動作型)のWHMはCTを拾う。CTは高圧→44系/低圧→72系、変流比は主幹電流に合わせる(茂泉様)。
        _n5=[r for r in rows if str(r.get('code','')) in ('70303','70313','70323','70333','70306','70316','70326','70336','70393','70396')]
        _hasct=any((str(r.get('code',''))[:2]=='72') or (str(r.get('code','')).startswith('44') and 'CT' in byCode.get(str(r.get('code','')),{}).get('name','')) for r in rows)
        if _n5 and not _hasct:
            # 主幹電流を取得(M)コード名 M)MCB/ELB 3P NNА、または主幹rawのAT)。
            _amps=[]
            for r in rows:
                _mn=byCode.get(str(r.get('code','')),{}).get('name','')
                _mm=re.search(r'3P\s*(\d+)A', _mn)
                if _mn.startswith('M)') and _mm: _amps.append(int(_mm.group(1)))
                _rr=re.search(r'主(幹|開閉器).{0,20}?(\d{2,4})\s*A[TF]?', str(r.get('raw','')))
                if _rr: _amps.append(int(_rr.group(2)))
            _amp=max(_amps) if _amps else None
            _is_hv=bool(re.search(r'受電|受変電|高圧|饋電', norm(p.get('panel',''))))
            _ct=''
            if _amp is not None:
                if _is_hv:
                    _ct='44121' if _amp<=40 else '44122' if _amp<=75 else '44123'
                else:
                    for a,c in [(10,'72000'),(15,'72001'),(100,'72002'),(200,'72003'),(300,'72004'),(400,'72005'),(500,'72006'),(600,'72007')]:
                        if _amp<=a: _ct=c; break
                    _ct=_ct or '72007'
            if _ct and _ct in byCode:
                rows.append(dict(code=_ct,name=byCode[_ct].get('name',''),conf='○',
                                 note='N/5A WHMのCTを拾う(コード表p41)・主幹%dA相当(%s)'%(_amp,'高圧44系' if _is_hv else '低圧72系'),
                                 raw='(N/5A WHM付随CT)',qty=str(len(_n5)),load_detail=False,feed=''))
            else:
                rows.append(dict(code='',name='',conf='△',note='N/5A WHMのCTを拾う(コード表p41)・主幹電流不明→変流比要確認',
                                 raw='(N/5A WHM付随CT)',qty=str(len(_n5)),load_detail=False,feed=''))
        # 分電盤(60系)・制御盤(50系): 各系統の主幹(M)MCB/M)ELB)ごとに、盤頭のM)LUGをペア計上。
        # M)LUG=盤頭の主幹端子(茂泉様確定)。容量=主幹のFA枠、系統=盤種。1系統=主幹1+M)LUG1。
        # ※TR盤(受変電40系)はM)LUGが低圧セット(17系)のexpandに内包されるので対象外(二重計上防止)。
        _kind=_panel_kind(p.get('panel',''))
        if _kind in ('ctrl','bunden'):
            _add=[]
            for r in rows:
                if r.get('load_detail'): continue
                cnm=str(byCode.get(str(r.get('code','')),{}).get('name',''))
                if re.match(r'M\)(MCB|ELB)', cnm):   # 主幹遮断器の行
                    fa=re.search(r'(\d{2,4})\s*A', cnm)
                    if fa:
                        lc=_lug_code('M)LUG 3P %sA'%fa.group(1), p.get('panel',''), {'main':'M)LUG'})
                        if lc: _add.append((lc, fa.group(1)))
            for lc,fa in _add:
                rows.append(dict(code=lc,name=byCode[lc].get('name',''),conf='○',
                                 note='盤頭の主幹端子M)LUG(主幹%sA・系統=盤種%s系)'%(fa,lc[:2]),
                                 raw='(盤頭 M)LUG)',qty='1',load_detail=False,feed=''))
        elif _kind=='haiden' and not _use_set:
            # 配電盤(受変電低圧40系・個別化): TR二次主幹端子M)LUG(TR二次容量→切上・40系)。
            _haslug=any(str(byCode.get(str(r.get('code','')),{}).get('name','')).startswith('M)LUG') for r in rows)
            if not _haslug:
                _names=[str(it.get('name','')) for it in p.get('items',[])]
                _tl=_tr_lug_from_names(p.get('panel',''), _names)
                if _tl:
                    rows.append(dict(code=_tl,name=byCode[_tl].get('name',''),conf='○',
                                     note='TR二次主幹端子M)LUG(配電盤個別化・TR二次容量→切上)',
                                     raw='(TR二次 M)LUG)',qty='1',load_detail=False,feed=''))
        # SPD用分離器の自動計上(茂泉様確定): SPD本体の頭には必ず分離器が付く(見積で拾われる)。
        # クラスI本体74131 → 分離器74113(3P)・本体は極数分(3)。クラスII本体74134/74136 → 分離器74123(3P)。
        # 系統(=SPD出現)ごとに本体1グループ+分離器1。本体行を検出して分離器をペア追加する。
        _SPD_SEP={'74131':'74113','74134':'74123','74136':'74123'}
        _sepadd=[]
        for r in rows:
            if r.get('load_detail'): continue
            bc=str(r.get('code',''))
            sep=_SPD_SEP.get(bc)
            if not sep: continue
            if bc=='74131':   # クラスI本体は1P型→3P系統では極数分(既定3)
                if str(r.get('qty','')).strip() in ('','1'): r['qty']='3'
            if sep in byCode: _sepadd.append(sep)
        for sep in _sepadd:
            rows.append(dict(code=sep,name=byCode[sep].get('name',''),conf='○',
                             note='SPD用分離器(SPD本体の頭に付く分離器・系統ごと1)',
                             raw='(SPD用分離器)',qty='1',load_detail=False,feed=''))
        # 制御盤の制御一式セット(茂泉様確定): 制御盤で主幹(M)MCB/M)ELB)またはM)LUGが選定されたら、
        # 付随の制御一式セット(既定21310=BZ/AUX-RY/PBS/T-RY内包)を自動計上。個別部品は拾わずセットに内包。
        # T-RY無し=21309を候補提示(確認ゲート)。制御内器具は主幹に付随=常にセットとして拾われる。
        if _kind=='ctrl' and '21310' in byCode:
            _has_main = any(re.match(r'M\)(MCB|ELB|LUG)', str(byCode.get(str(r.get('code','')),{}).get('name','')))
                            for r in rows if not r.get('load_detail'))
            _has_set = any(str(r.get('code','')) in ('21309','21310') for r in rows)
            if _has_main and not _has_set:
                # セットに内包される個別部品(BZ/AUX-RY/PBS/T-RY)が個別抽出されていれば抑制(二重計上防止)。
                _incl={'74001','73000','73001','71041','73220'}
                rows=[r for r in rows if str(r.get('code','')) not in _incl or r.get('load_detail')]
                _cands=[{'code':c,'name':byCode.get(c,{}).get('name',''),'volt':''} for c in ('21310','21309') if c in byCode]
                rows.append(dict(code='21310',name=byCode['21310'].get('name',''),conf='○',
                                 note='制御盤の制御一式(主幹選定に付随・BZ/AUX-RY/PBS/T-RY内包)。T-RY無し=21309を確認ゲートで選択',
                                 raw='(制御盤 制御一式)',qty='1',is_setcode=True,candidates=_cands,
                                 set_attrs={},load_detail=False,feed='',deviations=[]))
        # 予備スペース(SP)分岐(茂泉様確定): 予備/スペースの分岐は実配線が無く端子台不要→系統は
        # 受変電(haiden)=40系/その他(制御盤・分電盤)=60系に強制し、SPコード(4桁目→9)へ変換。
        # 実見積書: 受変電 予備→40193/40293、制御盤 予備→60593(MCB)/60596(ELB) で確認。
        _sp_series = '40' if _kind=='haiden' else '60'
        for r in rows:
            if r.get('load_detail'): continue
            _rawn = str(r.get('raw','')) + '|' + str(r.get('name',''))
            # SP(空きスペース)は「スペース」明記のみ。裸の「予備」は実装済スペアブレーカー(通常コード)
            # の場合があり(例「予備 MCCB 3P225 定格電流可調整型」)、SPに誤変換しない。
            if not re.search(r'スペース|\(SP\)', _rawn): continue
            if 'SPD' in _rawn.upper(): continue          # SPDは対象外
            _c=str(r.get('code',''))
            _cnm=byCode.get(_c,{}).get('name','')
            _is_elb=('ELB' in _cnm)
            # 分電盤のコンパクト分岐(2P50A系=60012/60014)の予備 → コンパクトSP(MCB=60028/ELB=60029)。
            if _kind=='bunden' and (_cnm.startswith('B)MCB(コンパクト') or _cnm.startswith('B)ELB(コンパクト')
                                    or re.search(r'2P\s*50', _cnm)):
                _spc='60029' if _is_elb else '60028'
                if _spc in byCode:
                    r['code']=_spc; r['name']=byCode[_spc]['name']; r['conf']='○'
                    r['note']='予備スペース分岐(分電盤コンパクトSP)'
                continue
            if not (_cnm.startswith('B)MCB') or _cnm.startswith('B)ELB')): continue  # 分岐MCB/ELBのみ
            if len(_c)!=5 or _c[3] not in ('2','3','4'): continue     # 標準構造のみ(400xx系は除外)
            _spc=_sp_series+_c[2:3]+'9'+_c[4]                          # 系統強制+4桁目→9
            if _spc in byCode:
                r['code']=_spc; r['name']=byCode[_spc]['name']; r['conf']='○'
                r['note']='予備スペース分岐(端子台なし→%s系SP)'%_sp_series
        out.append(dict(panel=p.get('panel',''),rows=rows))
    # 受変電(高圧)図面の標準付属品(1図面に各1): テストプラグ98800・CH取付金具79030。
    # 実見積書7案件すべての受変電で計上(テストプラグ=全7案件×1、CH取付金具=×1〜2)。決定的な同時計上。
    # 高圧受電機器(VCB/VCS/DS/高圧LBS/LA=43系の受変電コード)が1つでもあれば受変電図面と判定し、
    # 未計上なら各1を○で追加(取りこぼし解消)。低圧分電/制御盤のみの図面では発火しない。
    _allcodes=set(); _hv_panel=None
    for _pp in out:
        for _rr in _pp['rows']:
            _cc=str(_rr.get('code',''))
            if _cc: _allcodes.add(_cc)
            if _hv_panel is None and re.match(r'43(01\d|10\d|11\d|32\d|347|42\d|46\d)$', _cc):
                _hv_panel=_pp
    if _hv_panel is not None:
        for _acc in ('98800','79030'):
            if _acc in byCode and _acc not in _allcodes:
                _hv_panel['rows'].append(dict(code=_acc,name=byCode[_acc].get('name',''),conf='○',
                    note='受変電図面の標準付属品(1図面各1・実見積書7案件で計上)',load_detail=False))
    _DRAWING_KIND.set(None)   # 後続処理へ図面種別ヒントを漏らさない
    return out

def make_excel(panels):
    FONT='Meiryo'; thin=Side(style='thin',color='BBBBBB'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
    hf=PatternFill('solid',start_color='1E3A28')
    cf={'◎':PatternFill('solid',start_color='E8F0E8'),'○':PatternFill('solid',start_color='FFF8E0'),'△':PatternFill('solid',start_color='FCE4E4')}
    wb=Workbook(); ws=wb.active; ws.title='選定結果'
    # 社内プロのフィードバック用: 候補コードリスト＋確認ボックス(確認/修正コード/コメント)を追加。
    ws.append(['盤','抽出部品(図面表記)','数量','選定コード','正式品名(DB)','判定','候補コード(上位)','根拠・確認事項','確認','プロ修正コード','コメント'])
    for c in ws[1]:
        c.font=Font(name=FONT,bold=True,color='FFFFFF'); c.fill=hf; c.border=bd; c.alignment=Alignment(horizontal='center',wrap_text=True)
    nok=nw=nc=ndetail=0; prev=None
    def _cand_str(r):
        # 「機器未特定/辞書に無い」等の未同定行は、候補が汎用フォールバックで無関係→表示しない。
        if re.search(r'機器未特定|辞書に無い|DB該当なし|該当機器', str(r.get('note',''))):
            return ''
        cs=r.get('candidates') or []
        parts=[]
        for c in cs[:5]:
            cc=c.get('code','')
            if cc and cc!=r.get('code'):
                parts.append('%s(%s)'%(cc, byCode.get(cc,{}).get('name','')[:16]))
        return ' / '.join(parts)
    for p in panels:
        for r in p['rows']:
            is_detail=r.get('load_detail')
            ws.append([p['panel'],r['raw'],r['qty'],r['code'] or '—',byCode.get(r['code'],{}).get('name','') if r['code'] else '',
                       r['conf'],_cand_str(r),r['note'],'','',''])
            row=ws[ws.max_row]
            for c in row: c.font=Font(name=FONT,size=9); c.border=bd; c.alignment=Alignment(vertical='center',wrap_text=True)
            if p['panel']!=prev: row[0].font=Font(name=FONT,size=9,bold=True); row[0].fill=PatternFill('solid',start_color='F0F0F0'); prev=p['panel']
            if is_detail:
                # 負荷明細行: グレーでインデント表示、計上対象外
                for c in row: c.font=Font(name=FONT,size=9,italic=True,color='999999')
                row[1].alignment=Alignment(vertical='center',wrap_text=True,indent=2)
                ndetail+=1
                continue
            row[5].fill=cf.get(r['conf'],PatternFill()); row[5].alignment=Alignment(horizontal='center'); row[3].font=Font(name=FONT,size=9,bold=True)
            row[8].fill=PatternFill('solid',start_color='EAF1FB')  # 確認列を薄青で目立たせる
            if r['conf']=='◎': nok+=1
            elif r['conf']=='○': nw+=1
            else: nc+=1
    # 「確認」列(I列)にドロップダウン(未確認/OK/要修正)を設定=チェックボックス代わり
    dv=DataValidation(type='list', formula1='"未確認,OK,要修正"', allow_blank=True)
    dv.prompt='この選定でよければOK、違えば要修正を選び、右に正しいコードとコメントを記入'; dv.promptTitle='確認'
    ws.add_data_validation(dv); dv.add('I2:I%d'%ws.max_row)
    for i,w in enumerate([18,32,6,12,26,7,26,34,10,14,28],1): ws.column_dimensions[chr(64+i)].width=w
    ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:K{ws.max_row}'
    tot=nok+nw+nc or 1
    ws2=wb.create_sheet('集計',0)
    ws2.append(['積算コード選定システム 結果（社内レビュー用）']); ws2['A1'].font=Font(name=FONT,bold=True,size=13); ws2.append([])
    for lab,val in [('抽出機器数',tot),('◎ 確定',nok),('○ ほぼ確定',nw),('△ 要確認',nc),('自動確定率',f'{round((nok+nw)/tot*100)}%'),('負荷明細(計上対象外)',ndetail)]:
        ws2.append([lab,val])
    ws2.append([])
    ws2.append(['◎=確定 / ○=ほぼ確定 / △=要確認'])
    ws2.append(['レビュー方法: 「選定結果」シートの各行を確認し、「確認」列で 未確認/OK/要修正 を選択。'])
    ws2.append(['  違う場合は「候補コード(上位)」を参考に「プロ修正コード」へ正しいコードを記入し「コメント」に理由を。'])
    ws2.append(['  △(要確認)は図面から確定できなかった箇所です。特にご確認をお願いします。'])
    ws2.column_dimensions['A'].width=64; ws2.column_dimensions['B'].width=12
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ===== ルート =====
@app.route('/')
@login_required
def index(): return Response(INDEX_HTML, mimetype='text/html')

@app.route('/api/health')
def health(): return jsonify(ok=True, db=len(DB), key=bool(os.environ.get('ANTHROPIC_API_KEY')))

# 【段階1】抽出のみ：図面→盤・機器リスト（コード選定はまだしない）
@app.route('/api/extract', methods=['POST'])
@login_required
def api_extract():
    # 複数ファイル対応: getlistで全ファイルを受け取り、各々抽出してpanelsを統合。
    files=request.files.getlist('file')
    if not files:
        f=request.files.get('file')
        files=[f] if f else []
    if not files:
        return jsonify(error='ファイルがありません'),400
    all_panels=[]; errors=[]; nfiles=0
    for f in files:
        if not f or not f.filename: continue
        fname=f.filename; raw=f.read(); low=fname.lower()
        if not low.endswith(('.pdf','.png','.jpg','.jpeg','.dxf','.zip')):
            errors.append(f'{fname}: 非対応形式(PDF/PNG/JPG/DXF/ZIP)')
            continue
        try:
            data=extract_panels(fname, raw)
            for p in data.get('panels',[]):
                # どのファイル由来か分かるよう、盤名にファイル名を付記(任意)
                all_panels.append(p)
            nfiles+=1
        except Exception as e:
            errors.append(f'{fname}: {e}')
    if not all_panels and errors:
        return jsonify(error=' / '.join(errors)),500
    # 【抽出後・選定前の仕様確認ゲート】配電盤セット盤には、計器種別・VCB操作方式等の
    # 「図面に明記されず特記仕様/客先打合せで決まる」仕様の確認フォームを付ける。
    # 社員はコード選定の前にここを確定してから選定に進む(誤った既定値での◎誤答を防ぐ)。
    for p in all_panels:
        attrs=dict(sc_classify(p.get('panel','')))
        sa=p.get('set_attrs') or {}
        for k,v in sa.items():
            if v and k!='settype': attrs[k]=v
        if sa.get('settype'): attrs['settype']=sa['settype']
        if attrs.get('settype'):
            p['sc_gate']=sc_confirm_form(attrs)
    nitems=sum(len(p.get('items',[])) for p in all_panels)
    nset=sum(1 for p in all_panels if p.get('sc_gate'))
    return jsonify(panels=all_panels, count=nitems, npanels=len(all_panels),
                   nfiles=nfiles, warnings=errors, nset=nset)

# 【段階2】選定：抽出済みの盤・機器リスト→コード選定（◎○△）
@app.route('/api/select', methods=['POST'])
@login_required
def api_select():
    data=request.get_json()
    if not data or 'panels' not in data:
        return jsonify(error='抽出データがありません'),400
    try:
        panels=select_from_extracted({'panels':data['panels']})
    except Exception as e:
        return jsonify(error=str(e)),500
    c={'◎':0,'○':0,'△':0}
    ndetail=0
    for p in panels:
        for r in p['rows']:
            r['official']=byCode.get(r['code'],{}).get('name','') if r['code'] else ''
            r['name']=r.get('display') or r.get('raw','')
            # 候補を「コード（品名）」形式で最大5件用意(フロントのプルダウン用)
            cand_list=r.get('candidates',[]) or []
            r['cand_opts']=[{'code':cc['code'],
                             'label':f"{cc['code']}（{cc.get('name','')}）"}
                            for cc in cand_list[:5]]
            # 負荷明細行(conf='—')は計上対象外。集計母数に含めない。
            if r.get('load_detail') or r['conf'] not in c:
                ndetail+=1
                continue
            c[r['conf']]+=1
    tot=sum(c.values()) or 1
    # 確実ダウンロード用: 直近の選定結果をセッション単位でメモリ保持。
    # (POST+Blob方式はiOSで落ちるため、GETでファイル取得できるようにする)
    try:
        sid=session.get('sid')
        if not sid:
            sid=secrets.token_hex(8); session['sid']=sid
        _LAST_RESULT[sid]=panels
    except Exception:
        pass
    return jsonify(panels=panels, summary=dict(total=sum(c.values()),ok=c['◎'],warn=c['○'],chk=c['△'],
        detail=ndetail, rate=round((c['◎']+c['○'])/tot*100)))

# 直近の選定結果(セッションID→panels)。プロセス内メモリ。再起動で消えるが実用上十分。
_LAST_RESULT={}

@app.route('/api/excel', methods=['POST'])
@login_required
def api_excel():
    panels=request.get_json().get('panels',[])
    buf=make_excel(panels)
    ts=datetime.datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf,as_attachment=True,download_name=f'estimate_code_{ts}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# 確実ダウンロード(GET): ブラウザでURLを開くだけで保存できる。iOS/PC問わず動作。
@app.route('/api/excel/download', methods=['GET'])
@login_required
def api_excel_download():
    sid=session.get('sid')
    panels=_LAST_RESULT.get(sid)
    if not panels:
        return Response('選定結果がありません。先に図面を読み込み、コード選定を実行してください。',
                        status=404, mimetype='text/plain; charset=utf-8')
    buf=make_excel(panels)
    ts=datetime.datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf,as_attachment=True,download_name=f'estimate_code_{ts}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

INDEX_HTML=open(os.path.join(HERE,'index.html'),encoding='utf-8').read() if os.path.exists(os.path.join(HERE,'index.html')) else '<h1>index.html がありません</h1>'

if __name__=='__main__':
    port=int(os.environ.get('PORT','8000'))
    print(f'積算コード選定システム起動: http://localhost:{port}  (DB {len(DB)}件 / APIキー {"OK" if os.environ.get("ANTHROPIC_API_KEY") else "未設定"})')
    app.run(host='0.0.0.0',port=port,debug=False)
