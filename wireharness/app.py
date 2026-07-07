#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ワイヤーハーネス製作支援システム v0.1  Flask版（図面 → 電線リスト / 切断・端末加工表 / 結束・製作図）

概要:
  制御盤・分電盤の結線図（展開接続図・端子台結線図・単線/複線図）をアップロードし、
  Claude Vision で「From-To の結線（配線）」を抽出。以下を自動生成する。
    1. 電線リスト   … 号線・接続元/先(機器記号:端子)・電線種類/サイズ/色・長さ
    2. 切断・端末加工表 … 電線ごとの切断長・両端の端子(圧着)・マーク(号線)
    3. 結束・製作図 … 幹線に対する分岐・区間本数を図化した SVG（製作作業指示）
    4. 部材集計   … 電線(種類/サイズ/色別 総長)・端子(種類別 本数)

設計方針（既存の積算システムとは独立した別アプリ）:
  - AI は「図面に描かれている結線」を読むだけ。存在しない配線を創作しない。
  - 読み取れない箇所は unclear=true にして人の確認対象にする。
  - 端子(圧着)種類はサイズから定石で推定し、あくまで候補として提示（画面で修正可）。

起動:
  export ANTHROPIC_API_KEY=sk-ant-...
  pip install -r requirements.txt
  python3 app.py        # → http://localhost:8010
"""
import os, re, io, json, base64, unicodedata, datetime, zipfile, tempfile, hmac, secrets, functools
from flask import Flask, request, jsonify, send_file, Response, session, redirect

HERE = os.path.dirname(os.path.abspath(__file__))
MODEL = os.environ.get('SELECT_MODEL', 'claude-opus-4-8')

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 40 * 1024 * 1024  # 40MB
app.secret_key = os.environ.get('APP_SECRET', secrets.token_hex(16))
APP_PASSWORD = os.environ.get('APP_PASSWORD', '')   # 未設定なら認証オフ（ローカル試用向け）

# 直近の生成結果をセッション単位で保持（Excel の GET ダウンロード用）
_LAST = {}


# ============================================================
#  認証（既存アプリと同じ簡易パスワード方式・任意）
# ============================================================
def login_required(f):
    @functools.wraps(f)
    def w(*a, **k):
        if not APP_PASSWORD:
            return f(*a, **k)
        if session.get('auth'):
            return f(*a, **k)
        if request.path.startswith('/api/'):
            return jsonify(error='未認証'), 401
        return redirect('/login')
    return w


LOGIN_HTML = """<!DOCTYPE html><html lang=ja><head><meta charset=UTF-8>
<meta name=viewport content="width=device-width,initial-scale=1"><title>ログイン</title>
<style>body{font-family:'Yu Gothic',Meiryo,sans-serif;background:#eef2f5;display:flex;
align-items:center;justify-content:center;height:100vh;margin:0}
.box{background:#fff;border:1px solid #cfd8dc;border-radius:10px;padding:32px;width:320px}
h1{font-size:16px;color:#123b5e;margin:0 0 18px}input{width:100%;padding:10px;border:1px solid #cfd8dc;
border-radius:6px;font-size:14px;box-sizing:border-box}button{width:100%;margin-top:12px;padding:11px;
background:#123b5e;color:#fff;border:0;border-radius:6px;font-size:14px;font-weight:700;cursor:pointer}
.e{color:#c0504d;font-size:12px;margin-top:10px}</style></head>
<body><form class=box method=post action=/login>
<h1>ワイヤーハーネス製作支援システム</h1>
<input type=password name=pw placeholder=パスワード autofocus>
<button>ログイン</button>
{ERR}</form></body></html>"""


@app.route('/login', methods=['GET', 'POST'])
def login():
    if not APP_PASSWORD:
        return redirect('/')
    if request.method == 'POST':
        if hmac.compare_digest(request.form.get('pw', ''), APP_PASSWORD):
            session['auth'] = True
            return redirect('/')
        return Response(LOGIN_HTML.replace('{ERR}', '<div class=e>パスワードが違います</div>'), mimetype='text/html')
    return Response(LOGIN_HTML.replace('{ERR}', ''), mimetype='text/html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


def client():
    if not os.environ.get('ANTHROPIC_API_KEY'):
        raise RuntimeError('ANTHROPIC_API_KEY 未設定')
    from anthropic import Anthropic
    return Anthropic()


def _sid():
    if not session.get('sid'):
        session['sid'] = secrets.token_hex(8)
    return session['sid']


# ============================================================
#  フェーズ1: 図面から「結線（配線）」を抽出（Vision）
# ============================================================
EXTRACT_PROMPT = """あなたは制御盤・分電盤の結線図（展開接続図・端子台結線図・実体配線図・単線/複線結線図）を読む配線設計の専門家です。
この図面から、盤内で実際に配線される「電線1本ごとの接続（From-To）」を抽出してください。目的はワイヤーハーネス（電線加工）の製作データ作成です。

【抽出する単位＝電線1本（配線1本）】
1本の電線は「接続元(機器記号:端子番号)」から「接続先(機器記号:端子番号)」までを結ぶ。
図面上の1本の線・1つの結線行が電線1本に対応する。分岐（渡り）がある場合は区間ごとに1本として分ける。

【各電線について読み取る項目】
- mark    : 号線番号／電線番号（図面に線番・記号があれば。無ければ空欄）
- from_dev, from_term : 接続元の機器記号と端子（例 MCCB1 の R相、端子台 X1 の 1番）
- to_dev,  to_term    : 接続先の機器記号と端子
- type : 電線種類（KIV / IV / HIV / AWG / CVV 等。読み取れれば）
- size : 電線サイズ（例 0.75 / 1.25 / 2.0 / 3.5 / 5.5、単位 sq。読み取れれば）
- color: 電線色（赤/白/黒/緑/黄 等。相色・記号があれば）
- length: 配線長さ mm（図面に寸法があれば数値。無ければ空欄。推測しない）
- note : 補足（渡り線、より線、シールド等）

【機器記号・端子の読み方】
- 機器記号: MCCB, ELB, MC(電磁接触器), THR, X(補助リレー), PB(押ボタン), PL(表示灯),
  TB/X/端子台 の番号, R/S/T/N(相), U/V/W, A1/A2(コイル), 13/14, 21/22(接点番号) 等。
- 端子番号: 端子台なら 1,2,3...、機器なら R/S/T, A1/A2, 1/2, 13/14 等そのまま。
- 記号が読めても端子が不明な場合は from_term/to_term を空にし unclear=true。

【厳守】
- 図面に描かれていない配線を創作しない。線をたどって確認できたものだけを出す。
- 読み取りが曖昧・つぶれて不確かな行は unclear=true を付ける（捨てない）。
- 数量は出さない（1行=電線1本）。同じ接続が複数本並列なら本数分の行にする。

【出力（JSONのみ・前後の説明文やコードフェンス不要）】
{
  "harnesses": [
    {
      "name": "盤名またはハーネス名（図面のタイトル・盤名。無ければ '盤1'）",
      "wires": [
        {"mark":"1","from_dev":"MCCB1","from_term":"R","to_dev":"MC1","to_term":"R",
         "type":"KIV","size":"2.0","color":"赤","length":"","note":"","unclear":false}
      ]
    }
  ]
}
"""


def _stream_json(cli, content):
    with cli.messages.stream(model=MODEL, max_tokens=64000, messages=[{"role": "user", "content": content}]) as stream:
        msg = stream.get_final_message()
    txt = "".join(b.text for b in msg.content if b.type == "text").strip()
    txt = re.sub(r'^```(json)?|```$', '', txt, flags=re.M).strip()
    try:
        return json.loads(txt)
    except Exception:
        m = re.search(r'\{.*\}', txt, re.S)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                pass
        if getattr(msg, 'stop_reason', None) == 'max_tokens':
            raise RuntimeError('配線数が多く抽出結果が出力上限を超えました。図面を分割して投入してください。')
        raise RuntimeError('抽出JSONの解析に失敗しました')


def extract_image(cli, data_bytes, media):
    src = {"type": "base64", "media_type": media, "data": base64.standard_b64encode(data_bytes).decode()}
    block = {"type": "document", "source": src} if media == "application/pdf" else {"type": "image", "source": src}
    return _stream_json(cli, [block, {"type": "text", "text": EXTRACT_PROMPT}])


def extract_pdf_hires(cli, data_bytes):
    """PDFを高解像度画像に変換してページごとに抽出（細かい線番・端子番号の読み取り精度向上）。"""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return extract_image(cli, data_bytes, "application/pdf")
    all_h = []
    doc = fitz.open(stream=data_bytes, filetype="pdf")
    for page in doc:
        pix = page.get_pixmap(matrix=fitz.Matrix(3, 3))
        png = pix.tobytes("png")
        res = extract_image(cli, png, "image/png")
        all_h += res.get("harnesses", [])
    return {"harnesses": all_h}


def extract_dxf(cli, data_bytes):
    import ezdxf
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as tf:
        tf.write(data_bytes)
        path = tf.name
    try:
        doc = ezdxf.readfile(path)
    except Exception as e:
        raise RuntimeError(f"DXF読込失敗: {e}")
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass
    texts = []
    for e in doc.modelspace():
        try:
            if e.dxftype() == "TEXT":
                texts.append((round(e.dxf.insert.x, 1), round(e.dxf.insert.y, 1), e.dxf.text))
            elif e.dxftype() == "MTEXT":
                texts.append((round(e.dxf.insert.x, 1), round(e.dxf.insert.y, 1), e.text))
        except Exception:
            continue
    if not texts:
        return {"harnesses": []}
    texts.sort(key=lambda v: (-v[1], v[0]))
    lines = "\n".join(f"({x},{y}) {t}" for x, y, t in texts)
    prompt = EXTRACT_PROMPT + ("\n\n以下はDXF図面から抽出した文字（座標(x,y)付き）です。"
                               "座標の近さから端子台・機器と結線の対応を判断し、電線ごとに構造化してください。\n\n" + lines[:60000])
    return _stream_json(cli, prompt)


def extract_one(cli, fname, data_bytes):
    low = fname.lower()
    if low.endswith(".pdf"):
        return extract_pdf_hires(cli, data_bytes)
    if low.endswith(".png"):
        return extract_image(cli, data_bytes, "image/png")
    if low.endswith((".jpg", ".jpeg")):
        return extract_image(cli, data_bytes, "image/jpeg")
    if low.endswith(".dxf"):
        return extract_dxf(cli, data_bytes)
    return None


def extract_input(cli, fname, data_bytes):
    low = fname.lower()
    if low.endswith(".zip"):
        all_h = []
        with zipfile.ZipFile(io.BytesIO(data_bytes)) as z:
            for info in z.infolist():
                if info.is_dir():
                    continue
                inner = info.filename
                if inner.startswith("__MACOSX") or "/." in inner:
                    continue
                if not inner.lower().endswith((".pdf", ".png", ".jpg", ".jpeg", ".dxf")):
                    continue
                base = os.path.basename(inner)
                try:
                    res = extract_one(cli, inner, z.read(info))
                except Exception as e:
                    all_h.append({"name": f"[抽出失敗] {base}",
                                  "wires": [{"mark": "", "from_dev": str(e), "from_term": "", "to_dev": "", "to_term": "",
                                             "type": "", "size": "", "color": "", "length": "", "note": "", "unclear": True}]})
                    continue
                for h in (res or {}).get("harnesses", []):
                    h["name"] = f"{base} / {h.get('name', '')}"
                    all_h.append(h)
        return {"harnesses": all_h}
    res = extract_one(cli, fname, data_bytes)
    if res is None:
        raise RuntimeError("未対応の形式です（PDF/PNG/JPG/DXF/ZIP のみ）")
    return res


# ============================================================
#  正規化・端子(圧着)推定・切断長
# ============================================================
def _num(v):
    try:
        return float(re.sub(r'[^\d.]', '', str(v)))
    except Exception:
        return None


def norm_wire(w):
    """抽出/入力された1本の電線を正規化。欠損キーを補完し、端子候補と切断長を付与。"""
    d = {
        "mark": str(w.get("mark", "") or "").strip(),
        "from_dev": str(w.get("from_dev", "") or "").strip(),
        "from_term": str(w.get("from_term", "") or "").strip(),
        "to_dev": str(w.get("to_dev", "") or "").strip(),
        "to_term": str(w.get("to_term", "") or "").strip(),
        "type": str(w.get("type", "") or "").strip(),
        "size": str(w.get("size", "") or "").strip(),
        "color": str(w.get("color", "") or "").strip(),
        "length": str(w.get("length", "") or "").strip(),
        "note": str(w.get("note", "") or "").strip(),
        "unclear": bool(w.get("unclear", False)),
    }
    # 端末(端子)は指定があれば尊重、無ければサイズから推定
    d["from_end"] = str(w.get("from_end", "") or "").strip() or suggest_terminal(d["size"])
    d["to_end"] = str(w.get("to_end", "") or "").strip() or suggest_terminal(d["size"])
    return d


# 電線サイズ(sq) → 圧着端子バレル呼び（丸端子 R形の定石）
_BARREL = [(1.25, "1.25"), (2.0, "2"), (3.5, "3.5"), (5.5, "5.5"),
           (8.0, "8"), (14.0, "14"), (22.0, "22"), (38.0, "38"), (60.0, "60")]
DEFAULT_STUD = os.environ.get('WH_STUD', '4')  # 既定のねじ径 M4


def suggest_terminal(size):
    """サイズから丸端子(R形)の呼びを推定。例 1.25sq → R1.25-4。あくまで候補。"""
    v = _num(size)
    if v is None:
        return ""
    for hi, barrel in _BARREL:
        if v <= hi + 1e-6:
            return f"R{barrel}-{DEFAULT_STUD}"
    return f"R{int(v)}-{DEFAULT_STUD}"


def cut_length(w):
    """切断長 = 配線長 + 両端の端末むき/圧着代（既定 各50mm）。長さ未入力なら空。"""
    L = _num(w.get("length"))
    if L is None:
        return ""
    allow = int(os.environ.get('WH_END_ALLOW', '50'))
    return str(int(round(L + allow * 2)))


# ============================================================
#  結束・製作図（SVG）: 幹線＋分岐＋区間本数のコーム図
# ============================================================
def _esc(s):
    return (str(s)).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def make_diagram_svg(name, wires):
    """機器を左→右に並べた幹線（結束幹線）に、各機器の分岐を描くコーム図。
    区間ごとに通過する電線本数を表示し、ハーネスの束構成を可視化する。"""
    # 機器の出現順（from→to）でユニーク化
    order = []
    for w in wires:
        for dev in (w["from_dev"], w["to_dev"]):
            if dev and dev not in order:
                order.append(dev)
    if not order:
        return '<svg xmlns="http://www.w3.org/2000/svg" width="480" height="80">' \
               '<text x="16" y="44" font-size="13" fill="#c0504d">機器記号が読み取れず、製作図を描けません。</text></svg>'
    idx = {d: i for i, d in enumerate(order)}
    n = len(order)
    # レイアウト
    MX, TOP = 60, 70
    slot = 150
    W = MX * 2 + slot * max(n - 1, 1)
    trunk_y = TOP
    dev_y = TOP + 90
    H = dev_y + 130

    # 区間(セグメント)通過本数を集計
    seg_cnt = [0] * max(n - 1, 1)
    dev_marks = {d: [] for d in order}
    for w in wires:
        fi, ti = idx.get(w["from_dev"]), idx.get(w["to_dev"])
        tag = w["mark"] or "・"
        if w["from_dev"]:
            dev_marks[w["from_dev"]].append(tag)
        if w["to_dev"] and w["to_dev"] != w["from_dev"]:
            dev_marks[w["to_dev"]].append(tag)
        if fi is None or ti is None:
            continue
        lo, hi = sorted((fi, ti))
        for s in range(lo, hi):
            seg_cnt[s] += 1

    def sx(i):
        return MX + slot * i

    p = ['<svg xmlns="http://www.w3.org/2000/svg" width="%d" height="%d" font-family="sans-serif">' % (W, H)]
    p.append('<rect x="0" y="0" width="%d" height="%d" fill="#ffffff"/>' % (W, H))
    p.append('<text x="16" y="28" font-size="15" font-weight="700" fill="#123b5e">%s ／ 結束・製作図</text>' % _esc(name))
    p.append('<text x="16" y="46" font-size="11" fill="#777">幹線に沿った機器分岐と、区間ごとの結束本数。寸法は電線リストの長さに準拠。</text>')
    # 幹線（太線）
    p.append('<line x1="%d" y1="%d" x2="%d" y2="%d" stroke="#123b5e" stroke-width="6" stroke-linecap="round"/>'
             % (sx(0), trunk_y, sx(n - 1), trunk_y))
    # 区間本数ラベル
    for s in range(max(n - 1, 1)):
        if n < 2:
            break
        mxs = (sx(s) + sx(s + 1)) / 2
        cnt = seg_cnt[s]
        p.append('<rect x="%d" y="%d" width="52" height="20" rx="4" fill="#eef2f5" stroke="#9db4c7"/>'
                 % (mxs - 26, trunk_y - 30))
        p.append('<text x="%d" y="%d" font-size="11" text-anchor="middle" fill="#123b5e">%d本</text>'
                 % (mxs, trunk_y - 16, cnt))
    # 各機器: 分岐線＋ボックス＋号線
    for d in order:
        x = sx(idx[d])
        p.append('<circle cx="%d" cy="%d" r="5" fill="#123b5e"/>' % (x, trunk_y))
        p.append('<line x1="%d" y1="%d" x2="%d" y2="%d" stroke="#123b5e" stroke-width="2"/>' % (x, trunk_y, x, dev_y - 24))
        p.append('<rect x="%d" y="%d" width="120" height="28" rx="5" fill="#123b5e"/>' % (x - 60, dev_y - 24))
        p.append('<text x="%d" y="%d" font-size="12" font-weight="700" text-anchor="middle" fill="#fff">%s</text>'
                 % (x, dev_y - 5, _esc(d)))
        marks = dev_marks.get(d, [])
        show = "／".join(marks[:8]) + ("…" if len(marks) > 8 else "")
        p.append('<text x="%d" y="%d" font-size="10" text-anchor="middle" fill="#555">号線 %s</text>'
                 % (x, dev_y + 14, _esc(show)))
        p.append('<text x="%d" y="%d" font-size="10" text-anchor="middle" fill="#888">計%d本</text>'
                 % (x, dev_y + 30, len(marks)))
    p.append('</svg>')
    return "".join(p)


# ============================================================
#  部材集計
# ============================================================
def aggregate(harnesses):
    wire_agg = {}   # (type,size,color) -> {"len": mm合計, "count": 本数}
    term_agg = {}   # 端子呼び -> 本数
    total = 0
    for h in harnesses:
        for w in h["wires"]:
            total += 1
            key = (w["type"], w["size"], w["color"])
            a = wire_agg.setdefault(key, {"len": 0, "count": 0, "has_len": False})
            a["count"] += 1
            L = _num(w.get("length"))
            if L is not None:
                a["len"] += L
                a["has_len"] = True
            for end in (w["from_end"], w["to_end"]):
                if end:
                    term_agg[end] = term_agg.get(end, 0) + 1
    wires = []
    for (t, s, c), a in sorted(wire_agg.items()):
        wires.append({"type": t, "size": s, "color": c, "count": a["count"],
                      "length_mm": int(round(a["len"])) if a["has_len"] else "",
                      "length_m": round(a["len"] / 1000.0, 2) if a["has_len"] else ""})
    terms = [{"terminal": k, "count": v} for k, v in sorted(term_agg.items())]
    return {"wires": wires, "terminals": terms, "total_wires": total}


def build_result(harnesses):
    """正規化 → 各ハーネスに切断長・端子候補・SVG・集計を付けて返す。"""
    out = []
    for h in harnesses:
        name = h.get("name", "") or "盤1"
        wires = [norm_wire(w) for w in h.get("wires", [])]
        for w in wires:
            w["cut"] = cut_length(w)
        out.append({"name": name, "wires": wires, "svg": make_diagram_svg(name, wires)})
    agg = aggregate(out)
    n_unclear = sum(1 for h in out for w in h["wires"] if w["unclear"])
    return {"harnesses": out, "aggregate": agg,
            "summary": {"harnesses": len(out), "wires": agg["total_wires"], "unclear": n_unclear}}


# ============================================================
#  Excel 出力（電線リスト / 切断・端末加工表 / 部材集計）
# ============================================================
def make_excel(result):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    thin = Side(style='thin', color='B0BEC5')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill('solid', fgColor='123B5E')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    pnl_fill = PatternFill('solid', fgColor='DCE6F1')
    ctr = Alignment(horizontal='center', vertical='center')

    def style_header(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = ctr
            cell.border = bd

    # --- Sheet1: 電線リスト ---
    ws = wb.active
    ws.title = '電線リスト'
    cols1 = ['ハーネス', '号線', '接続元 機器', '端子', '接続先 機器', '端子', '電線種類', 'サイズsq', '色', '長さmm', '備考']
    ws.append(cols1)
    style_header(ws, 1, len(cols1))
    for h in result['harnesses']:
        for w in h['wires']:
            ws.append([h['name'], w['mark'], w['from_dev'], w['from_term'], w['to_dev'], w['to_term'],
                       w['type'], w['size'], w['color'], w['length'], w['note'] + (' ※要確認' if w['unclear'] else '')])
    widths1 = [22, 8, 14, 8, 14, 8, 10, 8, 8, 9, 18]
    for i, wd in enumerate(widths1, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = wd
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(cols1)):
        for cell in row:
            cell.border = bd

    # --- Sheet2: 切断・端末加工表 ---
    ws2 = wb.create_sheet('切断・端末加工表')
    cols2 = ['ハーネス', '号線', '電線種類', 'サイズsq', '色', '切断長mm', '接続元 端子(圧着)', '接続先 端子(圧着)', 'マーク(号線)', '備考']
    ws2.append(cols2)
    style_header(ws2, 1, len(cols2))
    for h in result['harnesses']:
        for w in h['wires']:
            ws2.append([h['name'], w['mark'], w['type'], w['size'], w['color'], w['cut'],
                        w['from_end'], w['to_end'], w['mark'], w['note']])
    for i, wd in enumerate([22, 8, 10, 8, 8, 9, 16, 16, 10, 16], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = wd
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(cols2)):
        for cell in row:
            cell.border = bd

    # --- Sheet3: 部材集計 ---
    ws3 = wb.create_sheet('部材集計')
    ws3.append(['◆ 電線（種類・サイズ・色 別）'])
    ws3.cell(row=1, column=1).font = Font(bold=True, size=11, color='123B5E')
    ws3.append(['電線種類', 'サイズsq', '色', '本数', '総長m', '総長mm'])
    style_header(ws3, 2, 6)
    r = 3
    for w in result['aggregate']['wires']:
        ws3.append([w['type'], w['size'], w['color'], w['count'], w['length_m'], w['length_mm']])
        r += 1
    ws3.append([])
    r += 1
    ws3.cell(row=r, column=1, value='◆ 端子（圧着）').font = Font(bold=True, size=11, color='123B5E')
    r += 1
    ws3.append(['端子呼び', '本数'])
    style_header(ws3, r, 2)
    for t in result['aggregate']['terminals']:
        ws3.append([t['terminal'], t['count']])
    for i, wd in enumerate([12, 10, 8, 8, 10, 10], 1):
        ws3.column_dimensions[ws3.cell(row=2, column=i).column_letter].width = wd

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
#  ルーティング
# ============================================================
@app.route('/')
@login_required
def index():
    return Response(INDEX_HTML, mimetype='text/html')


@app.route('/api/health')
def health():
    return jsonify(ok=True, key=bool(os.environ.get('ANTHROPIC_API_KEY')), model=MODEL)


@app.route('/api/extract', methods=['POST'])
@login_required
def api_extract():
    files = request.files.getlist('file')
    if not files:
        return jsonify(error='ファイルがありません'), 400
    try:
        cli = client()
    except Exception as e:
        return jsonify(error=str(e)), 400
    all_h = []
    for f in files:
        data = f.read()
        try:
            res = extract_input(cli, f.filename, data)
        except Exception as e:
            return jsonify(error=f'{f.filename}: {e}'), 400
        all_h += res.get('harnesses', [])
    if not all_h:
        return jsonify(error='結線を抽出できませんでした。結線図（端子台結線図・展開接続図）を投入してください。'), 400
    result = build_result(all_h)
    _LAST[_sid()] = result
    return jsonify(result)


@app.route('/api/rebuild', methods=['POST'])
@login_required
def api_rebuild():
    """画面で編集した電線データから、切断長・端子・SVG・集計を再計算する。"""
    data = request.get_json(force=True)
    harnesses = data.get('harnesses', [])
    result = build_result(harnesses)
    _LAST[_sid()] = result
    return jsonify(result)


@app.route('/api/excel/download', methods=['GET'])
@login_required
def api_excel_download():
    result = _LAST.get(_sid())
    if not result:
        return Response('データがありません。先に図面を読み込んでください。', status=404, mimetype='text/plain; charset=utf-8')
    buf = make_excel(result)
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    return send_file(buf, as_attachment=True, download_name=f'ワイヤーハーネス製作_{ts}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


INDEX_HTML = open(os.path.join(HERE, 'index.html'), encoding='utf-8').read() \
    if os.path.exists(os.path.join(HERE, 'index.html')) else '<h1>index.html がありません</h1>'


if __name__ == '__main__':
    port = int(os.environ.get('PORT', '8010'))
    print(f'ワイヤーハーネス製作支援システム起動: http://localhost:{port}  '
          f'(APIキー {"OK" if os.environ.get("ANTHROPIC_API_KEY") else "未設定"})')
    app.run(host='0.0.0.0', port=port, debug=False)
