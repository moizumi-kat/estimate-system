#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
過去図面ライブラリ / 類似図面検索エンジン v1.0

役割:
  抽出済み（または選定済み）の図面データを「過去図面ライブラリ」として蓄積し、
  搭載機器の内容（機器種別・数量・電圧・定格・DB選定コード）を特徴量として
  IDF重み付きコサイン類似度で類似図面を検索する。

設計方針:
  - AIや外部モデルに依存しない。図面から既に抽出済みの構造化データ（盤・機器）だけで
    完結する軽量な内容ベース検索。app.py の抽出/選定パイプラインの出力をそのまま扱う。
  - 抽出段階（items: 品名/数量/電圧）でも選定段階（rows: コード付き）でも保存でき、
    双方が「機器種別トークン(k:)」で相互に照合できる。
  - 保存先は JSON ファイル 1 つ（db.json と同じくファイルDB方式）。
"""
import os, re, json, math, unicodedata, datetime, secrets

HERE = os.path.dirname(os.path.abspath(__file__))
LIBRARY_PATH = os.environ.get('LIBRARY_PATH', os.path.join(HERE, 'library.json'))

# --- 参照データ（db.json / parts_aliases.json）を独立に読み込む(app.pyへの循環参照を避ける) ---
try:
    _DB = json.load(open(os.path.join(HERE, 'db.json'), encoding='utf-8'))
except Exception:
    _DB = []
_byCode = {d.get('code'): d for d in _DB}

try:
    _ALIASES = json.load(open(os.path.join(HERE, 'parts_aliases.json'), encoding='utf-8'))
except Exception:
    _ALIASES = {}


def _norm(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFKC', str(s)).lower()
    return re.sub(r'[ー\-\s　()（）\[\]、,．.]', '', s)


# ===== 機器種別（kind）判定 =====
# parts_aliases.json（正式名→別名・prefixフラグ）から判定ルールを構築。
# 長い別名を優先（"vct" が "ct" に誤マッチしないよう）。
_KIND_RULES = []
for _canon, _info in _ALIASES.items():
    _prefix = str(_info.get('prefix', '')).lower() == 'true'
    for _al in _info.get('aliases', []):
        na = _norm(_al)
        if na:
            _KIND_RULES.append((na, _canon, _prefix))
_KIND_RULES.sort(key=lambda x: len(x[0]), reverse=True)


def detect_kind(name):
    """品名文字列から機器種別（VCB/DS/TR/MCB等の正式名）を推定。無ければ ''。"""
    n = _norm(name)
    if not n:
        return ''
    for al, canon, prefix in _KIND_RULES:
        if prefix:
            if n.startswith(al):
                return canon
        elif al in n:
            return canon
    return ''


def _volt_class(v):
    """電圧表記を代表クラス(HV/400V/200V/100V)に丸める。"""
    n = _norm(v)
    if not n:
        return ''
    if 'kv' in n or re.search(r'6600|3300|22000|高圧', n):
        return 'HV'
    if '400' in n:
        return '400V'
    if '200' in n or '210' in n:
        return '200V'
    if '100' in n or '105' in n:
        return '100V'
    return n[:6]


def _rating_bucket(s):
    """定格(容量/電流)を 'kva:100' / 'a:300' のようなトークンに正規化。"""
    n = _norm(s)
    m = re.search(r'(\d+(?:\.\d+)?)(kva|kvar|kw|af|a|v)?', n)
    if not m:
        return ''
    num, unit = m.group(1), (m.group(2) or 'x')
    return f'{unit}:{num}'


# ===== 金額重み（コード表の金額 furukawa を単価として利用）=====
# furukawa列 = コード表の金額。「金額調整用(マイナス)」行が負値であることから金額列と確定。
_PRICE_FLOOR = 1.0     # 安価な小物・金額欠損でも重みを完全にゼロにしないための下限
_KIND_PRICE_DEFAULT = 20.0


def _amount_of(code):
    """コード → コード表の金額(furukawa)の原値。無効(欠損/非数値)は None。"""
    d = _byCode.get(code)
    if not d:
        return None
    try:
        return float(d.get('furukawa', ''))
    except Exception:
        return None


def _price_of(code):
    """類似度の金額重み。コード表の金額を用い、負(調整項目)・欠損は下限に丸める。"""
    v = _amount_of(code)
    if v is None:
        return _PRICE_FLOOR
    return max(v, _PRICE_FLOOR)


# 機器種別ごとの代表単価（DB内の同種別コードのfurukawa中央値）。
# コード無しの発注資材データでも「高額主機器の一致」を金額重みで評価できる。
def _median(xs):
    xs = sorted(xs)
    n = len(xs)
    if n == 0:
        return None
    return xs[n // 2] if n % 2 else (xs[n // 2 - 1] + xs[n // 2]) / 2.0


_KIND_PRICE = {}
_tmp = {}
for _d in _DB:
    try:
        _v = float(_d.get('furukawa', ''))
    except Exception:
        continue
    if _v <= 0:
        continue
    _k = detect_kind(_d.get('name', ''))
    if _k:
        _tmp.setdefault(_k, []).append(_v)
for _k, _vs in _tmp.items():
    m = _median(_vs)
    if m:
        _KIND_PRICE[_k] = m
del _tmp


def _token_price(t):
    """トークンの金額重み。code>kind>その他(電圧/定格/名称) の順で重い。"""
    if t.startswith('c:'):
        return _price_of(t[2:])
    if t.startswith('k:'):
        return _KIND_PRICE.get(t[2:], _KIND_PRICE_DEFAULT)
    return _PRICE_FLOOR  # v:/r:/n: は金額軸では小さく


# ===== 図面 → 特徴量トークン =====
def _item_name(item):
    return (item.get('display') or item.get('name') or item.get('raw')
            or item.get('official') or '').strip()


def _item_qty(item):
    try:
        q = int(re.sub(r'[^\d]', '', str(item.get('qty', '')))) if item.get('qty') not in (None, '') else 1
    except Exception:
        q = 1
    return q if q > 0 else 1


def _item_tokens(item):
    """機器1点 → {token: weight}。抽出item・選定rowの双方の形に対応。"""
    if item.get('load_detail'):
        return {}
    name = _item_name(item)
    code = item.get('code')
    if not name and not code:
        return {}
    qty = _item_qty(item)
    toks = {}
    kind = detect_kind(name)
    volt = item.get('volt') or ''
    cap = ''

    # 選定済みなら DB 実在コードを強い特徴量に。DBから種別/電圧/定格も補完。
    if code and code in _byCode:
        toks['c:' + str(code)] = toks.get('c:' + str(code), 0) + qty
        d = _byCode[code]
        if not kind:
            kind = detect_kind(d.get('name', ''))
        volt = volt or d.get('volt', '')
        cap = d.get('cap') or d.get('af') or ''

    # 機器種別トークン(k:)。抽出段階・選定段階を橋渡しする主要シグナル。
    if kind:
        toks['k:' + kind] = toks.get('k:' + kind, 0) + qty
    else:
        head = _norm(name)[:12]
        if head:
            toks['n:' + head] = toks.get('n:' + head, 0) + qty

    # 電圧クラス(v:)
    vc = _volt_class(volt) or _volt_class(name)
    if vc:
        toks['v:' + vc] = toks.get('v:' + vc, 0) + qty * 0.5

    # 定格バケット(r:) 変圧器kVA・遮断器A等
    rb = _rating_bucket(cap) or _rating_bucket(item.get('kw', ''))
    if rb:
        toks['r:' + rb] = toks.get('r:' + rb, 0) + qty * 0.5

    return toks


def _iter_items(panels):
    for p in (panels or []):
        its = p.get('rows') if p.get('rows') is not None else p.get('items', [])
        for it in (its or []):
            yield it


def drawing_features(panels):
    """図面(盤リスト) → 特徴量ベクトル {token: weight}。"""
    tokens = {}
    for it in _iter_items(panels):
        for k, w in _item_tokens(it).items():
            tokens[k] = tokens.get(k, 0) + w
    return tokens


def drawing_summary(panels):
    """人が読める要約(機器種別ごとの数量・盤数・機器点数・コード表金額合計)。"""
    kinds = {}
    nitems = 0
    amount = 0.0      # コード付き機器の金額(furukawa×数量)合計
    coded = 0         # コードが確定している機器点数
    for it in _iter_items(panels):
        if it.get('load_detail'):
            continue
        name = _item_name(it)
        code = it.get('code')
        if not name and not code:
            continue
        kind = detect_kind(name) or detect_kind(_byCode.get(code, {}).get('name', ''))
        key = kind or (name[:12] if name else str(code))
        qty = _item_qty(it)
        kinds[key] = kinds.get(key, 0) + qty
        nitems += 1
        if code:
            a = _amount_of(code)
            if a is not None and a > 0:
                amount += a * qty
                coded += 1
    top = sorted(kinds.items(), key=lambda x: -x[1])
    return {
        'nitems': nitems,
        'npanels': len(panels or []),
        'kinds': [{'kind': k, 'qty': v} for k, v in top],
        'amount': round(amount),   # コード表金額の概算合計
        'coded': coded,
    }


# ===== 類似度（IDF重み付きコサイン）=====
def _build_idf(library, extra_tokens=None):
    n = len(library) + (1 if extra_tokens else 0)
    df = {}
    for e in library:
        for k in (e.get('features') or {}):
            df[k] = df.get(k, 0) + 1
    if extra_tokens:
        for k in extra_tokens:
            df[k] = df.get(k, 0) + 1

    def idf(k):
        return math.log((n + 1) / (df.get(k, 0) + 1)) + 1.0
    return idf


def _cosine(a, b, idf, prefixes=None, money=False):
    """トークン種別(prefixes)を絞り、IDF・任意で金額重みを掛けたコサイン類似度。"""
    def sel(v):
        if prefixes is None:
            return v
        return {k: w for k, w in v.items() if k[:2] in prefixes}
    a, b = sel(a), sel(b)
    keys = set(a) | set(b)
    dot = na = nb = 0.0
    for k in keys:
        m = _token_price(k) if money else 1.0
        wa = a.get(k, 0) * idf(k) * m
        wb = b.get(k, 0) * idf(k) * m
        dot += wa * wb
        na += wa * wa
        nb += wb * wb
    if na == 0 or nb == 0:
        return 0.0
    return dot / math.sqrt(na * nb)


def _code_reuse(q, d):
    """クエリの確定コードのうち候補図面にも存在する分の金額比率と実額。
    「この見積の何割(金額ベース)を過去図面から流用できるか」を表す。
    戻り値: (比率0-1, 流用可能金額, クエリ総額) / コード無しなら (None,0,0)。"""
    qc = {k[2:]: w for k, w in q.items() if k.startswith('c:')}
    if not qc:
        return None, 0.0, 0.0
    dc = {k[2:] for k in d if k.startswith('c:')}
    num = sum((_amount_of(c) or 0) * w for c, w in qc.items() if c in dc and (_amount_of(c) or 0) > 0)
    den = sum((_amount_of(c) or 0) * w for c, w in qc.items() if (_amount_of(c) or 0) > 0)
    ratio = (num / den) if den else 0.0
    return ratio, num, den


def search_similar(query_panels, top=10, exclude_id=None):
    """query図面に類似する過去図面を、金額重み＋多軸スコアで降順に返す。"""
    library = load_library()
    q = drawing_features(query_panels)
    if not q:
        return []
    idf = _build_idf(library, q)
    qk = {k[2:] for k in q if k.startswith('k:')}
    results = []
    for e in library:
        if exclude_id and e.get('id') == exclude_id:
            continue
        f = e.get('features') or {}
        # 多軸スコア(0-1)
        major = _cosine(q, f, idf, prefixes={'c:', 'k:'}, money=True)   # 主要機器構成(金額重み)
        cls = _cosine(q, f, idf, prefixes={'v:', 'r:'})                 # 電圧・容量クラス
        reuse, reuse_amt, q_amt = _code_reuse(q, f)                     # コード流用率・流用金額
        # 総合点: 主要機器構成を主軸に、クラス一致・コード流用を加味
        if reuse is None:
            total = 0.72 * major + 0.28 * cls
        else:
            total = 0.55 * major + 0.20 * cls + 0.25 * reuse
        ek = {k[2:] for k in f if k.startswith('k:')}
        results.append({
            'id': e.get('id'),
            'title': e.get('title', ''),
            'note': e.get('note', ''),
            'date': e.get('date', ''),
            'source': e.get('source', ''),
            'spec': e.get('spec', {}),
            'score': round(total * 100, 1),
            'axes': {
                'major': round(major * 100),
                'class': round(cls * 100),
                'reuse': None if reuse is None else round(reuse * 100),
            },
            'reuse_amount': round(reuse_amt),
            'query_amount': round(q_amt),
            'summary': e.get('summary', {}),
            'matched_kinds': sorted(qk & ek),
            'lib_only_kinds': sorted(ek - qk),
            'query_only_kinds': sorted(qk - ek),
        })
    results.sort(key=lambda x: -x['score'])
    return results[:top]


# ===== ライブラリCRUD（JSONファイルDB）=====
def load_library():
    try:
        with open(LIBRARY_PATH, encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []


def save_library(lib):
    tmp = LIBRARY_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(lib, f, ensure_ascii=False, indent=1)
    os.replace(tmp, LIBRARY_PATH)


def add_drawing(title, note, panels, source='extract', created=None, spec=None):
    lib = load_library()
    entry = {
        'id': secrets.token_hex(6),
        'title': (title or '無題').strip() or '無題',
        'note': (note or '').strip(),
        'date': created or datetime.date.today().isoformat(),
        'source': source,
        'spec': spec or {},          # 基本仕様書由来(受電電圧・容量・盤種別・形式 等)
        'panels': panels,
        'features': drawing_features(panels),
        'summary': drawing_summary(panels),
    }
    lib.append(entry)
    save_library(lib)
    return entry


def list_drawings():
    return [{
        'id': e.get('id'),
        'title': e.get('title', ''),
        'note': e.get('note', ''),
        'date': e.get('date', ''),
        'source': e.get('source', ''),
        'spec': e.get('spec', {}),
        'summary': e.get('summary', {}),
    } for e in load_library()]


# ===== 発注資材データ(Excel/CSV)の取込 =====
# ヘッダ名から列を自動判定。品名/型式・数量・(あれば)積算コードを拾う。
_COL_HINTS = {
    'name': ['品名仕様', '品名', '名称', '機器名', '器具名', '形式', '型式', '規格', '仕様',
             'item', 'name', 'material', '資材'],
    'qty':  ['数量', '員数', '数', 'qty', 'quantity', '台数', '個数'],
    'code': ['積算コード', 'コード', 'code', '品番', '部品コード'],
}


def _norm_header(s):
    return re.sub(r'[\s　:：]', '', str(s or '')).lower()


def _map_columns(headers):
    """ヘッダ行 → {役割: 列index}。見つからない役割は欠落。"""
    nh = [_norm_header(h) for h in headers]
    colmap = {}
    for role, hints in _COL_HINTS.items():
        for i, h in enumerate(nh):
            if not h:
                continue
            if any(hint in h for hint in hints):
                colmap[role] = i
                break
    return colmap


def rows_to_panels(rows, panel_name='発注資材', colmap=None):
    """表形式データ(行の配列, 先頭行ヘッダ) → panels 構造。"""
    if not rows:
        return [], {}
    headers = rows[0]
    colmap = colmap or _map_columns(headers)
    ni = colmap.get('name')
    qi = colmap.get('qty')
    ci = colmap.get('code')
    if ni is None:
        # 品名列が判別できない場合は先頭の非数値列を品名とみなす
        ni = 0
    items = []
    for r in rows[1:]:
        if not r:
            continue
        def cell(idx):
            return str(r[idx]).strip() if idx is not None and idx < len(r) and r[idx] is not None else ''
        name = cell(ni)
        if not name:
            continue
        qty = cell(qi) or '1'
        code = cell(ci)
        it = {'name': name, 'qty': qty}
        if code and code in _byCode:
            it['code'] = code
        items.append(it)
    panels = [{'panel': panel_name, 'items': items}]
    return panels, colmap


def parse_csv(text):
    import csv, io as _io
    # 区切り自動判定(カンマ/タブ)
    sample = text[:2000]
    delim = '\t' if sample.count('\t') > sample.count(',') else ','
    rdr = csv.reader(_io.StringIO(text), delimiter=delim)
    return [list(r) for r in rdr]


def parse_xlsx(data_bytes):
    from openpyxl import load_workbook
    import io as _io
    wb = load_workbook(_io.BytesIO(data_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(['' if c is None else c for c in row])
    return rows


def import_material_file(filename, data_bytes, title='', note='', spec=None, panel_name='発注資材'):
    """Excel/CSVの発注資材データを取り込み、ライブラリ登録して返す。"""
    low = (filename or '').lower()
    if low.endswith(('.xlsx', '.xlsm')):
        rows = parse_xlsx(data_bytes)
    elif low.endswith(('.csv', '.txt', '.tsv')):
        try:
            text = data_bytes.decode('utf-8-sig')
        except Exception:
            text = data_bytes.decode('cp932', errors='replace')
        rows = parse_csv(text)
    else:
        raise ValueError('対応形式は .xlsx / .csv です')
    panels, colmap = rows_to_panels(rows, panel_name=panel_name)
    if not panels or not panels[0]['items']:
        raise ValueError('資材行を読み取れませんでした（品名列を確認してください）')
    entry = add_drawing(title or filename, note, panels, source='material', spec=spec)
    entry = dict(entry)
    entry['colmap'] = {k: int(v) for k, v in colmap.items()}
    entry['nitems'] = len(panels[0]['items'])
    return entry


def get_drawing(did):
    for e in load_library():
        if e.get('id') == did:
            return e
    return None


def delete_drawing(did):
    lib = load_library()
    new = [e for e in lib if e.get('id') != did]
    if len(new) == len(lib):
        return False
    save_library(new)
    return True
